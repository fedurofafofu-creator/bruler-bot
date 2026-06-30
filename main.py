import asyncio
import logging
import json
import uuid
import re
import io
import time
import pytz
import gspread
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo
from google.oauth2.service_account import Credentials

from telegram import Update, Bot, InlineKeyboardButton, InlineKeyboardMarkup, InputFile
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    ConversationHandler, ContextTypes, filters, CallbackQueryHandler
)
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
import os

logging.basicConfig(format="%(asctime)s [%(levelname)s] %(name)s: %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

# ── CONFIG ────────────────────────────────────────────────────────────────────
BOT_TOKEN      = os.getenv("BOT_TOKEN", "8767537310:AAHK1-RmvgH6yF6ZShQFq3A1DeLU0uFsAMs")
SPREADSHEET_ID = os.getenv("SPREADSHEET_ID", "1LCmoydfS73DKwjQcwMSnpOZxRHiKMAMjOZRRk0AwYnE")
ADMIN_CHAT_ID  = int(os.getenv("ADMIN_CHAT_ID", "-5486975908"))
TZ             = ZoneInfo("Europe/Moscow")
APZ            = pytz.timezone("Europe/Moscow")
PROCESS_STARTED_AT = datetime.now(TZ)  # для /ping — видно когда контейнер последний раз перезапускался

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

DEPT_LIST = ["Маркетинг", "Продажи", "Бухгалтерия", "Производство и дизайн", "Руководство"]

# Частые русские имена (для определения порядка Имя/Фамилия при нормализации)
COMMON_FIRST_NAMES = {
    "александр","алексей","анатолий","андрей","антон","аркадий","артём","артем",
    "борис","вадим","валентин","валерий","василий","виктор","виталий","владимир",
    "владислав","геннадий","георгий","григорий","даниил","денис","дмитрий","евгений",
    "егор","иван","игорь","илья","кирилл","константин","леонид","максим","максимилиан",
    "михаил","никита","николай","олег","павел","петр","пётр","роман","руслан",
    "семен","семён","сергей","степан","тимофей","тимур","федор","фёдор","юрий","ярослав",
    "александра","алина","алла","анастасия","анжела","анна","антонина","валентина",
    "валерия","вера","вероника","виктория","галина","дарья","диана","евгения",
    "екатерина","елена","елизавета","зоя","ирина","карина","кристина","ксения",
    "лариса","лидия","любовь","людмила","маргарита","марина","мария","надежда",
    "наталья","оксана","ольга","полина","раиса","светлана","софия","софья",
    "тамара","татьяна","юлия","яна",
}

def normalize_full_name(raw: str) -> str:
    """Приводит ввод к формату 'Имя Фамилия'. Распознаёт частые имена,
    переставляет слова если первое слово похоже на фамилию."""
    parts = raw.strip().split()
    if len(parts) < 2:
        return raw.strip().title()
    parts = [p.capitalize() for p in parts[:2]]
    first_lower = parts[0].lower()
    second_lower = parts[1].lower()
    if first_lower in COMMON_FIRST_NAMES:
        return f"{parts[0]} {parts[1]}"
    if second_lower in COMMON_FIRST_NAMES:
        return f"{parts[1]} {parts[0]}"
    # не распознали — оставляем как ввёл пользователь
    return f"{parts[0]} {parts[1]}"

# Стартовые значения — далее отдел хранится в Google Sheets (колонка department)
# и может быть переопределён через кнопки при регистрации / командой /setdept
DEPARTMENTS_SEED = {
    "7070230704": "Руководство",
    "7198542902": "Руководство",
    "8151347813": "Продажи",
    "195676845":  "Маркетинг",
    "8069881891": "Бухгалтерия",
    "458764300":  "Производство и дизайн",
    "860192861":  "Производство и дизайн",
    "89555212":   "Производство и дизайн",
    "549232571":  "Маркетинг",
}

def get_dept(tg_id) -> str:
    """Отдел сотрудника: сначала смотрим в Sheets, иначе — seed-словарь, иначе 'Без отдела'."""
    r = emp_by_id(tg_id)
    if r and r.get("department"):
        return r["department"]
    return DEPARTMENTS_SEED.get(str(tg_id), "Без отдела")

# ── GOOGLE SHEETS ─────────────────────────────────────────────────────────────
_gc = None
_ss = None
_ws_cache = {}        # name -> Worksheet object, кэшируется навсегда (объект не меняется)
_data_cache = {}       # name -> (timestamp, rows), короткий TTL чтобы не дёргать API на каждый вызов
_headers_checked = set()  # name -> уже проверяли заголовок в этом запуске процесса
DATA_CACHE_TTL = 8     # секунд — этого достаточно чтобы пережить серию из 5-6 команд подряд

def gc():
    global _gc
    if _gc is None:
        raw = os.getenv("GOOGLE_CREDENTIALS_JSON", "")
        if raw:
            info = json.loads(raw)
        else:
            with open("credentials.json") as f:
                info = json.load(f)
        creds = Credentials.from_service_account_info(info, scopes=SCOPES)
        _gc = gspread.authorize(creds)
    return _gc

def ss():
    global _ss
    if _ss is None:
        _ss = gc().open_by_key(SPREADSHEET_ID)
    return _ss

def sheet(name):
    """Кэширует объект Worksheet — избегает повторного fetch_sheet_metadata
    при каждом вызове (это отдельный API-запрос, который быстро съедает квоту)."""
    if name in _ws_cache:
        return _ws_cache[name]
    try:
        ws = ss().worksheet(name)
    except gspread.WorksheetNotFound:
        ws = ss().add_worksheet(title=name, rows=2000, cols=20)
    _ws_cache[name] = ws
    return ws

def invalidate_cache(name=None):
    """Сбрасывает кэш данных после любой записи, чтобы следующее чтение было свежим."""
    if name:
        _data_cache.pop(name, None)
    else:
        _data_cache.clear()

def ensure_headers(ws, headers, force=False):
    """
    Гарантирует корректный, не дублирующийся заголовок в первой строке.
    Проверяется только один раз за время жизни процесса на каждый лист
    (через _headers_checked), а не на каждый вызов emp_sheet()/tasks_sheet().
    Раньше ws.row_values(1) был отдельным API-запросом при КАЖДОМ обращении
    к листу — при 100+ последовательных вызовах get_dept() в одном цикле это
    моментально съедало квоту Google Sheets, даже с кэшем данных на месте.
    force=True используется явно в /fixsheets для принудительной проверки.
    """
    cache_key = ws.title
    if not force and cache_key in _headers_checked:
        return
    vals = ws.row_values(1)
    if vals != headers:
        # перезаписываем строку заголовка целиком (без сдвига остальных строк)
        ws.update('A1', [headers])
    _headers_checked.add(cache_key)

def safe_records(ws, headers):
    """
    Читает данные листа вручную по позиции колонки, без gspread.get_all_records().
    Устойчиво к повреждённому заголовку и расхождению числа колонок после миграции схемы.
    Кэширует результат на DATA_CACHE_TTL секунд — несколько команд подряд от одного
    или разных пользователей переиспользуют один и тот же снимок данных, что резко
    снижает число запросов к Google Sheets API и защищает от 429 Quota exceeded.
    При срабатывании квоты делает 2 повторные попытки с задержкой; если квота всё
    ещё превышена — отдаёт последний известный кэш (даже просроченный), а не падает.
    """
    cache_key = ws.title
    cached = _data_cache.get(cache_key)
    if cached and (datetime.now(TZ) - cached[0]).total_seconds() < DATA_CACHE_TTL:
        return cached[1]

    last_error = None
    for attempt in range(3):
        try:
            ensure_headers(ws, headers)
            all_values = ws.get_all_values()
            n = len(headers)
            result = []
            if len(all_values) >= 2:
                for row in all_values[1:]:
                    if len(row) < n:
                        row = row + [""] * (n - len(row))
                    elif len(row) > n:
                        row = row[:n]
                    result.append({headers[i]: row[i] for i in range(n)})
            _data_cache[cache_key] = (datetime.now(TZ), result)
            return result
        except gspread.exceptions.APIError as e:
            last_error = e
            if "429" in str(e) or "Quota exceeded" in str(e):
                # Блокирующий sleep здесь осознанно: это редкий аварийный путь
                # (квота превышена), а не штатный режим — благодаря кэшу выше
                # обычные обращения сюда вообще не доходят.
                time.sleep(2 * (attempt + 1))
                continue
            raise

    # квота не отпустила за 3 попытки — отдаём устаревший кэш, если он есть,
    # лучше показать чуть старые данные чем уронить команду полностью
    if cached:
        logger.warning(f"Sheets quota exceeded for {cache_key}, serving stale cache")
        return cached[1]
    logger.error(f"Sheets quota exceeded for {cache_key}, no cache available")
    raise last_error

# ── EMPLOYEES ─────────────────────────────────────────────────────────────────
EMP_H = ["tg_id","username","full_name","role","registered_at","department","founder_digest_pref","shift_schedule"]

def emp_sheet():
    ws = sheet("employees"); ensure_headers(ws, EMP_H); return ws

def emp_all():
    return safe_records(emp_sheet(), EMP_H)

def emp_employees():
    """Все, кто проходит ежедневный цикл план→EOD: employee, dept_head и admin —
    руководители тоже работают и должны отчитываться о своих задачах."""
    return [r for r in emp_all() if r["role"] in ("employee", "dept_head", "admin")]

def emp_admins():
    return [r for r in emp_all() if r["role"] == "admin"]

def emp_dept_heads():
    return [r for r in emp_all() if r["role"] == "dept_head"]

def emp_by_id(tg_id):
    for r in emp_all():
        if str(r["tg_id"]) == str(tg_id):
            return r
    return None

def emp_registered(tg_id):
    return emp_by_id(tg_id) is not None

def emp_is_admin(tg_id):
    r = emp_by_id(tg_id)
    return r is not None and r["role"] == "admin"

def emp_is_dept_head(tg_id):
    r = emp_by_id(tg_id)
    return r is not None and r["role"] == "dept_head"

def emp_is_founder(tg_id):
    """Учредитель — read-only роль, видит всё по компании, не управляет ничем,
    не проходит обучение, не пишет планы/отчёты, не получает напоминания
    о задачах. Только просмотр результатов и сводок."""
    r = emp_by_id(tg_id)
    return r is not None and r["role"] == "founder"

def emp_has_management_access(tg_id):
    """Полный доступ к панели на чтение: admin (полное управление),
    dept_head (управление своим отделом) или founder (только просмотр,
    без управляющих кнопок — see is_strict_admin_check и build_founder_keyboard)."""
    r = emp_by_id(tg_id)
    return r is not None and r["role"] in ("admin", "dept_head", "founder")

def emp_managed_dept(tg_id) -> str:
    """Отдел, который видит dept_head. Для admin и founder — пусто (видят всё)."""
    r = emp_by_id(tg_id)
    if not r:
        return ""
    if r["role"] in ("admin", "founder"):
        return ""
    return get_dept(tg_id)

def emp_register(tg_id, username, full_name, role="employee", department=""):
    if emp_registered(tg_id):
        return False
    emp_sheet().append_row([tg_id, username or "", full_name, role,
                             datetime.now(TZ).strftime("%Y-%m-%d %H:%M"), department, "", ""])
    invalidate_cache("employees")
    return True

def emp_set_admin(tg_id):
    ws = emp_sheet()
    for i, r in enumerate(safe_records(ws, EMP_H), start=2):
        if str(r["tg_id"]) == str(tg_id):
            ws.update_cell(i, EMP_H.index("role") + 1, "admin")
            invalidate_cache("employees")
            return True
    return False

def emp_set_role(tg_id, role: str):
    ws = emp_sheet()
    for i, r in enumerate(safe_records(ws, EMP_H), start=2):
        if str(r["tg_id"]) == str(tg_id):
            ws.update_cell(i, EMP_H.index("role") + 1, role)
            invalidate_cache("employees")
            return True
    return False

def emp_delete(tg_id) -> bool:
    """
    Удаляет сотрудника из листа employees. Его прошлые задачи, планы и
    отчёты в других листах НЕ удаляются — остаются для истории/отчётности,
    но сам человек больше не появится в emp_employees(), не получит
    автоматических напоминаний и не сможет пользоваться ботом без повторной
    регистрации через /start.
    """
    ws = emp_sheet()
    records = safe_records(ws, EMP_H)
    for i, r in enumerate(records, start=2):
        if str(r["tg_id"]) == str(tg_id):
            ws.delete_rows(i)
            invalidate_cache("employees")
            return True
    return False

def emp_set_founder_digest_pref(tg_id, pref: str):
    """pref: 'manual' (только заходит сам), 'digest' (только авторассылка),
    'both' (и то, и другое)."""
    ws = emp_sheet()
    for i, r in enumerate(safe_records(ws, EMP_H), start=2):
        if str(r["tg_id"]) == str(tg_id):
            ws.update_cell(i, EMP_H.index("founder_digest_pref") + 1, pref)
            invalidate_cache("employees")
            return True
    return False

# Графики работы: "" (пусто) или "standard" = 10:00–19:00, "shifted" = 11:00–22:00.
# Назначается вручную admin через /setshift — для должностей с другим графиком
# (например продавцы в магазине/шоуруме).
SHIFT_SCHEDULES = {
    "standard": {"start_hour": 10, "start_minute": 0, "end_hour": 19, "end_minute": 0},
    "shifted":  {"start_hour": 11, "start_minute": 0, "end_hour": 22, "end_minute": 0},
}

def emp_shift(tg_id) -> dict:
    """Возвращает словарь {start_hour, start_minute, end_hour, end_minute}
    для сотрудника — standard по умолчанию, shifted если назначено явно."""
    r = emp_by_id(tg_id)
    code = (r.get("shift_schedule") if r else "") or "standard"
    return SHIFT_SCHEDULES.get(code, SHIFT_SCHEDULES["standard"])

def emp_set_shift_schedule(tg_id, code: str):
    """code: 'standard' или 'shifted'."""
    ws = emp_sheet()
    for i, r in enumerate(safe_records(ws, EMP_H), start=2):
        if str(r["tg_id"]) == str(tg_id):
            ws.update_cell(i, EMP_H.index("shift_schedule") + 1, code)
            invalidate_cache("employees")
            return True
    return False

def emp_set_department(tg_id, department: str):
    ws = emp_sheet()
    for i, r in enumerate(safe_records(ws, EMP_H), start=2):
        if str(r["tg_id"]) == str(tg_id):
            ws.update_cell(i, EMP_H.index("department") + 1, department)
            invalidate_cache("employees")
            return True
    return False

def emp_set_full_name(tg_id, full_name: str):
    ws = emp_sheet()
    for i, r in enumerate(safe_records(ws, EMP_H), start=2):
        if str(r["tg_id"]) == str(tg_id):
            ws.update_cell(i, EMP_H.index("full_name") + 1, full_name)
            invalidate_cache("employees")
            return True
    return False

# ── PLANS & REPORTS ───────────────────────────────────────────────────────────
PH = ["date","tg_id","full_name","plan_text","submitted_at"]
RH = ["date","tg_id","full_name","report_text","submitted_at"]

def plans_sheet():
    ws = sheet("plans"); ensure_headers(ws, PH); return ws
def reports_sheet():
    ws = sheet("reports"); ensure_headers(ws, RH); return ws

def today_date():
    """Текущая дата по московскому времени, как объект date (не datetime).
    Использовать вместо date.today() везде в проекте — date.today() наивный
    и берёт дату по таймзоне сервера (UTC на Railway), что давало расхождение
    с реальным московским временем до 3 часов на границе суток."""
    return datetime.now(TZ).date()

def today_str():
    return datetime.now(TZ).strftime("%Y-%m-%d")

_DEADLINE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")

def is_overdue(deadline: str, today: str = None) -> bool:
    """
    Безопасная проверка 'дедлайн в прошлом', устойчивая к неправильному
    формату даты. Прямое сравнение строк (deadline < today) ломается, если
    где-то в таблице дедлайн оказался не в формате ГГГГ-ММ-ДД (например
    ДД.ММ.ГГГГ после ручной правки в Google Sheets) — тогда дальняя дата в
    неверном формате может "проиграть" сегодняшней по первому символу и
    ошибочно посчитаться просроченной. Если формат не распознан — задача
    НЕ считается просроченной (безопасный отказ, не ложная просрочка).
    """
    if not deadline or not _DEADLINE_RE.match(deadline):
        return False
    if today is None:
        today = today_str()
    return deadline < today

def save_plan(tg_id, name, text):
    plans_sheet().append_row([today_str(), tg_id, name, text,
                               datetime.now(TZ).strftime("%Y-%m-%d %H:%M")])
    invalidate_cache("plans")

def save_report(tg_id, name, text):
    reports_sheet().append_row([today_str(), tg_id, name, text,
                                 datetime.now(TZ).strftime("%Y-%m-%d %H:%M")])
    invalidate_cache("reports")

def plans_today():
    return [r for r in safe_records(plans_sheet(), PH) if r["date"] == today_str()]

def reports_today():
    return [r for r in safe_records(reports_sheet(), RH) if r["date"] == today_str()]

def has_plan_today(tg_id):
    return any(str(r["tg_id"]) == str(tg_id) for r in plans_today())

def has_report_today(tg_id):
    return any(str(r["tg_id"]) == str(tg_id) for r in reports_today())

# ── РАБОЧИЙ ДЕНЬ: СТАРТ/КОНЕЦ ────────────────────────────────────────────────
WH = ["date", "tg_id", "full_name", "start_at", "end_at"]

def workday_sheet():
    ws = sheet("workday"); ensure_headers(ws, WH); return ws

def workday_today() -> list:
    return [r for r in safe_records(workday_sheet(), WH) if r["date"] == today_str()]

def workday_row_for(tg_id):
    for r in workday_today():
        if str(r["tg_id"]) == str(tg_id):
            return r
    return None

def workday_started(tg_id) -> bool:
    r = workday_row_for(tg_id)
    return bool(r and r.get("start_at"))

def workday_ended(tg_id) -> bool:
    r = workday_row_for(tg_id)
    return bool(r and r.get("end_at"))

def workday_mark_start(tg_id, full_name):
    """Отмечает начало рабочего дня. Если строка на сегодня уже есть — не
    дублирует и не перезаписывает существующий start_at (защита от повторного
    нажатия кнопки)."""
    if workday_started(tg_id):
        return False
    ws = workday_sheet()
    now = datetime.now(TZ).strftime("%Y-%m-%d %H:%M")
    ws.append_row([today_str(), tg_id, full_name, now, ""])
    invalidate_cache("workday")
    return True

def workday_mark_end(tg_id):
    """Отмечает конец рабочего дня в уже существующей строке на сегодня."""
    ws = workday_sheet()
    records = safe_records(ws, WH)
    for i, r in enumerate(records, start=2):
        if str(r["tg_id"]) == str(tg_id) and r["date"] == today_str():
            if r.get("end_at"):
                return False  # уже отмечен
            ws.update_cell(i, WH.index("end_at") + 1, datetime.now(TZ).strftime("%Y-%m-%d %H:%M"))
            invalidate_cache("workday")
            return True
    return False

def records_for_week(ws_name, headers):
    ws = sheet(ws_name)
    iso = datetime.now(TZ).strftime("%Y-W%W")
    result = []
    for r in safe_records(ws, headers):
        try:
            d = datetime.strptime(r["date"], "%Y-%m-%d")
            if d.strftime("%Y-W%W") == iso:
                result.append(r)
        except Exception:
            pass
    return result

def records_for_month(ws_name, headers):
    ws = sheet(ws_name)
    month = datetime.now(TZ).strftime("%Y-%m")
    return [r for r in safe_records(ws, headers) if r.get("date","").startswith(month)]

def reports_for_period(date_from, date_to):
    ws = sheet("reports")
    return [r for r in safe_records(ws, RH)
            if r.get("date","") >= date_from and r.get("date","") <= date_to]

def plans_for_period(date_from, date_to):
    ws = sheet("plans")
    return [p for p in safe_records(ws, PH)
            if p.get("date","") >= date_from and p.get("date","") <= date_to]

# ── TASKS ─────────────────────────────────────────────────────────────────────
# status: open | in_progress | done | paused | overdue
TH = ["task_id","created_by_id","created_by_name","assigned_to_id",
      "assigned_to_name","title","deadline","status","created_at",
      "done_at","comment","result_link","source","channel","is_training"]

CHANNEL_LIST = ["Сайт", "Маркетплейсы", "Комиссионеры", "Опт", "Розница", "Bruler Studio", "Не применимо"]

# ── РЕЕСТР КОМАНД ─────────────────────────────────────────────────────────────
# Единый источник правды для /start, /help и кнопки "Обучение".
# role: "employee" — видно всем, "dept_head" — видно dept_head и admin,
# "admin" — видно только admin. Группы задают порядок и заголовки секций.
COMMAND_REGISTRY = [
    {
        "group": "📋 Ежедневная работа",
        "role": "employee",
        "items": [
            ("/startday", "начать рабочий день (запросит план)"),
            ("/plan", "план на день"),
            ("/mytasks", "мои активные задачи"),
            ("/changestatus", "сменить статус задачи"),
            ("/done", "отметить выполненной"),
            ("/endday", "актуализировать статусы и закрыть день"),
            ("/eod", "закрыть день вручную"),
        ],
    },
    {
        "group": "📌 Управление задачами",
        "role": "employee",
        "items": [
            ("/task", "поставить задачу"),
            ("/tag", "тег канала для задачи"),
            ("/edit", "изменить задачу"),
            ("/status", "карточка задачи по ID"),
        ],
    },
    {
        "group": "🎛 Панель руководителя",
        "role": "dept_head",
        "items": [
            ("/menu", "все функции отдела/команды кнопками"),
            ("/team", "кто сдал план сегодня"),
            ("/workday", "кто во сколько начал/закончил день"),
            ("/tasks_all", "все задачи текстом"),
            ("/checkstatuses", "запросить статусы прямо сейчас"),
            ("/learningreport", "прогресс обучения команды"),
        ],
    },
    {
        "group": "👑 Администрирование",
        "role": "admin",
        "items": [
            ("/setadmin", "выдать права администратора"),
            ("/setdepthead", "назначить руководителя отдела"),
            ("/setfounder", "назначить учредителя"),
            ("/setdept", "сменить отдел сотрудника"),
            ("/setshift", "назначить график (стандартный/сдвинутый)"),
            ("/recovertasks", "восстановить задачи из планов"),
            ("/deleteuser", "удалить сотрудника из команды"),
            ("/testdigest", "диагностика отправки сводки в группу"),
            ("/fixname", "исправить имя сотрудника"),
            ("/fixallnames", "нормализовать все имена"),
            ("/fixsheets", "аварийный ремонт таблиц"),
        ],
    },
]

def commands_for_role(role: str) -> list:
    """Возвращает список групп команд, видимых для роли employee/dept_head/admin."""
    visible_roles = {
        "employee":  {"employee"},
        "dept_head": {"employee", "dept_head"},
        "admin":     {"employee", "dept_head", "admin"},
    }.get(role, {"employee"})
    return [g for g in COMMAND_REGISTRY if g["role"] in visible_roles]

def format_commands_text(role: str) -> str:
    """Генерирует текст списка команд для /start — единственный источник правды,
    устраняет рассинхрон между тем, что написано в приветствии, и тем, что бот
    реально умеет (было 6 команд в тексте против 19+ зарегистрированных)."""
    lines = []
    for group in commands_for_role(role):
        lines.append(f"\n{group['group']}")
        for cmd, desc in group["items"]:
            lines.append(f"{cmd} — {desc}")
    return "\n".join(lines).strip()

# ── СЦЕНАРИИ ОБУЧЕНИЯ ─────────────────────────────────────────────────────────
# Короткие карточки "как сделать X", доступные через кнопку 📚 Обучение.
# Заменяют ручное объяснение новым сотрудникам — обучение через действие,
# не через текст: каждая карточка кончается кнопкой "Попробовать сейчас".
LEARN_SCENARIOS = [
    {
        "id": "startday", "role": "employee", "label": "☀️ Как начать рабочий день", "version": 1,
        "blurb": "Отмечаешь начало дня — бот сразу спрашивает план.",
        "text": (
            "☀️ <b>Как начать рабочий день</b>\n\n"
            "Отметь начало рабочего дня командой /startday в любой момент,"
            " когда реально приступаешь к работе.\n\n"
            "Если не отметишь сам — бот напомнит в 10:00 (или в 11:00, если "
            "у тебя сдвинутый график), и повторит напоминание ещё раз через "
            "30 минут.\n\n"
            "<i>Подсказка: сразу после старта бот попросит план на день —"
            " это один шаг, не два отдельных.</i>"
        ),
        "try_cmd": "/startday",
    },
    {
        "id": "plan", "role": "employee", "label": "📝 Как написать план", "version": 2,
        "blurb": "Пишешь список дел — каждая строка становится задачей.",
        "text": (
            "📝 <b>Как написать план</b>\n\n"
            "План запрашивается сразу после того, как ты отметил(а) начало "
            "рабочего дня через /startday — не в фиксированное время для всех.\n"
            "Просто напиши список дел — каждая строка станет отдельной задачей.\n\n"
            "Если срок не сегодня — добавь дату в конце строки: [25.06.2026]\n\n"
            "<i>Пример:</i>\n"
            "Согласовать смету\n"
            "Позвонить поставщику [25.06.2026]\n\n"
            "<i>Подсказка: если есть просроченные задачи из прошлых дней,"
            " бот сразу предложит их актуализировать.</i>"
        ),
        "try_cmd": "/plan",
    },
    {
        "id": "eod", "role": "employee", "label": "✅ Как закрыть день", "version": 2,
        "blurb": "Проходишь по задачам, отмечаешь статус каждой, закрываешь день.",
        "text": (
            "✅ <b>Как закрыть день</b>\n\n"
            "Отметь конец рабочего дня командой /endday — это запустит опрос "
            "статуса по каждой задаче из плана.\n"
            "Для каждой: выбери статус кнопкой → выбери канал кнопкой → напиши комментарий.\n\n"
            "Если не отметишь сам — бот спросит в 19:00 (или в 22:00 для "
            "сдвинутого графика), с повтором через 30 минут.\n\n"
            "<i>Подсказка: день засчитывается закрытым только после того,"
            " как все статусы актуализированы — не раньше.</i>"
        ),
        "try_cmd": "/endday",
    },
    {
        "id": "task", "role": "employee", "label": "📌 Как поставить задачу", "version": 1,
        "blurb": "Поручаешь дело коллеге — он получает уведомление сразу.",
        "text": (
            "📌 <b>Как поставить задачу</b>\n\n"
            "Напиши /task → выбери отдел кнопкой → выбери сотрудника кнопкой →\n"
            "напиши название и срок текстом (например: Сделать макет | 25.06.2026).\n\n"
            "Исполнитель сразу получит уведомление."
        ),
        "try_cmd": "/task",
    },
    {
        "id": "tag", "role": "employee", "label": "🏷 Как тегировать канал", "version": 2,
        "blurb": "Привязываешь задачу к каналу продаж — или ставишь «Не применимо».",
        "text": (
            "🏷 <b>Как тегировать канал</b>\n\n"
            "Напиши /tag → выбери задачу из списка → выбери канал продаж кнопкой.\n\n"
            "Каналы: Сайт, Маркетплейсы, Комиссионеры, Опт, Розница, Bruler Studio.\n\n"
            "<i>Правило выбора:</i>\n"
            "— Если у задачи есть чёткая принадлежность к конкретному каналу "
            "продаж (например «Согласовать КП для Wildberries» → Маркетплейсы) "
            "— выбираешь этот канал.\n"
            "— Если задача не связана с конкретным каналом (внутренняя "
            "встреча, административная работа, общая задача не про продажи) "
            "— выбираешь <b>«Не применимо»</b>.\n\n"
            "<i>Подсказка: не угадывай и не оставляй пустым — если сомневаешься,"
            " «Не применимо» лучше, чем случайный канал.</i>"
        ),
        "try_cmd": "/tag",
    },
    {
        "id": "changestatus", "role": "employee", "label": "🔄 Как сменить статус задачи", "version": 1,
        "blurb": "Меняешь статус одной задачи прямо сейчас, без полного EOD.",
        "text": (
            "🔄 <b>Как сменить статус задачи</b>\n\n"
            "Напиши /changestatus → выбери задачу → выбери новый статус кнопкой.\n\n"
            "Работает в любой момент дня, не дожидаясь вечернего закрытия дня."
        ),
        "try_cmd": "/changestatus",
    },
    {
        "id": "menu", "role": "dept_head", "label": "🎛 Как пользоваться панелью", "version": 1,
        "blurb": "Смотришь задачи, просрочки и отчётность своего отдела.",
        "text": (
            "🎛 <b>Как пользоваться панелью /menu</b>\n\n"
            "Открывает кнопки: задачи отдела, просроченные, отчётность по сотрудникам,\n"
            "за неделю/месяц. Всё отфильтровано по твоему отделу автоматически."
        ),
        "try_cmd": "/menu",
    },
    {
        "id": "remind", "role": "dept_head", "label": "🔔 Как напомнить о задаче", "version": 1,
        "blurb": "Подталкиваешь сотрудника по конкретной задаче, выбрав её из списка.",
        "text": (
            "🔔 <b>Как напомнить сотруднику о задаче</b>\n\n"
            "В /menu нажми «🔔 Напомнить о задаче» → выбери задачу из списка отдела.\n\n"
            "Бот сразу отправит исполнителю напоминание с кнопкой /done."
        ),
        "try_cmd": "/menu",
    },
    {
        "id": "export", "role": "admin", "label": "📥 Как выгрузить отчёт", "version": 1,
        "blurb": "Получаешь Excel-файл со всеми задачами компании за период.",
        "text": (
            "📥 <b>Как выгрузить отчёт в Excel</b>\n\n"
            "В /menu нажми «📥 Экспорт отдела» → выбери период (неделя/месяц).\n\n"
            "Бот сформирует .xlsx со всеми задачами, статусами и тегами каналов."
        ),
        "try_cmd": "/menu",
    },
]

ROLE_LEARNING_SEQUENCE = {
    "employee":  ["startday", "plan", "eod", "task", "tag", "changestatus"],
    "dept_head": ["startday", "plan", "eod", "task", "tag", "changestatus", "menu", "remind"],
    "admin":     ["startday", "plan", "eod", "task", "tag", "changestatus", "menu", "remind", "export"],
}

def learn_scenarios_for_role(role: str) -> list:
    """Возвращает сценарии в строго заданном порядке для роли (своя полная
    последовательность для каждой роли, не общий список с надстройками)."""
    sequence = ROLE_LEARNING_SEQUENCE.get(role, ROLE_LEARNING_SEQUENCE["employee"])
    by_id = {s["id"]: s for s in LEARN_SCENARIOS}
    return [by_id[sid] for sid in sequence if sid in by_id]

# ── ЛИСТ ОБУЧЕНИЯ ─────────────────────────────────────────────────────────────
# Широкая таблица: одна строка на сотрудника, один столбец на каждый сценарий
# обучения + служебные колонки прогресса. Ячейка сценария содержит
# "✅ vN, ДД.ММ" если пройдено, иначе пусто.
LH_FIXED = ["tg_id", "full_name", "role", "department", "started_at", "completed_at", "last_reminder_at"]

def learning_columns() -> list:
    """Фиксированные колонки + одна колонка на каждый сценарий из LEARN_SCENARIOS."""
    return LH_FIXED + [s["id"] for s in LEARN_SCENARIOS]

def learning_sheet():
    ws = sheet("learning_progress"); ensure_headers(ws, learning_columns()); return ws

def learning_all() -> list:
    return safe_records(learning_sheet(), learning_columns())

def learning_format_cell(version: int, completed_at: str) -> str:
    date_part = completed_at.split(" ")[0] if completed_at else ""
    if len(date_part) == 10:
        y, m, d = date_part.split("-")
        date_part = f"{d}.{m}"
    return f"✅ v{version}, {date_part}".strip(", ")

def learning_find_or_create_row(tg_id, full_name, role, department) -> int:
    """Находит строку сотрудника в таблице обучения или создаёт новую,
    фиксируя started_at в момент первого обращения к обучению."""
    ws = learning_sheet()
    columns = learning_columns()
    records = safe_records(ws, columns)
    for i, r in enumerate(records, start=2):
        if str(r["tg_id"]) == str(tg_id):
            return i
    now = datetime.now(TZ).strftime("%Y-%m-%d %H:%M")
    blank_row = [tg_id, full_name, role, department, now, "", ""] + [""] * len(LEARN_SCENARIOS)
    ws.append_row(blank_row)
    invalidate_cache("learning_progress")
    return len(records) + 2

def learning_get_row_dict(tg_id):
    """Возвращает сырую строку прогресса (как словарь) или None, если сотрудник
    ещё не начинал обучение."""
    for r in learning_all():
        if str(r["tg_id"]) == str(tg_id):
            return r
    return None

def learning_set_field(tg_id, field, value):
    """Записывает значение в служебную колонку (started_at/completed_at/last_reminder_at)."""
    ws = learning_sheet()
    columns = learning_columns()
    if field not in columns:
        return
    col_index = columns.index(field) + 1
    records = safe_records(ws, columns)
    for i, r in enumerate(records, start=2):
        if str(r["tg_id"]) == str(tg_id):
            ws.update_cell(i, col_index, value)
            invalidate_cache("learning_progress")
            return

def learning_mark_done(tg_id, full_name, scenario_id, version):
    """
    Отмечает сценарий пройденным для сотрудника. Если в этой колонке уже стоит
    версия >= текущей — не перезаписывает.
    """
    u = emp_by_id(tg_id)
    role = u["role"] if u else "employee"
    department = get_dept(tg_id) if u else ""

    columns = learning_columns()
    if scenario_id not in columns:
        return
    col_index = columns.index(scenario_id) + 1

    row = learning_find_or_create_row(tg_id, full_name, role, department)
    ws = learning_sheet()

    try:
        current = ws.cell(row, col_index).value or ""
    except Exception:
        current = ""
    if current:
        m = re.search(r"v(\d+)", current)
        if m and int(m.group(1)) >= version:
            return

    now = datetime.now(TZ).strftime("%Y-%m-%d %H:%M")
    ws.update_cell(row, col_index, learning_format_cell(version, now))
    invalidate_cache("learning_progress")

def learning_completed_versions(tg_id) -> dict:
    """Возвращает {scenario_id: max_version_completed} для сотрудника."""
    result = {}
    for r in learning_all():
        if str(r["tg_id"]) != str(tg_id):
            continue
        for s in LEARN_SCENARIOS:
            cell_value = r.get(s["id"], "")
            if not cell_value:
                continue
            m = re.search(r"v(\d+)", cell_value)
            if m:
                result[s["id"]] = int(m.group(1))
        break
    return result

def learning_pending_scenarios(tg_id, role: str) -> list:
    """Сценарии, которые сотрудник ещё не прошёл в актуальной версии."""
    completed = learning_completed_versions(tg_id)
    available = learn_scenarios_for_role(role)
    return [s for s in available if completed.get(s["id"], 0) < s["version"]]

def learning_current_step_index(tg_id, role: str) -> int:
    """
    Возвращает индекс первого непройденного шага в последовательности роли
    (0-indexed). Если все шаги пройдены — возвращает len(sequence), то есть
    обучение полностью завершено. Это и есть механизм строгого порядка:
    шаг N+1 недоступен, пока не пройден шаг N.
    """
    sequence = learn_scenarios_for_role(role)
    completed = learning_completed_versions(tg_id)
    for i, s in enumerate(sequence):
        if completed.get(s["id"], 0) < s["version"]:
            return i
    return len(sequence)

def learning_is_complete(tg_id, role: str) -> bool:
    sequence = learn_scenarios_for_role(role)
    return learning_current_step_index(tg_id, role) >= len(sequence)

def learning_new_scenarios_since_completion(tg_id, role: str) -> list:
    """
    Для сотрудника, который ранее завершил всю последовательность на 100% —
    сценарии с версией выше уже отмеченной, появившиеся ПОСЛЕ завершения.
    Используется для push-уведомления о новой функции без повторного запуска
    всей цепочки (агент 2, риск 5 / решение пользователя).
    """
    if not learning_is_complete(tg_id, role):
        return []
    return learning_pending_scenarios(tg_id, role)

def delete_training_tasks(tg_id):
    """Удаляет все тренажёрные задачи сотрудника из листа tasks (по завершению
    обучения, или после 5 рабочих дней брошенного на середине прогресса)."""
    ws = tasks_sheet()
    records = safe_records(ws, TH)
    to_delete = [
        i + 2 for i, r in enumerate(records)
        if str(r["assigned_to_id"]) == str(tg_id)
        and str(r.get("is_training", "")).upper() == "TRUE"
    ]
    for row_i in sorted(to_delete, reverse=True):
        ws.delete_rows(row_i)
    if to_delete:
        invalidate_cache("tasks")
    return len(to_delete)

def count_working_days_between(start_str: str, end_date) -> int:
    """Считает рабочие дни (пн-пт) между датой start_str (YYYY-MM-DD HH:MM или
    YYYY-MM-DD) и end_date (объект date), не включая выходные."""
    if not start_str:
        return 0
    try:
        start_date = datetime.strptime(start_str[:10], "%Y-%m-%d").date()
    except ValueError:
        return 0
    if start_date >= end_date:
        return 0
    count = 0
    d = start_date
    while d < end_date:
        d += timedelta(days=1)
        if d.weekday() < 5:
            count += 1
    return count

def learning_abandoned_employees() -> list:
    """
    Сотрудники, которые начали обучение, не завершили его, и прогресс
    'застрял' (started_at есть, completed_at пусто). Возвращает список
    словарей строк прогресса — используется для ежедневного напоминания
    в 14:00 МСК и для автоудаления после 5 рабочих дней.
    Исключает founder — если человека повысили до учредителя посреди
    незавершённого обучения, напоминания ему больше не нужны (роль не
    проходит обучение вообще).
    """
    result = []
    for r in learning_all():
        if not (r.get("started_at") and not r.get("completed_at")):
            continue
        u = emp_by_id(int(r["tg_id"])) if r.get("tg_id") else None
        if u and u["role"] == "founder":
            continue
        result.append(r)
    return result

def tasks_sheet():
    ws = sheet("tasks"); ensure_headers(ws, TH); return ws

def tasks_all():
    return safe_records(tasks_sheet(), TH)

def tasks_all_real():
    """
    Все задачи КРОМЕ тренажёрных (is_training=TRUE). Это единственное место,
    где должен происходить такой фильтр — все агрегатные функции (дайджесты,
    /menu, экспорт, /team, аудиты) обязаны использовать именно эту функцию,
    а не tasks_all() напрямую, иначе тренажёрные задачи попадут в реальную
    отчётность и испортят цифры (агент 2, риск 1).
    """
    return [t for t in tasks_all() if str(t.get("is_training", "")).upper() != "TRUE"]

def tasks_all_training(tg_id=None):
    """Только тренажёрные задачи, опционально отфильтрованные по сотруднику —
    используется для автоудаления и напоминаний о брошенном обучении."""
    result = [t for t in tasks_all() if str(t.get("is_training", "")).upper() == "TRUE"]
    if tg_id is not None:
        result = [t for t in result if str(t["assigned_to_id"]) == str(tg_id)]
    return result

def task_by_id(tid):
    for r in tasks_all():
        if r["task_id"] == tid.upper():
            return r
    return None

def tasks_for_user(tg_id):
    """Не фильтрует тренажёрные задачи — используется в /mytasks, где сотрудник
    должен видеть свои тестовые задачи во время прохождения обучения."""
    return [r for r in tasks_all()
            if str(r["assigned_to_id"]) == str(tg_id) and r["status"] not in ("done", "cancelled")]

def tasks_overdue_for_user(tg_id):
    """Просроченные задачи конкретного сотрудника (не done, дедлайн в прошлом).
    Используется при внесении нового плана, чтобы предложить актуализировать
    статус старых просрочек прежде чем плодить новые задачи поверх них."""
    today = today_str()
    return [r for r in tasks_for_user(tg_id)
            if is_overdue(r.get("deadline",""), today)]

def tasks_open():
    return [r for r in tasks_all_real() if r["status"] in ("open","in_progress","paused")]

def tasks_overdue():
    today = today_str()
    return [r for r in tasks_all_real()
            if r["status"] not in ("done", "cancelled") and is_overdue(r.get("deadline",""), today)]

def tasks_due_tomorrow():
    tmr = (today_date() + timedelta(days=1)).strftime("%Y-%m-%d")
    return [r for r in tasks_all_real() if r.get("deadline") == tmr and r["status"] not in ("done", "cancelled")]

def tasks_due_today_list():
    return [r for r in tasks_all_real()
            if r.get("deadline") == today_str() and r["status"] not in ("done", "cancelled")]

def task_find_row(tid):
    for i, r in enumerate(safe_records(tasks_sheet(), TH), start=2):
        if r["task_id"] == tid.upper():
            return i
    return None

def task_create(by_id, by_name, to_id, to_name, title, deadline, source="manual", channel="", is_training=False):
    tid = str(uuid.uuid4())[:8].upper()
    tasks_sheet().append_row([
        tid, by_id, by_name, to_id, to_name, title,
        deadline, "open", datetime.now(TZ).strftime("%Y-%m-%d %H:%M"),
        "", "", "", source, channel, "TRUE" if is_training else ""
    ])
    invalidate_cache("tasks")
    return tid

def task_update_channel(tid, channel: str):
    row = task_find_row(tid)
    if not row: return False
    tasks_sheet().update_cell(row, TH.index("channel") + 1, channel)
    invalidate_cache("tasks")
    return True

def task_update_status(tid, status):
    row = task_find_row(tid)
    if not row: return False
    ws = tasks_sheet()
    ws.update_cell(row, TH.index("status") + 1, status)
    if status == "done":
        ws.update_cell(row, TH.index("done_at") + 1, datetime.now(TZ).strftime("%Y-%m-%d %H:%M"))
    invalidate_cache("tasks")
    return True

def task_update_comment(tid, comment, link=""):
    row = task_find_row(tid)
    if not row: return False
    ws = tasks_sheet()
    ws.update_cell(row, TH.index("comment") + 1, comment)
    if link:
        ws.update_cell(row, TH.index("result_link") + 1, link)
    invalidate_cache("tasks")
    return True

def task_update_title(tid, title: str):
    row = task_find_row(tid)
    if not row: return False
    tasks_sheet().update_cell(row, TH.index("title") + 1, title)
    invalidate_cache("tasks")
    return True

def task_update_deadline(tid, deadline: str):
    row = task_find_row(tid)
    if not row: return False
    tasks_sheet().update_cell(row, TH.index("deadline") + 1, deadline)
    invalidate_cache("tasks")
    return True

def tasks_today_for_user(tg_id):
    """Все активные задачи пользователя для EOD-опроса: задачи с дедлайном сегодня
    ИЛИ просроченные ИЛИ длительные (дедлайн в будущем, но уже в работе/приостановлены).
    Это даёт промежуточный статус каждый вечер даже для многодневных задач.
    Задачи со статусом 'cancelled' исключены навсегда — это и есть весь смысл
    статуса 'Отменено/Не актуально': один раз отмечена, больше не появляется
    в опросе (в отличие от 'Не начато', которая продолжает спрашиваться
    каждый вечер, пока не закрыта)."""
    today = today_str()
    return [r for r in tasks_all()
            if str(r["assigned_to_id"]) == str(tg_id)
            and r["status"] not in ("done", "cancelled")
            and (
                r.get("deadline","") == today           # дедлайн сегодня
                or is_overdue(r.get("deadline",""), today)  # просрочена
                or r["status"] in ("in_progress", "paused")  # длительная, уже стартовала
            )]

# ── PARSE PLAN ────────────────────────────────────────────────────────────────
def parse_plan_items(text: str) -> list:
    """
    Разбирает план на пункты. Каждая строка может опционально содержать дату
    в конце в формате [ДД.ММ.ГГГГ] или (ДД.ММ.ГГГГ) — если её нет, дедлайн = сегодня.
    Пример: "Согласовать смету [25.06.2026]" → задача с дедлайном 25.06.2026.
    Возвращает список (title, deadline) где deadline — строка YYYY-MM-DD.
    """
    items = []
    today = today_str()
    date_pattern = re.compile(r"[\[\(]\s*(\d{2})\.(\d{2})\.(\d{4})\s*[\]\)]\s*$")
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        line = re.sub(r"^\d+[\.\)]\s*", "", line)
        line = re.sub(r"^[-–—•*]\s*", "", line)
        line = re.sub(r"^[✅🪡💘🩵🔜👌📌⚡✔▪▸►→]\s*", "", line)
        line = line.strip()

        deadline = today
        m = date_pattern.search(line)
        if m:
            d, mo, y = m.groups()
            deadline = f"{y}-{mo}-{d}"
            line = date_pattern.sub("", line).strip()

        if len(line) > 3:
            items.append((line, deadline))
    return items

# ── CONVERSATION STATES ───────────────────────────────────────────────────────
S_NAME         = 1
S_DEPT         = 2    # выбор отдела при регистрации
S_PLAN         = 10
S_TASK         = 20
S_EOD_STATUS   = 30   # выбор статуса задачи конец дня
S_EOD_EXTRA    = 32   # были ли другие задачи
S_EOD_EXTRA_TEXT = 33 # текст других задач

# Храним состояние сессии EOD (end-of-day) в bot_data
# bot_data["eod"][tg_id] = {"tasks": [...], "current_idx": 0, "results": [...]}

def dept_keyboard(prefix: str, tg_id_for_callback: str = "") -> InlineKeyboardMarkup:
    """Клавиатура выбора отдела. prefix: 'reg_dept_' (при регистрации) или 'admset_dept_{tg_id}_'."""
    rows = []
    row = []
    for d in DEPT_LIST:
        cb = f"{prefix}{tg_id_for_callback}_{d}" if tg_id_for_callback else f"{prefix}{d}"
        row.append(InlineKeyboardButton(d, callback_data=cb))
        if len(row) == 2:
            rows.append(row); row = []
    if row: rows.append(row)
    return InlineKeyboardMarkup(rows)

async def cmd_ping(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Мгновенный health-check, без обращения к Google Sheets — проверяет,
    что бот вообще отвечает, до проверки конкретных функций. Показывает время
    запуска текущего процесса — если оно неожиданно старое, значит Railway
    не подхватил последний деплой и крутит старый контейнер."""
    uptime = datetime.now(TZ) - PROCESS_STARTED_AT
    hours = int(uptime.total_seconds() // 3600)
    minutes = int((uptime.total_seconds() % 3600) // 60)
    await update.effective_message.reply_text(
        f"🟢 Бот на связи.\n"
        f"Сейчас: {datetime.now(TZ).strftime('%d.%m.%Y %H:%M:%S')} МСК\n"
        f"Запущен: {PROCESS_STARTED_AT.strftime('%d.%m.%Y %H:%M:%S')} МСК "
        f"({hours}ч {minutes}мин назад)"
    )

async def cmd_testdigest(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """
    Диагностика отправки сводного отчёта в групповой чат (ADMIN_CHAT_ID).
    Раньше ошибка отправки только тихо логировалась (logger.warning) и
    проглатывалась — если бот не состоит в группе, потерял права писать,
    или ADMIN_CHAT_ID указывает не туда, сводка просто не приходила без
    единого видимого следа. Эта команда сразу показывает настоящую причину.
    """
    if not emp_is_admin(update.effective_user.id):
        await update.effective_message.reply_text("⛔ Только для администраторов."); return
    await update.effective_message.reply_text(
        f"🔍 Проверяю отправку в группу руководителей...\nADMIN_CHAT_ID: <code>{ADMIN_CHAT_ID}</code>",
        parse_mode="HTML"
    )
    try:
        chat = await ctx.bot.get_chat(ADMIN_CHAT_ID)
        await ctx.bot.send_message(
            ADMIN_CHAT_ID,
            f"✅ Тестовое сообщение от бота.\nОтправлено: {datetime.now(TZ).strftime('%d.%m.%Y %H:%M:%S')} МСК"
        )
        await update.effective_message.reply_text(
            f"✅ Успешно! Группа: «{chat.title}» (id {chat.id}). Сообщение доставлено."
        )
    except Exception as ex:
        await update.effective_message.reply_text(
            f"❌ Не удалось отправить в ADMIN_CHAT_ID={ADMIN_CHAT_ID}.\n\n"
            f"Причина: <code>{ex}</code>\n\n"
            f"Частые причины: бота удалили из группы, группу превратили в "
            f"супергруппу (id меняется), у бота нет прав писать сообщения, "
            f"или ADMIN_CHAT_ID в Variables на Railway указывает на старый чат.",
            parse_mode="HTML"
        )

async def cmd_help(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """/help — список команд по роли в любой момент, не только при регистрации.
    Использует тот же COMMAND_REGISTRY, что /start — единый источник правды."""
    tg_id = update.effective_user.id
    if not emp_registered(tg_id):
        await update.effective_message.reply_text("Сначала /start"); return
    u = emp_by_id(tg_id)
    commands_text = format_commands_text(u["role"])
    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("📚 Обучение", callback_data="learn_main"),
    ]])
    await update.effective_message.reply_text(
        f"📖 <b>Доступные команды:</b>\n{commands_text}",
        parse_mode="HTML", reply_markup=keyboard
    )

# ── /start ────────────────────────────────────────────────────────────────────
async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id
    if emp_registered(tg_id):
        u = emp_by_id(tg_id)

        if u["role"] == "founder":
            await update.effective_message.reply_text(
                f"👋 Привет, {u['full_name']}! Ты учредитель.\n\n"
                "Ты видишь результаты работы и задачи всей компании в режиме "
                "просмотра — планы, отчёты и закрытие дня тебя не касаются.\n\n"
                "/menu — открыть панель просмотра"
            )
            return ConversationHandler.END

        role_labels = {"admin": "руководитель", "dept_head": "руководитель отдела", "employee": "сотрудник"}
        role = role_labels.get(u["role"], "сотрудник")
        dept = get_dept(tg_id)
        commands_text = format_commands_text(u["role"])
        keyboard = InlineKeyboardMarkup([[
            InlineKeyboardButton("📚 Обучение", callback_data="learn_main"),
        ]])
        await update.effective_message.reply_text(
            f"👋 Привет, {u['full_name']}! Ты зарегистрирован как {role} ({dept}).\n\n"
            f"Вот что ты можешь делать:\n{commands_text}",
            reply_markup=keyboard
        )
        return ConversationHandler.END
    await update.effective_message.reply_text(
        "👋 Привет! Это трекер команды Brûler d'Amour.\nНапиши имя и фамилию:"
    )
    return S_NAME

async def recv_name(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw_name = update.message.text.strip()
    name = normalize_full_name(raw_name)
    ctx.user_data["reg_name"] = name
    await update.message.reply_text(
        f"Приятно познакомиться, {name}!\nВыбери свой отдел:",
        reply_markup=dept_keyboard("reg_dept_")
    )
    return ConversationHandler.END

async def cb_reg_dept(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Сотрудник выбрал отдел при регистрации."""
    query = update.callback_query
    await query.answer()
    tg_id = query.from_user.id
    dept = query.data.replace("reg_dept_", "", 1)
    name = ctx.user_data.get("reg_name") or query.from_user.full_name or "Без имени"
    username = query.from_user.username or ""

    emp_register(tg_id, username, name, role="employee", department=dept)

    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("📚 Начать обучение", callback_data="learn_main"),
    ]])
    await query.edit_message_text(
        f"✅ Готово, {name}! Отдел: <b>{dept}</b>\n\n"
        "Каждый день в 11:00 — план, в 19:00 — закрытие дня.\n\n"
        "Прежде чем начать — пройди короткое обучение, это займёт пару минут "
        "и покажет, как тут всё работает:",
        parse_mode="HTML",
        reply_markup=keyboard
    )

    # уведомляем всех руководителей о новом сотруднике
    notify_text = (
        f"🆕 <b>Новый сотрудник зарегистрирован</b>\n\n"
        f"Имя: {name}\n"
        f"Username: @{username}\n" if username else f"🆕 <b>Новый сотрудник зарегистрирован</b>\n\nИмя: {name}\n"
    )
    notify_text += f"Отдел (указал сам): {dept}\n\nМожешь изменить отдел или назначить руководителем отдела:"
    keyboard_rows = [
        [InlineKeyboardButton(f"📁 {d}", callback_data=f"admset_dept_{tg_id}_{d}") for d in DEPT_LIST[i:i+2]]
        for i in range(0, len(DEPT_LIST), 2)
    ]
    keyboard_rows.append([InlineKeyboardButton("⭐ Сделать руководителем отдела", callback_data=f"admset_head_{tg_id}")])
    keyboard = InlineKeyboardMarkup(keyboard_rows)

    for adm in emp_admins():
        try:
            await ctx.bot.send_message(int(adm["tg_id"]), notify_text, parse_mode="HTML", reply_markup=keyboard)
        except Exception as e:
            logger.warning(f"notify admin {adm['tg_id']}: {e}")

    return ConversationHandler.END

async def cb_admset_dept(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Руководитель меняет отдел сотрудника из уведомления о регистрации."""
    query = update.callback_query
    if not emp_is_admin(query.from_user.id):
        await query.answer("⛔ Только для администраторов.", show_alert=True); return
    await query.answer()
    # callback: admset_dept_{tg_id}_{Отдел}
    rest = query.data.replace("admset_dept_", "", 1)
    tg_id_str, dept = rest.split("_", 1)
    emp_set_department(int(tg_id_str), dept)
    emp = emp_by_id(int(tg_id_str))
    await query.edit_message_text(
        query.message.text_html + f"\n\n✅ Отдел установлен: <b>{dept}</b>",
        parse_mode="HTML"
    )

async def cb_admset_head(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Руководитель назначает сотрудника руководителем отдела из уведомления о регистрации."""
    query = update.callback_query
    if not emp_is_admin(query.from_user.id):
        await query.answer("⛔ Только для администраторов.", show_alert=True); return
    await query.answer()
    tg_id_str = query.data.replace("admset_head_", "", 1)
    emp_set_role(int(tg_id_str), "dept_head")
    emp = emp_by_id(int(tg_id_str))
    dept = get_dept(int(tg_id_str))
    await query.edit_message_text(
        query.message.text_html + f"\n\n⭐ Назначен руководителем отдела «{dept}»",
        parse_mode="HTML"
    )
    try:
        await ctx.bot.send_message(
            int(tg_id_str),
            f"⭐ Тебя назначили руководителем отдела «{dept}»!\n"
            f"Теперь доступна команда /menu — сводки и задачи по твоему отделу."
        )
    except Exception: pass

# ── /plan ─────────────────────────────────────────────────────────────────────
async def cmd_plan(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.callback_query:
        await update.callback_query.answer()
        # learntry_plan — точка входа в ConversationHandler напрямую (см. main()),
        # минует cb_learn_try, который единственный обычно выставляет флаг
        # обучения. Без этой строки recv_plan ниже всегда считал бы is_training
        # ложным, даже когда план запущен явно из карточки обучения.
        if update.callback_query.data == "learntry_plan":
            mark_learning_action(update.callback_query.from_user.id, "plan")
    if not emp_registered(update.effective_user.id):
        await update.effective_message.reply_text("Сначала /start")
        return ConversationHandler.END
    u = emp_by_id(update.effective_user.id)
    await update.effective_message.reply_text(
        f"📋 {u['full_name']}, напиши план на сегодня:\n\n"
        "Каждый пункт с новой строки — они автоматически станут задачами в трекере."
    )
    return S_PLAN

async def recv_plan(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id
    u = emp_by_id(tg_id)
    text = update.message.text.strip()
    today = today_str()
    is_training = _learning_in_progress.get(tg_id) == "plan"

    if not is_training:
        save_plan(tg_id, u["full_name"], text)

        # удаляем старые задачи из плана на сегодня (только реальные, не тренажёрные)
        ws = tasks_sheet()
        all_rows = safe_records(ws, TH)
        to_delete = [i+2 for i, r in enumerate(all_rows)
                     if str(r["assigned_to_id"]) == str(tg_id)
                     and r.get("created_at","").startswith(today)
                     and r.get("source","") == "plan"
                     and str(r.get("is_training","")).upper() != "TRUE"]
        for row_i in sorted(to_delete, reverse=True):
            ws.delete_rows(row_i)
        if to_delete:
            invalidate_cache("tasks")

    # создаём задачи из плана (тренажёрные, если запущено из обучения)
    items = parse_plan_items(text)
    created_ids = []
    for item, item_deadline in items:
        tid = task_create(tg_id, u["full_name"], tg_id, u["full_name"],
                          item, item_deadline, source="plan", is_training=is_training)
        created_ids.append((tid, item, item_deadline))

    if created_ids:
        task_lines = []
        for tid, title, dl in created_ids:
            dl_mark = f" 📅{fmt_dl(dl)}" if dl != today else ""
            task_lines.append(f"  • <code>{tid}</code> {title}{dl_mark}")
        task_list = "\n".join(task_lines)
        keyboard = InlineKeyboardMarkup([[
            InlineKeyboardButton("✅ Подтвердить задачи", callback_data=f"confirm_plan_{tg_id}"),
        ]])
        training_note = "\n\n<i>🎓 Это тренажёр — задачи учебные, реальный план не записан.</i>" if is_training else ""
        await update.message.reply_text(
            f"📋 План записан! Создано <b>{len(created_ids)} задач</b>:\n\n{task_list}{training_note}\n\n"
            f"Совет: укажи [ДД.ММ.ГГГГ] в конце строки, если срок не сегодня.\n"
            f"Нажми кнопку чтобы подтвердить — или добавь задачи вручную через /task",
            parse_mode="HTML", reply_markup=keyboard
        )
    else:
        learn_kb = pop_learning_continue_keyboard(tg_id)
        await update.message.reply_text("✅ План записан! Увидимся в 19:00 🌆", reply_markup=learn_kb)

    # актуализация статусов по старым просроченным задачам — не даём им
    # просто копиться под новым планом без внимания
    if not is_training:
        overdue = tasks_overdue_for_user(tg_id)
        if overdue:
            ob_lines = [f"⚠️ <b>У тебя {len(overdue)} просроченных задач из прошлых дней.</b>",
                        "Актуализируй статус, прежде чем начинать новый день:\n"]
            buttons = []
            for t in overdue[:15]:
                label = f"{t['title'][:38]}" + ("…" if len(t['title']) > 38 else "") + f" (до {fmt_dl(t['deadline'])})"
                buttons.append([InlineKeyboardButton(label, callback_data=f"chstatustask_{t['task_id']}")])
            await update.message.reply_text(
                "\n".join(ob_lines), parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup(buttons)
            )
            if len(overdue) > 15:
                await update.message.reply_text(f"…и ещё {len(overdue)-15}. Используй /changestatus для полного списка.")

    return ConversationHandler.END

async def cb_confirm_plan(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer("✅ Задачи активированы!")
    await query.edit_message_reply_markup(reply_markup=None)
    tg_id = query.data.split("_")[-1]
    tasks = tasks_today_for_user(int(tg_id))
    learn_kb = pop_learning_continue_keyboard(int(tg_id))
    await query.message.reply_text(
        f"✅ {len(tasks)} задач активны на сегодня.\n"
        f"/mytasks — посмотреть список\n"
        f"В 19:00 я попрошу статус по каждой.",
        reply_markup=learn_kb
    )

# ── END-OF-DAY FLOW ───────────────────────────────────────────────────────────
# Состояние активных EOD-сессий по tg_id. Module-level, а не ctx.bot_data,
# потому что start_eod_flow вызывается и из обработчиков (ctx доступен),
# и из фоновых заданий APScheduler (доступен только объект Bot, у которого
# нет bot_data — это атрибут Application, не Bot).
_eod_state = {}

# Отмечает, что пользователь запустил действие (план/задачу/тег и т.д.) из
# карточки обучения, через кнопку "Попробовать сейчас". Используется чтобы
# после РЕАЛЬНОГО завершения действия засчитать шаг пройденным и сразу
# предложить кнопку "Продолжить обучение", вместо того чтобы человек искал
# глазами старое сообщение с карточками выше.
_learning_in_progress = {}  # tg_id -> scenario_id

def mark_learning_action(tg_id, scenario_id):
    _learning_in_progress[tg_id] = scenario_id

def pop_learning_continue_keyboard(tg_id):
    """
    Если действие было запущено из обучения — засчитывает шаг пройденным
    (только теперь, по факту завершения действия, не по факту просмотра
    карточки) и возвращает клавиатуру с кнопкой 'Продолжить обучение'.
    Если это был последний шаг последовательности — автоматически удаляет
    все тренажёрные задачи сотрудника и фиксирует completed_at, чтобы
    больше не считать обучение 'брошенным на середине'.
    Если действие не из обучения — возвращает None, обычный поток не
    получает лишних кнопок и прогресс не засчитывается.
    """
    scenario_id = _learning_in_progress.pop(tg_id, None)
    if not scenario_id:
        return None
    scenario = next((s for s in LEARN_SCENARIOS if s["id"] == scenario_id), None)
    if not scenario:
        return None
    u = emp_by_id(tg_id)
    if not u:
        return None
    learning_mark_done(tg_id, u["full_name"], scenario["id"], scenario["version"])

    role = u["role"]
    if learning_is_complete(tg_id, role):
        delete_training_tasks(tg_id)
        learning_set_field(tg_id, "completed_at", datetime.now(TZ).strftime("%Y-%m-%d %H:%M"))
        return InlineKeyboardMarkup([[
            InlineKeyboardButton("🎉 Посмотреть итог", callback_data="learn_main"),
        ]])

    return InlineKeyboardMarkup([[
        InlineKeyboardButton("📚 Продолжить обучение", callback_data="learn_main"),
    ]])


def status_keyboard(task_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("✅ Выполнено",     callback_data=f"eods_done_{task_id}"),
            InlineKeyboardButton("🔄 В работе",      callback_data=f"eods_in_progress_{task_id}"),
        ],
        [
            InlineKeyboardButton("⏸ Приостановлено", callback_data=f"eods_paused_{task_id}"),
            InlineKeyboardButton("⬜ Не начато",      callback_data=f"eods_open_{task_id}"),
        ],
        [
            InlineKeyboardButton("🚫 Отменено / Не актуально", callback_data=f"eods_cancelled_{task_id}"),
        ],
    ])

async def start_eod_flow(bot: Bot, tg_id: int, training_task_ids=None):
    """
    Запускает опрос статусов задач для сотрудника.
    Важно: эта функция вызывается и из обработчиков команд (где доступен ctx.bot_data),
    и из фоновых заданий APScheduler (где доступен только сам bot, без ctx).
    Поэтому состояние EOD-сессии хранится в module-level _eod_state, а не в
    ctx.bot_data/bot.bot_data — это единое хранилище, видимое из обоих мест.

    training_task_ids: если указан, опрос проходит только по этим конкретным
    задачам (используется тренажёром обучения), а не по всем реальным
    активным задачам пользователя — это изолирует обучение от рабочих данных.
    """
    if training_task_ids:
        tasks = [task_by_id(tid) for tid in training_task_ids]
        tasks = [t for t in tasks if t]
    else:
        tasks = tasks_today_for_user(tg_id)
    if not tasks:
        try:
            learn_kb = pop_learning_continue_keyboard(tg_id)
            await bot.send_message(tg_id,
                "🌆 Рабочий день завершается!\n"
                "Задач на сегодня не было.",
                reply_markup=learn_kb)
        except Exception as e:
            logger.warning(f"eod no tasks {tg_id}: {e}")
        return

    if tg_id in _eod_state:
        # У человека уже есть активная EOD-сессия (например, он отвечает на
        # статусы, а тут параллельно сработал автоматический job или повторное
        # нажатие /endday) — не перезаписываем её новой пустой, иначе
        # накопленные results теряются и в конце получается 'Выполнено 0 из 0'
        # хотя реально было отмечено несколько задач.
        logger.info(f"start_eod_flow: session already active for {tg_id}, skipping re-init")
        return

    _eod_state[tg_id] = {
        "tasks": tasks,
        "current_idx": 0,
        "results": []
    }

    task = tasks[0]
    try:
        await bot.send_message(
            tg_id,
            f"🌆 Подводим итоги дня! У тебя <b>{len(tasks)} задач</b>.\n\n"
            f"Задача 1/{len(tasks)}:\n<b>{task['title']}</b>\n\n"
            "Выбери статус:",
            parse_mode="HTML",
            reply_markup=status_keyboard(task["task_id"])
        )
    except Exception as e:
        logger.warning(f"eod start {tg_id}: {e}")

async def cb_eod_status(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Callback: сотрудник выбрал статус задачи → сразу предлагаем выбрать канал."""
    query = update.callback_query
    await query.answer()
    tg_id = query.from_user.id

    parts = query.data.split("_")
    # format: eods_{status}_{task_id}
    # eods can be: eods_in_progress, eods_done, eods_paused, eods_open
    task_id = parts[-1]
    status = "_".join(parts[1:-1])  # handles in_progress

    task_update_status(task_id, status)

    eod = _eod_state.get(tg_id)
    if not eod:
        await query.edit_message_reply_markup(reply_markup=None)
        return

    eod["results"].append({"task_id": task_id, "status": status})

    status_labels = {
        "done": "✅ Выполнено",
        "in_progress": "🔄 В работе",
        "paused": "⏸ Приостановлено",
        "open": "⬜ Не начато",
        "cancelled": "🚫 Отменено"
    }
    label = status_labels.get(status, status)

    await query.edit_message_text(
        query.message.text + f"\n\n<b>Статус: {label}</b>",
        parse_mode="HTML"
    )

    if status == "cancelled":
        # Отменённая задача не требует канала/комментария — сразу к следующей.
        eod["current_idx"] += 1
        idx = eod["current_idx"]
        if idx < len(eod["tasks"]):
            next_task = eod["tasks"][idx]
            await query.message.reply_text(
                f"Задача {idx+1}/{len(eod['tasks'])}:\n<b>{next_task['title']}</b>\n\nВыбери статус:",
                parse_mode="HTML",
                reply_markup=status_keyboard(next_task["task_id"])
            )
        else:
            await ask_extra_tasks(query.message, ctx, tg_id)
        return

    task = task_by_id(task_id)
    if task and task.get("channel"):
        # Канал уже проставлен раньше (например, у длительной задачи,
        # которая проходит через EOD не первый раз) — не спрашиваем снова,
        # одного запроса на задачу достаточно. Сразу переходим к комментарию.
        ctx.bot_data.setdefault("eod_pending_comment", {})[tg_id] = task_id
        await query.message.reply_text(
            f"💬 Комментарий по задаче:\n<b>{eod['tasks'][eod['current_idx']]['title']}</b>\n\n"
            "Опиши результат / причину статуса.\n"
            "Если есть документ — прикрепи ссылку в конце через пробел или с новой строки.",
            parse_mode="HTML"
        )
        return

    # Сразу предлагаем выбрать канал продаж — без отдельного текстового комментария
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton(ch, callback_data=f"eodchannel_{task_id}_{ch}") for ch in CHANNEL_LIST[i:i+2]]
        for i in range(0, len(CHANNEL_LIST), 2)
    ])
    await query.message.reply_text(
        f"🏷 К какому каналу относится задача:\n<b>{eod['tasks'][eod['current_idx']]['title']}</b>",
        parse_mode="HTML",
        reply_markup=keyboard
    )

async def cb_eod_channel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Канал выбран кнопкой — сохраняем тег и запрашиваем обязательный комментарий."""
    query = update.callback_query
    await query.answer("✅ Записано")
    tg_id = query.from_user.id

    rest = query.data.replace("eodchannel_", "", 1)
    task_id, channel = rest.split("_", 1)
    task_update_channel(task_id, channel)

    await query.edit_message_text(
        query.message.text + f"\n\n<b>Канал: {channel}</b>",
        parse_mode="HTML"
    )

    eod = _eod_state.get(tg_id)
    if not eod:
        return

    ctx.bot_data.setdefault("eod_pending_comment", {})[tg_id] = task_id
    await query.message.reply_text(
        f"💬 Обязательный комментарий по задаче:\n<b>{eod['tasks'][eod['current_idx']]['title']}</b>\n\n"
        "Опиши результат / причину статуса.\n"
        "Если есть документ — прикрепи ссылку в конце через пробел или с новой строки.\n\n"
        "<i>Пример: Выгрузила отчёт, согласовала с Настей\nhttps://docs.google.com/...</i>",
        parse_mode="HTML"
    )

async def recv_eod_comment(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Получаем обязательный комментарий к задаче — последний шаг перед следующей задачей."""
    tg_id = update.effective_user.id
    pending = ctx.bot_data.get("eod_pending_comment", {})
    task_id = pending.get(tg_id)
    if not task_id:
        return  # не в этом потоке

    text = update.message.text.strip()

    # ищем ВСЕ ссылки в тексте — у задачи может быть несколько результатов
    # (документ + фото + переписка), раньше сохранялась только первая
    url_pattern = r'https?://\S+'
    links = re.findall(url_pattern, text)
    link = "\n".join(links) if links else ""
    comment = re.sub(url_pattern, "", text).strip()

    task_update_comment(task_id, comment, link)
    del pending[tg_id]

    eod = _eod_state.get(tg_id)
    if not eod:
        return

    eod["current_idx"] += 1
    idx = eod["current_idx"]

    if idx < len(eod["tasks"]):
        # следующая задача
        task = eod["tasks"][idx]
        await update.message.reply_text(
            f"Задача {idx+1}/{len(eod['tasks'])}:\n<b>{task['title']}</b>\n\nВыбери статус:",
            parse_mode="HTML",
            reply_markup=status_keyboard(task["task_id"])
        )
    else:
        # все задачи обработаны
        await ask_extra_tasks(update.message, ctx, tg_id)

async def ask_extra_tasks(message, ctx, tg_id):
    """Спрашиваем были ли другие задачи."""
    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ Да, были", callback_data="eod_extra_yes"),
        InlineKeyboardButton("❌ Нет", callback_data="eod_extra_no"),
    ]])
    await message.reply_text(
        "🎉 Все задачи из плана отмечены!\n\n"
        "Были ли сегодня <b>другие задачи</b>, не из плана?",
        parse_mode="HTML",
        reply_markup=keyboard
    )

async def cb_eod_extra_yes(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_reply_markup(reply_markup=None)
    ctx.bot_data.setdefault("eod_extra_pending", set()).add(query.from_user.id)
    await query.message.reply_text(
        "📝 Напиши что ещё сделал сегодня (каждое дело с новой строки):"
    )

async def cb_eod_extra_no(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_reply_markup(reply_markup=None)
    tg_id = query.from_user.id
    await finish_eod(query.message, ctx, tg_id)

async def recv_eod_extra(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Получаем текст доп. задач."""
    tg_id = update.effective_user.id
    extra_set = ctx.bot_data.get("eod_extra_pending", set())
    if tg_id not in extra_set:
        return

    extra_set.discard(tg_id)
    text = update.message.text.strip()
    items = parse_plan_items(text)
    u = emp_by_id(tg_id)

    eod = _eod_state.get(tg_id)
    for item, item_deadline in items:
        tid = task_create(tg_id, u["full_name"], tg_id, u["full_name"],
                          item, item_deadline, source="extra")
        task_update_status(tid, "done")
        # Без этой строки задачи сверх плана создавались и закрывались, но
        # никогда не попадали в eod['results'] — finish_eod считал только
        # исходный план, и счётчик 'Выполнено X из Y' полностью игнорировал
        # дополнительно добавленные задачи, даже если все они выполнены.
        if eod is not None:
            eod["results"].append({"task_id": tid, "status": "done"})

    await update.message.reply_text(f"✅ Добавлено ещё {len(items)} выполненных задач.")
    await finish_eod(update.message, ctx, tg_id)

async def finish_eod(message, ctx, tg_id):
    """Завершаем EOD. Автоматически сохраняем сводку как 'отчёт' — это даёт
    руководителям статистику 'кто закрыл день' без отдельного текстового
    отчёта от сотрудника, который мы убрали по их запросу."""
    eod = _eod_state.pop(tg_id, {})
    results = eod.get("results", [])
    done_c = sum(1 for r in results if r["status"] == "done")
    total_c = len(results)

    streak_line = ""
    if total_c > 0 and done_c == total_c:
        streak = calc_streak(tg_id)
        if streak >= 2:
            streak_line = f"\n🔥 {streak} дней подряд все задачи закрыты!"
        elif streak == 1:
            streak_line = "\n🌱 Отличное начало серии — продолжай завтра!"

    u = emp_by_id(tg_id)
    is_training_eod = _learning_in_progress.get(tg_id) == "eod"
    if u and not has_report_today(tg_id) and not is_training_eod:
        summary = f"EOD пройден: {done_c}/{total_c} задач выполнено"
        save_report(tg_id, u["full_name"], summary)

    # Конец рабочего дня фиксируется только теперь — после реальной
    # актуализации статусов, не просто по факту намерения закончить.
    if u and not is_training_eod:
        workday_mark_end(tg_id)

    learn_kb = pop_learning_continue_keyboard(tg_id)
    if is_training_eod:
        delete_training_tasks(tg_id)
    await message.reply_text(
        f"✅ День закрыт! Выполнено {done_c} из {total_c} задач.{streak_line}\n\n"
        "Хорошего вечера! 🌙",
        reply_markup=learn_kb
    )

# ── /eod — запуск вручную ────────────────────────────────────────────────────
async def cmd_eod(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id
    if not emp_registered(tg_id):
        await update.effective_message.reply_text("Сначала /start"); return

    is_training = False
    if update.callback_query and update.callback_query.data == "learntry_eod":
        mark_learning_action(tg_id, "eod")
        is_training = True

    if is_training:
        # Тренажёр: создаём 2 изолированные тестовые задачи вместо того,
        # чтобы запускать опрос по всем реальным активным задачам сотрудника
        # (раньше тут подставлялись настоящие рабочие задачи целиком — баг).
        u = emp_by_id(tg_id)
        today = today_str()
        demo_titles = ["Учебная задача: позвонить клиенту", "Учебная задача: подготовить отчёт"]
        demo_ids = []
        for title in demo_titles:
            tid = task_create(tg_id, u["full_name"], tg_id, u["full_name"],
                              title, today, source="plan", is_training=True)
            demo_ids.append(tid)
        await start_eod_flow(ctx.bot, tg_id, training_task_ids=demo_ids)
        return

    # Та же защита от утечки состояния, что в cmd_endday — иначе застрявший
    # флаг обучения мог бы заставить finish_eod принять настоящий EOD за тренажёр.
    _learning_in_progress.pop(tg_id, None)
    await start_eod_flow(ctx.bot, tg_id)

async def cmd_startday(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """
    /startday — сотрудник самостоятельно отмечает начало рабочего дня в любой
    момент. Сразу запускает запрос плана (заменяет старый фиксированный
    автозапрос в 11:00 — теперь план запрашивается по факту начала дня,
    не по единому времени для всех).
    """
    if update.callback_query:
        try:
            await update.callback_query.answer()
        except Exception as ex:
            # Telegram отклоняет answer() для callback_query старше нескольких
            # минут ('Query is too old'). Без этого try/except исключение
            # уходило бы наружу до выполнения остальной логики — кнопка
            # оставалась в состоянии 'Загрузка...' навсегда, хотя реальное
            # действие (отметка старта дня) могло бы пройти штатно дальше.
            logger.warning(f"startday answer expired: {ex}")
        # Кнопка 'Попробовать /startday' из карточки обучения зарегистрирована
        # как прямая точка входа в ConversationHandler (см. main()), минуя
        # cb_learn_try целиком — поэтому mark_learning_action там никогда не
        # вызывается для этого сценария. Выставляем флаг здесь явно, иначе
        # шаг обучения никогда не засчитывается и пользователь застревает.
        if update.callback_query.data == "learntry_startday":
            mark_learning_action(update.callback_query.from_user.id, "startday")

    tg_id = update.effective_user.id
    if not emp_registered(tg_id):
        await update.effective_message.reply_text("Сначала /start"); return ConversationHandler.END
    u = emp_by_id(tg_id)
    try:
        marked = workday_mark_start(tg_id, u["full_name"])
    except Exception as ex:
        logger.error(f"workday_mark_start failed for {tg_id}: {ex}")
        await update.effective_message.reply_text(
            "⚠️ Не получилось записать начало дня — проблема с таблицей. Попробуй ещё раз через минуту."
        )
        return ConversationHandler.END

    if not marked:
        # День уже был отмечен раньше — это нормально, если человек реально
        # начал работать до того, как зашёл в обучение/нажал кнопку повторно.
        # Шаг обучения всё равно нужно засчитать (цель — научить пользоваться
        # командой, а не заставить физически переотмечать уже начатый день) —
        # без этого пользователь застревал бы тут навсегда.
        in_learning = _learning_in_progress.get(tg_id) == "startday"
        await update.effective_message.reply_text("✅ Ты уже отметил(а) начало рабочего дня сегодня.")
        if in_learning:
            learn_kb = pop_learning_continue_keyboard(tg_id)
            await update.effective_message.reply_text(
                "Это нормально — в реальной работе ты обычно отмечаешь старт только один раз в день.",
                reply_markup=learn_kb
            )
        return ConversationHandler.END

    await update.effective_message.reply_text(f"☀️ Рабочий день начат! {datetime.now(TZ).strftime('%H:%M')} МСК")
    return await cmd_plan(update, ctx)

async def cmd_endday(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """
    /endday — запускает актуализацию статусов задач (EOD). Сама отметка конца
    рабочего дня фиксируется ПОСЛЕ завершения EOD-опроса (внутри finish_eod),
    не сразу при нажатии — конец дня засчитывается только когда статусы
    реально актуализированы, не просто по факту намерения закончить.
    """
    tg_id = update.effective_user.id
    if not emp_registered(tg_id):
        await update.effective_message.reply_text("Сначала /start"); return

    is_training = False
    if update.callback_query and update.callback_query.data == "learntry_eod":
        mark_learning_action(tg_id, "eod")
        is_training = True

    if is_training:
        # Тренажёр: создаём 2 изолированные тестовые задачи вместо того,
        # чтобы запускать опрос по всем реальным активным задачам сотрудника.
        # Кнопка 'Попробовать /endday' из карточки обучения вызывает именно
        # cmd_endday (не cmd_eod) — это была реальная причина бага: изоляция
        # тренажёра была добавлена в cmd_eod, которую этот путь не вызывает.
        u = emp_by_id(tg_id)
        today = today_str()
        demo_titles = ["Учебная задача: позвонить клиенту", "Учебная задача: подготовить отчёт"]
        demo_ids = []
        for title in demo_titles:
            tid = task_create(tg_id, u["full_name"], tg_id, u["full_name"],
                              title, today, source="plan", is_training=True)
            demo_ids.append(tid)
        await start_eod_flow(ctx.bot, tg_id, training_task_ids=demo_ids)
        return

    if workday_ended(tg_id):
        # День уже реально закрыт — это нормально, если человек уже работал
        # до обучения. Засчитываем шаг, не оставляем пользователя в тупике
        # (тот же класс проблемы, что был в cmd_startday).
        in_learning = _learning_in_progress.get(tg_id) == "eod"
        await update.effective_message.reply_text("✅ Ты уже завершил(а) рабочий день сегодня.")
        if in_learning:
            learn_kb = pop_learning_continue_keyboard(tg_id)
            await update.effective_message.reply_text(
                "Это нормально — в реальной работе день закрывается один раз.",
                reply_markup=learn_kb
            )
        return

    # Защита от утечки состояния: если человек раньше начинал, но не
    # завершил урок 'Как закрыть день' (бросил обучение на середине), флаг
    # _learning_in_progress[tg_id] == 'eod' мог остаться висеть навсегда —
    # pop_learning_continue_keyboard вызывается только при завершении урока.
    # Без явной очистки здесь finish_eod увидел бы этот старый флаг и принял
    # бы настоящее закрытие дня за тренажёр, тихо не сохранив реальный отчёт.
    _learning_in_progress.pop(tg_id, None)

    await start_eod_flow(ctx.bot, tg_id)

# ── /task ─────────────────────────────────────────────────────────────────────
def parse_task(text):
    m = re.match(r"@(\S+)\s+(.+?)\s*\|\s*(\d{2}\.\d{2}\.\d{4})", text)
    if not m: return None
    username, title, dl = m.groups()
    p = dl.split(".")
    return username, title.strip(), f"{p[2]}-{p[1]}-{p[0]}"

async def cmd_task(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """
    Новый флоу: /task без аргументов → выбор отдела кнопками → выбор сотрудника
    по имени кнопками → текстовый ввод названия задачи и срока.
    Старый текстовый формат @username Название | Дата всё ещё поддерживается
    для тех, кто уже привык.
    """
    if update.callback_query:
        await update.callback_query.answer()
        # Та же причина, что в cmd_plan/cmd_startday — learntry_task минует
        # cb_learn_try, нужно выставить флаг обучения здесь явно.
        if update.callback_query.data == "learntry_task":
            mark_learning_action(update.callback_query.from_user.id, "task")
    if not emp_registered(update.effective_user.id):
        await update.effective_message.reply_text("Сначала /start")
        return ConversationHandler.END
    args = " ".join(ctx.args) if getattr(ctx, "args", None) else ""
    if args and parse_task(args):
        return await do_create_task(update, ctx, parse_task(args))

    # диалоговый режим — выбор отдела
    depts_present = sorted(set(get_dept(e["tg_id"]) for e in emp_employees()))
    buttons = []
    row = []
    for d in depts_present:
        row.append(InlineKeyboardButton(d, callback_data=f"tasknew_dept_{d}"))
        if len(row) == 2:
            buttons.append(row); row = []
    if row: buttons.append(row)
    await update.effective_message.reply_text(
        "📌 Кому поставить задачу? Выбери отдел:",
        reply_markup=InlineKeyboardMarkup(buttons)
    )
    return ConversationHandler.END

async def cb_tasknew_dept(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Выбран отдел — показываем сотрудников этого отдела по именам."""
    query = update.callback_query; await query.answer()
    dept = query.data.replace("tasknew_dept_", "", 1)
    emps = [e for e in emp_employees() if get_dept(e["tg_id"]) == dept]
    if not emps:
        await query.message.reply_text("В этом отделе нет сотрудников."); return
    buttons = []
    row = []
    for e in emps:
        row.append(InlineKeyboardButton(e["full_name"], callback_data=f"tasknew_emp_{e['tg_id']}"))
        if len(row) == 2:
            buttons.append(row); row = []
    if row: buttons.append(row)
    await query.message.reply_text(f"Сотрудники отдела «{dept}»:", reply_markup=InlineKeyboardMarkup(buttons))

async def cb_tasknew_emp(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Выбран сотрудник — просим текст задачи и срок."""
    query = update.callback_query; await query.answer()
    assignee_tg_id = query.data.replace("tasknew_emp_", "", 1)
    assignee = emp_by_id(int(assignee_tg_id))
    if not assignee:
        await query.message.reply_text("Сотрудник не найден."); return
    ctx.user_data["tasknew_assignee"] = assignee_tg_id
    await query.message.reply_text(
        f"Задача для <b>{assignee['full_name']}</b>.\n\n"
        f"Напиши название и срок в формате:\n"
        f"<code>Название задачи | ДД.ММ.ГГГГ</code>\n\n"
        f"Если срок не важен — просто название, поставим на сегодня.",
        parse_mode="HTML"
    )
    ctx.bot_data.setdefault("tasknew_pending", {})[query.from_user.id] = assignee_tg_id

async def recv_tasknew_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Получаем текст задачи после выбора сотрудника кнопками."""
    tg_id = update.effective_user.id
    pending = ctx.bot_data.get("tasknew_pending", {})
    assignee_tg_id = pending.get(tg_id)
    if not assignee_tg_id:
        return  # не в этом потоке
    del pending[tg_id]

    text = update.message.text.strip()
    if "|" in text:
        title, date_part = text.split("|", 1)
        title = title.strip()
        m = re.match(r"^(\d{2})\.(\d{2})\.(\d{4})$", date_part.strip())
        if m:
            d, mo, y = m.groups()
            deadline = f"{y}-{mo}-{d}"
        else:
            deadline = today_str()
    else:
        title = text
        deadline = today_str()

    creator = emp_by_id(tg_id)
    assignee = emp_by_id(int(assignee_tg_id))
    is_training = _learning_in_progress.get(tg_id) == "task"
    tid = task_create(tg_id, creator["full_name"],
                      int(assignee_tg_id), assignee["full_name"], title, deadline,
                      is_training=is_training)
    dl_fmt = fmt_dl(deadline)
    try:
        await ctx.bot.send_message(
            int(assignee_tg_id),
            f"📌 Тебе поставлена задача!\n\n<b>{title}</b>\n"
            f"От: {creator['full_name']}\nСрок: {dl_fmt}\n"
            f"ID: <code>{tid}</code>\n\n/done {tid} — отметить выполненной",
            parse_mode="HTML"
        )
    except Exception: pass
    learn_kb = pop_learning_continue_keyboard(tg_id)
    training_note = "\n\n<i>🎓 Это тренажёр — уведомление исполнителю не имеет реальных последствий.</i>" if is_training else ""
    await update.message.reply_text(
        f"✅ Задача создана!\n<b>{title}</b>\nИсполнитель: {assignee['full_name']}\n"
        f"Срок: {dl_fmt}\nID: <code>{tid}</code>{training_note}", parse_mode="HTML",
        reply_markup=learn_kb
    )

async def recv_task(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    parsed = parse_task(update.message.text.strip())
    if not parsed:
        await update.message.reply_text(
            "Не понял формат:\n<code>@username Название | ДД.ММ.ГГГГ</code>",
            parse_mode="HTML"
        )
        return S_TASK
    return await do_create_task(update, ctx, parsed)

async def do_create_task(update, ctx, parsed):
    username, title, deadline = parsed
    creator = emp_by_id(update.effective_user.id)
    assignee = next((e for e in emp_all()
                     if e["username"].lstrip("@").lower() == username.lower()), None)
    if not assignee:
        await update.effective_message.reply_text(f"❌ @{username} не найден.")
        return ConversationHandler.END
    tid = task_create(update.effective_user.id, creator["full_name"],
                      assignee["tg_id"], assignee["full_name"], title, deadline)
    dl_fmt = deadline[8:]+"."+deadline[5:7]+"."+deadline[:4]
    try:
        await ctx.bot.send_message(
            int(assignee["tg_id"]),
            f"📌 Тебе поставлена задача!\n\n<b>{title}</b>\n"
            f"От: {creator['full_name']}\nСрок: {dl_fmt}\n"
            f"ID: <code>{tid}</code>\n\n/done {tid} — отметить выполненной",
            parse_mode="HTML"
        )
    except Exception: pass
    await update.effective_message.reply_text(
        f"✅ Задача создана!\n<b>{title}</b>\nИсполнитель: {assignee['full_name']}\n"
        f"Срок: {dl_fmt}\nID: <code>{tid}</code>", parse_mode="HTML"
    )
    return ConversationHandler.END

# ── /done /status /mytasks ────────────────────────────────────────────────────
async def cmd_done(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """
    /done — диалоговый выбор: своя активная задача кнопками, отмечается выполненной сразу.
    /done ID [комментарий] — старый текстовый вариант для тех, кто уже привык.
    """
    if not ctx.args:
        tg_id = update.effective_user.id
        my_tasks = tasks_for_user(tg_id)
        if not my_tasks:
            await update.effective_message.reply_text("У тебя нет активных задач."); return
        buttons = []
        for t in my_tasks[:30]:
            label = t["title"][:40] + ("…" if len(t["title"]) > 40 else "")
            buttons.append([InlineKeyboardButton(f"✅ {label}", callback_data=f"donetask_{t['task_id']}")])
        await update.effective_message.reply_text(
            "Какую задачу отметить выполненной?",
            reply_markup=InlineKeyboardMarkup(buttons)
        )
        return

    tid = ctx.args[0].upper()
    comment = " ".join(ctx.args[1:])
    await _mark_task_done(update.effective_message, ctx.bot, update.effective_user.id, tid, comment)

async def _mark_task_done(message, bot, tg_id, tid, comment=""):
    """Общая логика отметки задачи выполненной — используется и текстовой командой, и кнопкой."""
    task = task_by_id(tid)
    if not task:
        await message.reply_text(f"❌ Задача {tid} не найдена."); return
    if str(task["assigned_to_id"]) != str(tg_id) and not emp_is_admin(tg_id):
        await message.reply_text("❌ Это не твоя задача."); return
    task_update_status(tid, "done")
    if comment:
        task_update_comment(tid, comment)
    try:
        await bot.send_message(
            int(task["created_by_id"]),
            f"✅ Задача выполнена!\n<b>{task['title']}</b>\n"
            f"Выполнил: {task['assigned_to_name']}\nID: <code>{tid}</code>"
            + (f"\n{comment}" if comment else ""),
            parse_mode="HTML"
        )
    except Exception: pass
    await message.reply_text(
        f"✅ Задача <code>{tid}</code> выполнена!", parse_mode="HTML"
    )

async def cb_donetask_pick(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Кнопка из /done — отмечаем выбранную задачу выполненной."""
    query = update.callback_query; await query.answer("✅ Отмечено!")
    tid = query.data.replace("donetask_", "", 1)
    await query.edit_message_reply_markup(reply_markup=None)
    await _mark_task_done(query.message, ctx.bot, query.from_user.id, tid)

async def cmd_changestatus(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """
    /changestatus — диалоговый выбор: своя активная задача кнопками →
    любой из 4 статусов кнопками. Не запускает полный EOD-опрос по всем
    задачам — меняет статус только одной выбранной, можно вызывать в
    любой момент дня.
    """
    tg_id = update.effective_user.id
    my_tasks = tasks_for_user(tg_id)
    if not my_tasks:
        await update.effective_message.reply_text("У тебя нет активных задач."); return
    buttons = []
    sl = {"open":"⬜","in_progress":"🔄","paused":"⏸","cancelled":"🚫"}
    for t in my_tasks[:30]:
        icon = sl.get(t["status"], "⚪")
        label = f"{icon} {t['title'][:38]}" + ("…" if len(t['title']) > 38 else "")
        buttons.append([InlineKeyboardButton(label, callback_data=f"chstatustask_{t['task_id']}")])
    await update.effective_message.reply_text(
        "Какой задаче сменить статус?",
        reply_markup=InlineKeyboardMarkup(buttons)
    )

async def cb_changestatus_task_pick(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Выбрана задача — показываем 4 кнопки статуса (отдельный callback-префикс,
    не пересекается с EOD-потоком). Также убираем кнопку выбранной задачи из
    исходного списка, чтобы она не висела там кликабельной после смены статуса."""
    query = update.callback_query; await query.answer()
    tid = query.data.replace("chstatustask_", "", 1)
    task = task_by_id(tid)
    if not task:
        await query.message.reply_text("Задача не найдена."); return

    # убираем кнопку этой задачи из исходного списка выбора
    if query.message.reply_markup:
        new_rows = [
            row for row in query.message.reply_markup.inline_keyboard
            if not any(btn.callback_data == query.data for btn in row)
        ]
        try:
            await query.edit_message_reply_markup(
                reply_markup=InlineKeyboardMarkup(new_rows) if new_rows else None
            )
        except Exception:
            pass  # сообщение могло быть уже изменено параллельно — не критично

    keyboard = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("✅ Выполнено",     callback_data=f"chstatus_done_{tid}"),
            InlineKeyboardButton("🔄 В работе",      callback_data=f"chstatus_in_progress_{tid}"),
        ],
        [
            InlineKeyboardButton("⏸ Приостановлено", callback_data=f"chstatus_paused_{tid}"),
            InlineKeyboardButton("⬜ Не начато",      callback_data=f"chstatus_open_{tid}"),
        ],
        [
            InlineKeyboardButton("🚫 Отменено / Не актуально", callback_data=f"chstatus_cancelled_{tid}"),
        ],
    ])
    await query.message.reply_text(
        f"Новый статус для:\n<b>{task['title']}</b>",
        parse_mode="HTML",
        reply_markup=keyboard
    )

async def cb_changestatus_apply(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Применяем выбранный статус к задаче — без захода в полный EOD-поток."""
    query = update.callback_query
    parts = query.data.split("_")
    task_id = parts[-1]
    status = "_".join(parts[1:-1])  # handles in_progress

    task = task_by_id(task_id)
    if not task:
        await query.answer("Задача не найдена.", show_alert=True); return
    tg_id = query.from_user.id
    if str(task["assigned_to_id"]) != str(tg_id) and not emp_is_admin(tg_id):
        await query.answer("⛔ Это не твоя задача.", show_alert=True); return

    task_update_status(task_id, status)
    status_labels = {
        "done": "✅ Выполнено", "in_progress": "🔄 В работе",
        "paused": "⏸ Приостановлено", "open": "⬜ Не начато", "cancelled": "🚫 Отменено"
    }
    label = status_labels.get(status, status)
    await query.answer(f"Статус: {label}")
    await query.edit_message_text(
        f"<b>{task['title']}</b>\nНовый статус: {label}",
        parse_mode="HTML"
    )
    learn_kb = pop_learning_continue_keyboard(tg_id)
    if learn_kb:
        await query.message.reply_text("Продолжим?", reply_markup=learn_kb)

async def cmd_status(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not ctx.args:
        await update.effective_message.reply_text("Укажи ID: /status ABCD1234"); return
    task = task_by_id(ctx.args[0].upper())
    if not task:
        await update.effective_message.reply_text("❌ Не найдена."); return
    sl = {"open":"⬜ Не начато","in_progress":"🔄 В работе",
          "done":"✅ Выполнена","overdue":"🔴 Просрочена","paused":"⏸ Приостановлено","cancelled":"🚫 Отменено"}
    dl = task["deadline"]
    dl_fmt = dl[8:]+"."+dl[5:7]+"."+dl[:4] if dl else "—"
    link_line = ""
    if task.get("result_link"):
        link_lines = task["result_link"].split("\n")
        link_line = "\n" + "\n".join(f"🔗 {l}" for l in link_lines if l)
    comment_line = f"\n💬 {task['comment']}" if task.get("comment") else ""
    tag_line = f"\n🏷 {task['channel']}" if task.get("channel") else ""
    await update.effective_message.reply_text(
        f"📌 <b>{task['title']}</b>\n{sl.get(task['status'],task['status'])}\n"
        f"Исполнитель: {task['assigned_to_name']}\nПостановщик: {task['created_by_name']}\n"
        f"Срок: {dl_fmt}{tag_line}{comment_line}{link_line}\nID: <code>{task['task_id']}</code>",
        parse_mode="HTML"
    )

async def cmd_edit(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """
    /edit — диалоговый выбор: своя задача кнопками → поле (название/срок) кнопками → текст значения.
    /edit ID title Новое название
    /edit ID deadline ДД.ММ.ГГГГ
    Изменить может постановщик, исполнитель или администратор.
    """
    tg_id = update.effective_user.id
    if len(ctx.args) >= 3:
        tid = ctx.args[0].upper()
        field = ctx.args[1].lower()
        value = " ".join(ctx.args[2:])
        await _apply_edit(update.effective_message, tg_id, tid, field, value)
        return

    # диалоговый режим — задачи, которые можно редактировать: свои + те, что сам поставил
    my_tasks = tasks_for_user(tg_id)
    all_t = tasks_all()
    created_by_me = [t for t in all_t if str(t["created_by_id"]) == str(tg_id)
                     and t["status"] not in ("done", "cancelled") and t["task_id"] not in {x["task_id"] for x in my_tasks}]
    editable = my_tasks + created_by_me
    if not editable:
        await update.effective_message.reply_text("Нет задач, доступных для редактирования."); return

    buttons = []
    for t in editable[:30]:
        label = t["title"][:35] + ("…" if len(t["title"]) > 35 else "")
        buttons.append([InlineKeyboardButton(label, callback_data=f"edittask_{t['task_id']}")])
    await update.effective_message.reply_text(
        "Какую задачу изменить?",
        reply_markup=InlineKeyboardMarkup(buttons)
    )

async def cb_edittask_pick(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Выбрана задача — показываем выбор поля: название или срок."""
    query = update.callback_query; await query.answer()
    tid = query.data.replace("edittask_", "", 1)
    task = task_by_id(tid)
    if not task:
        await query.message.reply_text("Задача не найдена."); return
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("✏️ Название", callback_data=f"editfield_title_{tid}"),
         InlineKeyboardButton("📅 Срок", callback_data=f"editfield_deadline_{tid}")],
    ])
    await query.message.reply_text(
        f"Что изменить в задаче:\n<b>{task['title']}</b>",
        parse_mode="HTML", reply_markup=keyboard
    )

async def cb_editfield_pick(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Выбрано поле — просим новое значение текстом."""
    query = update.callback_query; await query.answer()
    rest = query.data.replace("editfield_", "", 1)
    field, tid = rest.split("_", 1)
    tg_id = query.from_user.id
    ctx.bot_data.setdefault("edit_pending", {})[tg_id] = (tid, field)
    hint = "Напиши новое название:" if field == "title" else "Напиши новый срок в формате ДД.ММ.ГГГГ:"
    await query.message.reply_text(hint)

async def recv_edit_value(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Получаем новое значение поля после выбора кнопками."""
    tg_id = update.effective_user.id
    pending = ctx.bot_data.get("edit_pending", {})
    if tg_id not in pending:
        return
    tid, field = pending.pop(tg_id)
    value = update.message.text.strip()
    await _apply_edit(update.message, tg_id, tid, field, value)

async def _apply_edit(message, tg_id, tid, field, value):
    """Общая логика применения правки — используется текстовой командой и кнопочным флоу."""
    task = task_by_id(tid)
    if not task:
        await message.reply_text(f"❌ Задача {tid} не найдена."); return

    allowed = (str(task["assigned_to_id"]) == str(tg_id)
               or str(task["created_by_id"]) == str(tg_id)
               or emp_is_admin(tg_id))
    if not allowed:
        await message.reply_text("❌ Можно редактировать только свои задачи."); return

    if field == "title":
        task_update_title(tid, value)
        await message.reply_text(f"✅ Название обновлено:\n<b>{value}</b>", parse_mode="HTML")
    elif field == "deadline":
        m = re.match(r"^(\d{2})\.(\d{2})\.(\d{4})$", value.strip())
        if not m:
            await message.reply_text("❌ Формат даты: ДД.ММ.ГГГГ"); return
        d, mo, y = m.groups()
        new_deadline = f"{y}-{mo}-{d}"
        task_update_deadline(tid, new_deadline)
        await message.reply_text(f"✅ Срок обновлён: {value}")
    else:
        await message.reply_text("Поле должно быть title или deadline.")

async def cmd_tag(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """
    /tag — диалоговый выбор: своя активная задача → канал, кнопками.
    /tag ID Канал — прямой текстовый вариант для опытных пользователей.
    """
    tg_id = update.effective_user.id
    if not emp_registered(tg_id):
        await update.effective_message.reply_text("Сначала /start"); return

    if ctx.args and len(ctx.args) >= 2:
        tid = ctx.args[0].upper()
        channel = " ".join(ctx.args[1:])
        if channel not in CHANNEL_LIST:
            await update.effective_message.reply_text(
                "Канал должен быть одним из: " + ", ".join(CHANNEL_LIST)
            )
            return
        task = task_by_id(tid)
        if not task:
            await update.effective_message.reply_text(f"❌ Задача {tid} не найдена."); return
        task_update_channel(tid, channel)
        await update.effective_message.reply_text(f"✅ Тег «{channel}» проставлен для задачи {tid}.")
        return

    my_tasks = tasks_for_user(tg_id)
    if not my_tasks:
        await update.effective_message.reply_text("У тебя нет активных задач для тегирования."); return

    buttons = []
    for t in my_tasks[:20]:
        label = t["title"][:35] + ("…" if len(t["title"]) > 35 else "")
        buttons.append([InlineKeyboardButton(label, callback_data=f"tagtask_{t['task_id']}")])
    await update.effective_message.reply_text(
        "Выбери задачу для тегирования каналом:",
        reply_markup=InlineKeyboardMarkup(buttons)
    )

async def cb_tagtask_pick(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Выбрана задача — показываем список каналов."""
    query = update.callback_query; await query.answer()
    tid = query.data.replace("tagtask_", "", 1)
    buttons = []
    row = []
    for ch in CHANNEL_LIST:
        row.append(InlineKeyboardButton(ch, callback_data=f"tagchannel_{tid}_{ch}"))
        if len(row) == 2:
            buttons.append(row); row = []
    if row: buttons.append(row)
    await query.message.reply_text("Выбери канал:", reply_markup=InlineKeyboardMarkup(buttons))

async def cb_tagchannel_pick(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Выбран канал — сохраняем тег."""
    query = update.callback_query; await query.answer()
    rest = query.data.replace("tagchannel_", "", 1)
    tid, channel = rest.split("_", 1)
    task_update_channel(tid, channel)
    task = task_by_id(tid)
    title = task["title"] if task else tid
    await query.edit_message_text(f"✅ Тег «{channel}» проставлен:\n{title}")
    learn_kb = pop_learning_continue_keyboard(query.from_user.id)
    if learn_kb:
        await query.message.reply_text("Продолжим?", reply_markup=learn_kb)

def calc_streak(tg_id: int) -> int:
    """Считает, сколько последних дней подряд у сотрудника ВСЕ задачи плана
    были закрыты статусом done (без 'paused'/'open' в конце дня)."""
    ws = tasks_sheet()
    all_t = safe_records(ws, TH)
    my_tasks = [t for t in all_t if str(t["assigned_to_id"]) == str(tg_id) and t.get("source") == "plan"]
    if not my_tasks:
        return 0
    by_date: dict = {}
    for t in my_tasks:
        d = t.get("deadline", "")
        if d:
            by_date.setdefault(d, []).append(t)

    streak = 0
    day = today_date()
    # сегодняшний день не считаем, если он ещё не закрыт — начинаем со вчера
    while True:
        d_str = day.strftime("%Y-%m-%d")
        day_tasks = by_date.get(d_str)
        if day_tasks is None:
            # нет задач в этот день — пропускаем день, не прерывая стрик (выходной/отсутствие плана)
            day -= timedelta(days=1)
            if (today_date() - day).days > 60:
                break
            continue
        if all(t["status"] == "done" for t in day_tasks):
            streak += 1
            day -= timedelta(days=1)
        else:
            break
    return streak

async def cmd_mytasks(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id
    tasks = tasks_for_user(tg_id)
    streak = calc_streak(tg_id)
    streak_line = f"🔥 Подряд закрытых дней: {streak}\n\n" if streak >= 2 else ""
    if not tasks:
        await update.effective_message.reply_text(f"{streak_line}✅ Нет активных задач!", parse_mode="HTML")
        return
    sl = {"open":"⬜","in_progress":"🔄","paused":"⏸","overdue":"🔴","cancelled":"🚫"}
    lines = [f"{streak_line}📋 <b>Твои задачи:</b>\n"] if streak_line else ["📋 <b>Твои задачи:</b>\n"]
    for t in tasks:
        dl = t["deadline"]
        dl_fmt = dl[8:]+"."+dl[5:7] if dl else "—"
        src = " <i>(план)</i>" if t.get("source") == "plan" else ""
        tag = f" 🏷{t['channel']}" if t.get("channel") else ""
        lines.append(f"{sl.get(t['status'],'⚪')} <code>{t['task_id']}</code> {t['title']} — до {dl_fmt}{src}{tag}")
    lines.append("\n🔄 /changestatus — сменить статус задачи\n✅ /done — отметить выполненной")
    await reply_long_text(update.effective_message, lines, parse_mode="HTML")

# ── ADMIN KEYBOARD ────────────────────────────────────────────────────────────
def _section_header(label: str):
    """Кнопка-разделитель — визуально выглядит как заголовок блока внутри
    клавиатуры, нажатие ничего не делает (callback 'noop'). Telegram не
    поддерживает настоящие заголовки секций в inline-кнопках, это лучшая
    доступная имитация, чтобы сгруппировать кнопки по смыслу."""
    return InlineKeyboardButton(f"▸ {label}", callback_data="noop")

def build_dept_head_keyboard():
    """
    Базовая панель для dept_head и admin: задачи отдела/команды, статусы,
    отчётность по своей зоне видимости. Без чисто административных функций
    (восстановление задач, экспорт, динамика) — это отдельная клавиатура admin.
    Кнопки сгруппированы по смыслу теми же блоками, что в справочнике:
    Задачи → Отчётность → Управление командой → Обучение.
    """
    return InlineKeyboardMarkup([
        [_section_header("ЗАДАЧИ")],
        [
            InlineKeyboardButton("📋 Задачи сегодня",   callback_data="tasks_today"),
            InlineKeyboardButton("📋 Все активные",     callback_data="show_all_tasks"),
        ],
        [
            InlineKeyboardButton("⚠️ Просроченные",     callback_data="overdue_pick"),
            InlineKeyboardButton("✅ Закрытые задачи",  callback_data="closed_period_pick"),
        ],
        [_section_header("ОТЧЁТНОСТЬ")],
        [
            InlineKeyboardButton("📊 По отделам",       callback_data="summary_depts"),
            InlineKeyboardButton("👤 По сотруднику",    callback_data="summary_person_list"),
        ],
        [
            InlineKeyboardButton("🕐 Рабочий день",     callback_data="workday_summary"),
            InlineKeyboardButton("📅 За неделю",        callback_data="period_week"),
        ],
        [
            InlineKeyboardButton("📅 За месяц",         callback_data="period_month"),
        ],
        [_section_header("УПРАВЛЕНИЕ КОМАНДОЙ")],
        [
            InlineKeyboardButton("🔔 Напомнить о задаче", callback_data="remind_task_pick"),
            InlineKeyboardButton("📨 Запросить статусы",   callback_data="checkstatuses_now"),
        ],
        [_section_header("ОБУЧЕНИЕ")],
        [
            InlineKeyboardButton("📚 Обучение", callback_data="learn_main"),
        ],
    ])

def build_admin_keyboard():
    """Полная панель для admin: всё из dept_head + чисто административные функции,
    собранные в отдельный блок «АДМИНИСТРИРОВАНИЕ» в конце."""
    rows = build_dept_head_keyboard().inline_keyboard
    rows = list(rows) + [
        [_section_header("АДМИНИСТРИРОВАНИЕ")],
        [
            InlineKeyboardButton("📥 Экспорт отдела",      callback_data="export_dept_pick"),
            InlineKeyboardButton("📈 Динамика",            callback_data="dynamics_dept"),
        ],
        [
            InlineKeyboardButton("🔧 Восстановить задачи", callback_data="recover_period_pick"),
        ],
        [
            InlineKeyboardButton("📣 Разослать новое обучение", callback_data="notify_learning_pick"),
        ],
        [
            InlineKeyboardButton("🗑 Удалить пользователя", callback_data="deleteuser_init"),
        ],
    ]
    return InlineKeyboardMarkup(rows)

def build_founder_keyboard():
    """
    Read-only панель для учредителя: видит результаты и задачи всей компании,
    как admin, но без единой управляющей кнопки — не напоминает, не
    восстанавливает задачи, не рассылает обучение, не запрашивает статусы.
    Только просмотр, сгруппированный по тем же блокам ЗАДАЧИ/ОТЧЁТНОСТЬ.
    """
    return InlineKeyboardMarkup([
        [_section_header("ЗАДАЧИ")],
        [
            InlineKeyboardButton("📋 Задачи сегодня",   callback_data="tasks_today"),
            InlineKeyboardButton("📋 Все активные",     callback_data="show_all_tasks"),
        ],
        [
            InlineKeyboardButton("⚠️ Просроченные",     callback_data="overdue_pick"),
            InlineKeyboardButton("✅ Закрытые задачи",  callback_data="closed_period_pick"),
        ],
        [_section_header("ОТЧЁТНОСТЬ")],
        [
            InlineKeyboardButton("📊 По отделам",       callback_data="summary_depts"),
            InlineKeyboardButton("👤 По сотруднику",    callback_data="summary_person_list"),
        ],
        [
            InlineKeyboardButton("🕐 Рабочий день",     callback_data="workday_summary"),
            InlineKeyboardButton("📅 За неделю",        callback_data="period_week"),
        ],
        [
            InlineKeyboardButton("📅 За месяц",         callback_data="period_month"),
        ],
        [
            InlineKeyboardButton("📥 Экспорт",          callback_data="export_dept_pick"),
            InlineKeyboardButton("📈 Динамика",         callback_data="dynamics_dept"),
        ],
    ])

def menu_keyboard_for(tg_id: int):
    """Возвращает правильную клавиатуру по роли: founder видит read-only панель
    по всей компании, dept_head видит базовую по своему отделу, admin видит
    полную с административным блоком."""
    if emp_is_founder(tg_id):
        return build_founder_keyboard()
    return build_admin_keyboard() if emp_is_admin(tg_id) else build_dept_head_keyboard()

def fmt_dl(deadline: str) -> str:
    if not deadline: return "—"
    return deadline[8:]+"."+deadline[5:7]+"."+deadline[:4]

TELEGRAM_MAX_LEN = 4000  # с запасом от реального лимита 4096

def split_lines_to_chunks(lines: list, max_len: int = TELEGRAM_MAX_LEN) -> list:
    """
    Группирует список строк в чанки, каждый не длиннее max_len символов
    (с учётом переводов строк), чтобы не упереться в лимит Telegram на
    длину одного сообщения ("Message is too long").
    """
    chunks = []
    current = []
    current_len = 0
    for line in lines:
        line_len = len(line) + 1  # +1 за \n
        if current_len + line_len > max_len and current:
            chunks.append(current)
            current = []
            current_len = 0
        current.append(line)
        current_len += line_len
    if current:
        chunks.append(current)
    return chunks or [[""]]

async def reply_long_text(message, lines: list, parse_mode="HTML", reply_markup=None):
    """
    Отправляет список строк, разбивая на несколько сообщений если суммарный
    текст превышает лимит Telegram. Кнопки (reply_markup) ставятся только
    под последним сообщением.
    """
    chunks = split_lines_to_chunks(lines)
    for i, chunk in enumerate(chunks):
        is_last = (i == len(chunks) - 1)
        await message.reply_text(
            "\n".join(chunk),
            parse_mode=parse_mode,
            reply_markup=reply_markup if is_last else None
        )

async def send_long_text(bot, chat_id, lines: list, parse_mode="HTML", reply_markup=None):
    """Аналог reply_long_text для рассылок по chat_id (фоновые задания, дайджесты)."""
    chunks = split_lines_to_chunks(lines)
    for i, chunk in enumerate(chunks):
        is_last = (i == len(chunks) - 1)
        await bot.send_message(
            chat_id,
            "\n".join(chunk),
            parse_mode=parse_mode,
            reply_markup=reply_markup if is_last else None
        )

def is_admin_check(query) -> bool:
    """Допускает dept_head и admin — для общих функций панели руководителя."""
    return emp_has_management_access(query.from_user.id)

def is_strict_admin_check(query) -> bool:
    """Только admin — для управляющих действий (рассылка обучения,
    восстановление задач), которых ни dept_head, ни founder не должны
    видеть вообще, даже если узнают callback_data каким-то иным путём."""
    return emp_is_admin(query.from_user.id)

def is_readonly_admin_check(query) -> bool:
    """admin или founder — для read-only функций уровня компании (экспорт,
    динамика), которые не управляют ничем, только показывают данные.
    founder должен видеть их (видит всё как admin), dept_head — нет
    (видит только свой отдел через обычный is_admin_check)."""
    tg_id = query.from_user.id
    return emp_is_admin(tg_id) or emp_is_founder(tg_id)

def dept_filter_for(query) -> str:
    """Пустая строка для admin (видит всё), иначе название отдела."""
    return emp_managed_dept(query.from_user.id)

def tasks_for_date(target_date: str) -> list:
    return [t for t in tasks_all_real() if t.get("deadline","") == target_date and t["status"] not in ("done", "cancelled")]

# ── ADMIN COMMANDS ────────────────────────────────────────────────────────────
async def cmd_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id
    if not emp_has_management_access(tg_id):
        await update.effective_message.reply_text("⛔ Только для руководителей.")
        return
    dept_filter = emp_managed_dept(tg_id)  # "" для admin, название отдела для dept_head
    today = today_date().strftime("%d.%m.%Y")
    employees = emp_employees()
    if dept_filter:
        employees = [e for e in employees if get_dept(e["tg_id"]) == dept_filter]

    plan_ids   = {str(p["tg_id"]) for p in plans_today()}
    report_ids = {str(r["tg_id"]) for r in reports_today()}
    over       = tasks_overdue()
    open_t     = tasks_open()
    if dept_filter:
        over   = [t for t in over   if get_dept(t["assigned_to_id"]) == dept_filter]
        open_t = [t for t in open_t if get_dept(t["assigned_to_id"]) == dept_filter]
        my_plan_ids   = {str(e["tg_id"]) for e in employees} & plan_ids
        my_report_ids = {str(e["tg_id"]) for e in employees} & report_ids
    else:
        my_plan_ids, my_report_ids = plan_ids, report_ids

    dept_label = f" — {dept_filter}" if dept_filter else ""
    text = (
        f"🎛 <b>Панель управления{dept_label} — {today}</b>\n\n"
        f"📋 Планов: {len(my_plan_ids)}/{len(employees)}\n"
        f"📝 День закрыт (EOD): {len(my_report_ids)}/{len(employees)}\n"
        f"✅ Задач активных: {len(open_t)}\n"
        + (f"⚠️ Просроченных: {len(over)}\n" if over else "")
    )

    # У 'menu' нет конкретного действия внутри — сам просмотр панели и есть
    # цель урока, поэтому он засчитывается сразу при открытии. У 'remind' и
    # 'export', наоборот, есть конкретное действие внутри панели (нажать
    # 'Напомнить о задаче' / 'Экспорт') — для них здесь только выставляем
    # флаг (mark_learning_action), а засчитываем уже на реальной точке
    # завершения (cb_remind_task_send / экспорт), иначе шаг засчитывался бы
    # до того, как человек реально попробовал именно ту функцию, про которую
    # урок, и защита от отправки настоящим людям не успевала бы сработать.
    learn_kb = None
    if update.callback_query and update.callback_query.data == "learntry_menu":
        mark_learning_action(tg_id, "menu")
        learn_kb = pop_learning_continue_keyboard(tg_id)
    elif update.callback_query and update.callback_query.data in (
        "learntry_remind", "learntry_export"
    ):
        scenario_id = update.callback_query.data.replace("learntry_", "", 1)
        mark_learning_action(tg_id, scenario_id)

    await update.effective_message.reply_text(
        text, parse_mode="HTML", reply_markup=menu_keyboard_for(tg_id)
    )
    if learn_kb:
        await update.effective_message.reply_text(
            "Панель открыта — можешь нажимать любые кнопки, попробовать их все. "
            "Когда закончишь изучать — продолжи обучение:",
            reply_markup=learn_kb
        )

async def cmd_fixsheets(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Аварийная команда: чинит заголовки во всех листах (принудительно, игнорируя кэш проверки)."""
    if not emp_is_admin(update.effective_user.id):
        await update.effective_message.reply_text("⛔ Только для администраторов."); return
    try:
        _headers_checked.clear()  # сбрасываем, чтобы ensure_headers реально перепроверил все листы
        ensure_headers(emp_sheet(), EMP_H, force=True)
        ensure_headers(plans_sheet(), PH, force=True)
        ensure_headers(reports_sheet(), RH, force=True)
        ensure_headers(tasks_sheet(), TH, force=True)
        invalidate_cache()  # сбрасываем и кэш данных, раз заголовки могли поменяться
        await update.effective_message.reply_text("✅ Заголовки во всех листах исправлены.")
    except Exception as e:
        await update.effective_message.reply_text(f"❌ Ошибка: {e}")


async def cmd_makeadmin(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id
    if not emp_registered(tg_id):
        await update.effective_message.reply_text("Сначала /start"); return
    admins = emp_admins()
    if admins and not emp_is_admin(tg_id):
        await update.effective_message.reply_text("⛔ Уже есть администраторы."); return
    emp_set_admin(tg_id)
    await update.effective_message.reply_text("✅ Ты теперь администратор!")

async def cmd_setadmin(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not emp_is_admin(update.effective_user.id):
        await update.effective_message.reply_text("⛔ Только для руководителей."); return
    if not ctx.args:
        await update.effective_message.reply_text("Укажи: /setadmin @username"); return
    username = ctx.args[0].lstrip("@")
    for e in emp_all():
        if e["username"].lstrip("@").lower() == username.lower():
            emp_set_admin(int(e["tg_id"]))
            await update.effective_message.reply_text(f"✅ {e['full_name']} теперь администратор.")
            return
    await update.effective_message.reply_text(f"❌ @{username} не найден.")

async def cmd_setdepthead(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """
    /setdepthead — диалоговый выбор: отдел → сотрудник из активных, кнопками.
    /setdepthead @username — старый текстовый вариант для тех, кто уже привык.
    """
    if not emp_is_admin(update.effective_user.id):
        await update.effective_message.reply_text("⛔ Только для администраторов."); return

    if ctx.args:
        username = ctx.args[0].lstrip("@")
        for e in emp_all():
            if e["username"].lstrip("@").lower() == username.lower():
                emp_set_role(int(e["tg_id"]), "dept_head")
                dept = get_dept(e["tg_id"])
                await update.effective_message.reply_text(
                    f"✅ {e['full_name']} теперь руководитель подразделения «{dept}».\n"
                    f"Видит планы/отчёты/задачи только своего отдела через /menu"
                )
                return
        await update.effective_message.reply_text(f"❌ @{username} не найден.")
        return

    # диалоговый режим — сначала отдел
    depts_present = sorted(set(get_dept(e["tg_id"]) for e in emp_employees()))
    if not depts_present:
        await update.effective_message.reply_text("Нет зарегистрированных сотрудников."); return
    buttons = []
    row = []
    for d in depts_present:
        row.append(InlineKeyboardButton(d, callback_data=f"sethead_dept_{d}"))
        if len(row) == 2:
            buttons.append(row); row = []
    if row: buttons.append(row)
    await update.effective_message.reply_text(
        "⭐ Назначение руководителя отдела.\nВыбери отдел:",
        reply_markup=InlineKeyboardMarkup(buttons)
    )

async def cb_sethead_dept(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Выбран отдел — показываем активных сотрудников этого отдела."""
    query = update.callback_query; await query.answer()
    if not emp_is_admin(query.from_user.id):
        await query.answer("⛔ Только для администраторов.", show_alert=True); return
    dept = query.data.replace("sethead_dept_", "", 1)
    emps = [e for e in emp_employees() if get_dept(e["tg_id"]) == dept]
    if not emps:
        await query.message.reply_text(f"В отделе «{dept}» нет сотрудников."); return
    buttons = []
    row = []
    for e in emps:
        role_mark = " ⭐" if e["role"] == "dept_head" else ""
        row.append(InlineKeyboardButton(e["full_name"]+role_mark, callback_data=f"sethead_emp_{e['tg_id']}"))
        if len(row) == 2:
            buttons.append(row); row = []
    if row: buttons.append(row)
    await query.message.reply_text(
        f"Сотрудники отдела «{dept}»:\n(⭐ — уже руководитель)",
        reply_markup=InlineKeyboardMarkup(buttons)
    )

async def cb_sethead_emp(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Выбран сотрудник — назначаем руководителем отдела."""
    query = update.callback_query; await query.answer()
    if not emp_is_admin(query.from_user.id):
        await query.answer("⛔ Только для администраторов.", show_alert=True); return
    target_tg_id = query.data.replace("sethead_emp_", "", 1)
    emp_set_role(int(target_tg_id), "dept_head")
    emp = emp_by_id(int(target_tg_id))
    dept = get_dept(int(target_tg_id))
    await query.edit_message_text(
        f"✅ {emp['full_name']} теперь руководитель подразделения «{dept}».\n"
        f"Видит планы/отчёты/задачи только своего отдела через /menu"
    )
    try:
        await ctx.bot.send_message(
            int(target_tg_id),
            f"⭐ Тебя назначили руководителем отдела «{dept}»!\n"
            f"Теперь доступна команда /menu — сводки и задачи по твоему отделу."
        )
    except Exception: pass

async def cmd_setfounder(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """
    /setfounder @username — назначить учредителем. Человек должен сначала
    сам написать /start и зарегистрироваться (выбрать любой отдел — это
    значения не имеет, роль founder его перекроет), затем admin назначает
    эту команду. После назначения у учредителя спрашивается, хочет ли он
    самостоятельно заходить в /menu, или получать автоматическую ежедневную
    сводку, или и то и другое.
    """
    if not emp_is_admin(update.effective_user.id):
        await update.effective_message.reply_text("⛔ Только для администраторов."); return
    if not ctx.args:
        await update.effective_message.reply_text("Укажи: /setfounder @username"); return
    username = ctx.args[0].lstrip("@")
    for e in emp_all():
        if e["username"].lstrip("@").lower() == username.lower():
            emp_set_role(int(e["tg_id"]), "founder")
            await update.effective_message.reply_text(
                f"✅ {e['full_name']} теперь учредитель.\n"
                f"Видит результаты и задачи всей компании в режиме просмотра."
            )
            try:
                keyboard = InlineKeyboardMarkup([
                    [InlineKeyboardButton("📊 Заходить самостоятельно", callback_data="founder_pref_manual")],
                    [InlineKeyboardButton("📬 Получать ежедневную сводку", callback_data="founder_pref_digest")],
                    [InlineKeyboardButton("📊📬 И то, и другое", callback_data="founder_pref_both")],
                ])
                await ctx.bot.send_message(
                    int(e["tg_id"]),
                    "👋 Тебя назначили учредителем!\n\n"
                    "Ты видишь результаты работы и задачи всей компании в "
                    "режиме просмотра — без необходимости писать планы, отчёты "
                    "или закрывать день.\n\n"
                    "Как тебе удобнее получать информацию?",
                    reply_markup=keyboard
                )
            except Exception as ex:
                logger.warning(f"setfounder notify {e['tg_id']}: {ex}")
            return
    await update.effective_message.reply_text(f"❌ @{username} не найден.")

async def cb_founder_pref(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Учредитель выбрал предпочтение по уведомлениям."""
    query = update.callback_query; await query.answer("Записано")
    pref = query.data.replace("founder_pref_", "", 1)  # manual|digest|both
    tg_id = query.from_user.id
    emp_set_founder_digest_pref(tg_id, pref)
    labels = {
        "manual": "Будешь заходить в /menu самостоятельно, когда захочешь.",
        "digest": "Будешь получать ежедневную сводку автоматически в 19:35.",
        "both": "Будешь и получать ежедневную сводку, и можешь заходить в /menu в любой момент.",
    }
    await query.edit_message_text(f"✅ Готово!\n{labels.get(pref, '')}")

async def cmd_setdept(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """
    /setdept — диалоговый выбор: сотрудник из списка → новый отдел, кнопками.
    /setdept @username — старый текстовый вариант для тех, кто уже привык.
    """
    if not emp_is_admin(update.effective_user.id):
        await update.effective_message.reply_text("⛔ Только для администраторов."); return

    if ctx.args:
        username = ctx.args[0].lstrip("@")
        for e in emp_all():
            if e["username"].lstrip("@").lower() == username.lower():
                await update.effective_message.reply_text(
                    f"Выбери отдел для {e['full_name']}:",
                    reply_markup=dept_keyboard("admset_dept_", str(e["tg_id"]))
                )
                return
        await update.effective_message.reply_text(f"❌ @{username} не найден.")
        return

    # диалоговый режим — список всех сотрудников по именам
    emps = emp_employees()
    if not emps:
        await update.effective_message.reply_text("Нет зарегистрированных сотрудников."); return
    buttons = []
    row = []
    for e in emps:
        row.append(InlineKeyboardButton(e["full_name"], callback_data=f"setdeptemp_{e['tg_id']}"))
        if len(row) == 2:
            buttons.append(row); row = []
    if row: buttons.append(row)
    await update.effective_message.reply_text(
        "Кому сменить отдел? Выбери сотрудника:",
        reply_markup=InlineKeyboardMarkup(buttons)
    )

async def cb_setdeptemp_pick(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Выбран сотрудник — показываем кнопки отделов."""
    query = update.callback_query; await query.answer()
    if not emp_is_admin(query.from_user.id):
        await query.answer("⛔ Только для администраторов.", show_alert=True); return
    target_tg_id = query.data.replace("setdeptemp_", "", 1)
    emp = emp_by_id(int(target_tg_id))
    if not emp:
        await query.message.reply_text("Сотрудник не найден."); return
    await query.message.reply_text(
        f"Выбери отдел для {emp['full_name']}:",
        reply_markup=dept_keyboard("admset_dept_", target_tg_id)
    )

async def cmd_setshift(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """
    /setshift — диалоговый выбор: сотрудник из списка → график (стандартный
    10:00–19:00 / сдвинутый 11:00–22:00), кнопками. Назначает, во сколько
    конкретному сотруднику бот будет спрашивать старт/конец рабочего дня.
    """
    if not emp_is_admin(update.effective_user.id):
        await update.effective_message.reply_text("⛔ Только для администраторов."); return
    emps = emp_employees()
    if not emps:
        await update.effective_message.reply_text("Нет зарегистрированных сотрудников."); return
    buttons = []
    row = []
    for e in emps:
        shift = emp_shift(int(e["tg_id"]))
        mark = " 🌙" if shift["start_hour"] == 11 else ""
        row.append(InlineKeyboardButton(e["full_name"]+mark, callback_data=f"setshiftemp_{e['tg_id']}"))
        if len(row) == 2:
            buttons.append(row); row = []
    if row: buttons.append(row)
    await update.effective_message.reply_text(
        "Кому изменить график? (🌙 — уже сдвинутый 11:00–22:00)",
        reply_markup=InlineKeyboardMarkup(buttons)
    )

async def cb_setshiftemp_pick(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Выбран сотрудник — показываем кнопки графика."""
    query = update.callback_query; await query.answer()
    if not emp_is_admin(query.from_user.id):
        await query.answer("⛔ Только для администраторов.", show_alert=True); return
    target_tg_id = query.data.replace("setshiftemp_", "", 1)
    emp = emp_by_id(int(target_tg_id))
    if not emp:
        await query.message.reply_text("Сотрудник не найден."); return
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("☀️ Стандартный 10:00–19:00", callback_data=f"setshiftval_standard_{target_tg_id}")],
        [InlineKeyboardButton("🌙 Сдвинутый 11:00–22:00", callback_data=f"setshiftval_shifted_{target_tg_id}")],
    ])
    await query.message.reply_text(f"Выбери график для {emp['full_name']}:", reply_markup=keyboard)

async def cb_setshiftval_pick(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Применяем выбранный график."""
    query = update.callback_query; await query.answer("✅ Сохранено")
    if not emp_is_admin(query.from_user.id):
        await query.answer("⛔ Только для администраторов.", show_alert=True); return
    rest = query.data.replace("setshiftval_", "", 1)
    code, target_tg_id = rest.split("_", 1)
    emp_set_shift_schedule(int(target_tg_id), code)
    emp = emp_by_id(int(target_tg_id))
    label = "Стандартный (10:00–19:00)" if code == "standard" else "Сдвинутый (11:00–22:00)"
    await query.edit_message_text(f"✅ {emp['full_name']}: график «{label}»")

async def cmd_deleteuser(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """
    /deleteuser — диалоговый выбор сотрудника кнопками, с явным подтверждением
    перед удалением (деструктивная операция — нужна защита от случайного
    нажатия). История задач/планов/отчётов человека не удаляется, остаётся
    для отчётности.
    """
    if update.callback_query:
        await update.callback_query.answer()
    if not emp_is_admin(update.effective_user.id):
        await update.effective_message.reply_text("⛔ Только для администраторов."); return
    emps = emp_employees()
    if not emps:
        await update.effective_message.reply_text("Нет зарегистрированных сотрудников."); return
    buttons = []
    row = []
    for e in emps:
        row.append(InlineKeyboardButton(e["full_name"], callback_data=f"deluser_pick_{e['tg_id']}"))
        if len(row) == 2:
            buttons.append(row); row = []
    if row: buttons.append(row)
    await update.effective_message.reply_text(
        "Кого удалить из команды?",
        reply_markup=InlineKeyboardMarkup(buttons)
    )

async def cb_deluser_pick(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Выбран сотрудник — запрашиваем явное подтверждение перед удалением."""
    query = update.callback_query; await query.answer()
    if not emp_is_admin(query.from_user.id):
        await query.answer("⛔ Только для администраторов.", show_alert=True); return
    target_tg_id = query.data.replace("deluser_pick_", "", 1)
    emp = emp_by_id(int(target_tg_id))
    if not emp:
        await query.message.reply_text("Сотрудник не найден."); return
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("❌ Да, удалить", callback_data=f"deluser_confirm_{target_tg_id}")],
        [InlineKeyboardButton("◀️ Отмена", callback_data="back_main")],
    ])
    await query.message.reply_text(
        f"⚠️ Удалить {emp['full_name']} из команды?\n\n"
        f"Человек больше не сможет пользоваться ботом без повторной регистрации. "
        f"Его прошлые задачи, планы и отчёты останутся в таблицах для истории.",
        reply_markup=keyboard
    )

async def cb_deluser_confirm(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Подтверждено — удаляем сотрудника."""
    query = update.callback_query; await query.answer("Удалено")
    if not emp_is_admin(query.from_user.id):
        await query.answer("⛔ Только для администраторов.", show_alert=True); return
    target_tg_id = query.data.replace("deluser_confirm_", "", 1)
    emp = emp_by_id(int(target_tg_id))
    name = emp["full_name"] if emp else target_tg_id
    if emp_delete(int(target_tg_id)):
        await query.edit_message_text(f"✅ {name} удалён(а) из команды.")
    else:
        await query.edit_message_text("❌ Не удалось удалить — сотрудник не найден.")

async def cmd_fixname(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """/fixname @username Имя Фамилия — вручную исправить порядок имени/фамилии."""
    if not emp_is_admin(update.effective_user.id):
        await update.effective_message.reply_text("⛔ Только для администраторов."); return
    if len(ctx.args) < 2:
        await update.effective_message.reply_text("Укажи: /fixname @username Имя Фамилия"); return
    username = ctx.args[0].lstrip("@")
    new_name = " ".join(ctx.args[1:])
    for e in emp_all():
        if e["username"].lstrip("@").lower() == username.lower():
            emp_set_full_name(int(e["tg_id"]), new_name)
            await update.effective_message.reply_text(f"✅ Имя обновлено: {new_name}")
            return
    await update.effective_message.reply_text(f"❌ @{username} не найден.")

async def cmd_fixallnames(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """/fixallnames — прогоняет нормализацию имён по всем сотрудникам сразу."""
    if not emp_is_admin(update.effective_user.id):
        await update.effective_message.reply_text("⛔ Только для администраторов."); return
    changed = []
    for e in emp_all():
        old_name = e["full_name"]
        new_name = normalize_full_name(old_name)
        if new_name != old_name:
            emp_set_full_name(int(e["tg_id"]), new_name)
            changed.append(f"{old_name} → {new_name}")
    if changed:
        await update.effective_message.reply_text(
            "✅ Имена нормализованы:\n" + "\n".join(changed)
        )
    else:
        await update.effective_message.reply_text("Все имена уже в формате «Имя Фамилия».")

async def cmd_team(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id
    if not emp_has_management_access(tg_id):
        await update.effective_message.reply_text("⛔ Только для руководителей."); return
    dept_filter = emp_managed_dept(tg_id)
    employees = emp_employees()
    if dept_filter:
        employees = [e for e in employees if get_dept(e["tg_id"]) == dept_filter]
    plan_ids  = {str(p["tg_id"]) for p in plans_today()}
    report_ids = {str(r["tg_id"]) for r in reports_today()}
    dept_label = f" — {dept_filter}" if dept_filter else ""
    lines = [f"📊 <b>Команда{dept_label} {datetime.now(TZ).strftime('%d.%m.%Y')}:</b>\n"]
    for e in employees:
        tid = str(e["tg_id"])
        lines.append(
            f"{'✅' if tid in report_ids else '❌'} <b>{e['full_name']}</b>  "
            f"план {'✅' if tid in plan_ids else '❌'}  EOD {'✅' if tid in report_ids else '—'}"
        )
    await reply_long_text(update.effective_message, lines, parse_mode="HTML")

async def cmd_workday(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """/workday — текстовая версия сводки 'кто во сколько начал/закончил день',
    альтернатива кнопке 🕐 Рабочий день для тех, кто предпочитает команды."""
    tg_id = update.effective_user.id
    if not emp_has_management_access(tg_id):
        await update.effective_message.reply_text("⛔ Только для руководителей."); return
    dept_filter = emp_managed_dept(tg_id)
    employees = emp_employees()
    if dept_filter:
        employees = [e for e in employees if get_dept(e["tg_id"]) == dept_filter]

    today_rows = {str(r["tg_id"]): r for r in workday_today()}
    today_fmt = today_date().strftime("%d.%m.%Y")
    dept_label = f" — {dept_filter}" if dept_filter else ""

    closed, in_progress, not_started = [], [], []
    for e in employees:
        tid = str(e["tg_id"])
        row = today_rows.get(tid)
        if not row or not row.get("start_at"):
            not_started.append(e)
        elif not row.get("end_at"):
            in_progress.append((e, row))
        else:
            closed.append((e, row))

    def fmt_time(ts: str) -> str:
        return ts.split(" ")[-1] if ts else "—"

    lines = [f"🕐 <b>Рабочий день{dept_label} — {today_fmt}</b>\n"]
    if in_progress:
        lines.append(f"🟢 <b>Сейчас на смене ({len(in_progress)}):</b>")
        for e, row in sorted(in_progress, key=lambda x: x[1].get("start_at", "")):
            lines.append(f"  {e['full_name']}: начал {fmt_time(row['start_at'])}")
        lines.append("")
    if closed:
        lines.append(f"⚪ <b>День закрыт ({len(closed)}):</b>")
        for e, row in sorted(closed, key=lambda x: x[1].get("start_at", "")):
            lines.append(f"  {e['full_name']}: {fmt_time(row['start_at'])}–{fmt_time(row['end_at'])}")
        lines.append("")
    if not_started:
        lines.append(f"🔴 <b>Не начали день ({len(not_started)}):</b>")
        for e in not_started:
            lines.append(f"  {e['full_name']}")
    if not employees:
        lines.append("Нет сотрудников для отображения.")

    await reply_long_text(update.effective_message, lines, parse_mode="HTML")

async def cmd_tasks_all(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id
    if not emp_has_management_access(tg_id):
        await update.effective_message.reply_text("⛔ Только для руководителей."); return
    dept_filter = emp_managed_dept(tg_id)
    open_t = tasks_open()
    over_t = tasks_overdue()
    if dept_filter:
        open_t = [t for t in open_t if get_dept(t["assigned_to_id"]) == dept_filter]
        over_t = [t for t in over_t if get_dept(t["assigned_to_id"]) == dept_filter]
    dept_label = f" — {dept_filter}" if dept_filter else ""
    lines = [f"📋 <b>Активные задачи{dept_label}:</b>\n"]
    if not open_t:
        lines.append("Нет активных задач.")
    by_dept: dict = {}
    for t in open_t:
        dept = get_dept(t["assigned_to_id"])
        by_dept.setdefault(dept, []).append(t)
    for dept, dt_list in sorted(by_dept.items()):
        lines.append(f"<b>— {dept} —</b>")
        for t in dt_list:
            icon = "🔴" if t in over_t else "🔵"
            src = " <i>(план)</i>" if t.get("source") == "plan" else ""
            lines.append(f"  {icon} <code>{t['task_id']}</code> {t['assigned_to_name']}: {t['title']}{src}")
        lines.append("")
    await reply_long_text(update.effective_message, lines, parse_mode="HTML")

async def cmd_checkstatuses(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Запросить статусы по текущим задачам у всей команды (или своего отдела) прямо сейчас."""
    tg_id = update.effective_user.id
    if not emp_has_management_access(tg_id):
        await update.effective_message.reply_text("⛔ Только для руководителей."); return
    dept_filter = emp_managed_dept(tg_id)
    employees = emp_employees()
    if dept_filter:
        employees = [e for e in employees if get_dept(e["tg_id"]) == dept_filter]

    if not employees:
        await update.effective_message.reply_text("Нет сотрудников для запроса."); return

    sent = 0
    for e in employees:
        try:
            await start_eod_flow(ctx.bot, int(e["tg_id"]))
            sent += 1
        except Exception as ex:
            logger.warning(f"checkstatuses {e['tg_id']}: {ex}")

    dept_label = f" отдела «{dept_filter}»" if dept_filter else " команды"
    await update.effective_message.reply_text(
        f"📨 Запрос статусов отправлен {sent} сотрудникам{dept_label}."
    )

def recover_tasks_from_plans(date_from: str, date_to: str, dept_filter: str = ""):
    """
    Сравнивает планы за период с уже существующими задачами (source=plan)
    и досоздаёт недостающие. Возвращает (recovered_count, by_person, affected_dates).
    """
    ws_plans = sheet("plans"); ensure_headers(ws_plans, PH)
    all_plans = [p for p in safe_records(ws_plans, PH)
                 if date_from <= p.get("date","") <= date_to]
    if dept_filter:
        all_plans = [p for p in all_plans if get_dept(p["tg_id"]) == dept_filter]

    all_t = tasks_all_real()
    # ключ существующей задачи из плана: (tg_id, date, title) — чтобы не дублировать
    existing_keys = set()
    for t in all_t:
        if t.get("source") == "plan":
            key = (str(t["assigned_to_id"]), t.get("deadline",""), t.get("title","").strip())
            existing_keys.add(key)

    recovered = 0
    by_person: dict = {}
    affected_dates = set()

    # берём последнюю запись плана на каждую (сотрудник, дата) — план может быть переписан несколько раз за день
    latest_plan: dict = {}
    for p in all_plans:
        key = (str(p["tg_id"]), p["date"])
        # submitted_at растёт по времени добавления строк, последняя запись с этим ключом — самая свежая
        latest_plan[key] = p

    for (tg_id, plan_date), p in latest_plan.items():
        items = parse_plan_items(p["plan_text"])
        for item, item_deadline in items:
            key = (tg_id, item_deadline, item.strip())
            if key in existing_keys:
                continue
            task_create(int(tg_id), p["full_name"], int(tg_id), p["full_name"],
                       item, item_deadline, source="plan")
            existing_keys.add(key)
            recovered += 1
            by_person[p["full_name"]] = by_person.get(p["full_name"], 0) + 1
            affected_dates.add(item_deadline)

    return recovered, by_person, sorted(affected_dates)

async def cmd_recovertasks(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """/recovertasks [N] — восстановить задачи из планов за последние N дней (по умолчанию 7).
    Только для admin — dept_head не должен иметь доступа к этой функции."""
    tg_id = update.effective_user.id
    if not emp_is_admin(tg_id):
        await update.effective_message.reply_text("⛔ Только для администраторов."); return
    dept_filter = emp_managed_dept(tg_id)

    days_back = 7
    if ctx.args and ctx.args[0].isdigit():
        days_back = int(ctx.args[0])

    date_to = today_date().strftime("%Y-%m-%d")
    date_from = (today_date() - timedelta(days=days_back)).strftime("%Y-%m-%d")

    await update.effective_message.reply_text(
        f"🔧 Сканирую планы с {date_from} по {date_to}..."
    )

    recovered, by_person, affected_dates = recover_tasks_from_plans(date_from, date_to, dept_filter)

    if recovered == 0:
        await update.effective_message.reply_text("✅ Расхождений не найдено — все задачи из планов на месте.")
        return

    lines = [f"🔧 <b>Восстановлено задач: {recovered}</b>\n"]
    for person, cnt in sorted(by_person.items()):
        lines.append(f"  {person}: +{cnt}")
    lines.append(f"\nЗатронутые даты: {', '.join(affected_dates)}")
    await reply_long_text(update.effective_message, lines, parse_mode="HTML")

    # отправляем уведомление всем администраторам (в т.ч. инициатору, если он не получил выше)
    notify = (
        f"🔧 <b>Восстановление задач выполнено</b>\n"
        f"Инициатор: {emp_by_id(tg_id)['full_name']}\n"
        f"Период: {date_from} — {date_to}\n"
        f"Восстановлено задач: {recovered}\n"
    )
    for adm in emp_admins():
        if int(adm["tg_id"]) == tg_id:
            continue
        try:
            await ctx.bot.send_message(int(adm["tg_id"]), notify, parse_mode="HTML")
        except Exception as ex:
            logger.warning(f"recover notify admin {adm['tg_id']}: {ex}")

    # сразу предлагаем актуализировать статусы у всех затронутых
    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("📨 Запросить статусы сейчас", callback_data="checkstatuses_now")
    ]])
    await update.effective_message.reply_text(
        "Хочешь сразу запросить актуальные статусы по восстановленным задачам?",
        reply_markup=keyboard
    )

async def cb_checkstatuses_now(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Кнопка-дублёр /checkstatuses, вызываемая после восстановления задач."""
    query = update.callback_query
    if not emp_has_management_access(query.from_user.id):
        await query.answer("⛔ Только для руководителей.", show_alert=True); return
    await query.answer("Запускаю опрос...")
    dept_filter = emp_managed_dept(query.from_user.id)
    employees = emp_employees()
    if dept_filter:
        employees = [e for e in employees if get_dept(e["tg_id"]) == dept_filter]

    sent = 0
    for e in employees:
        try:
            await start_eod_flow(ctx.bot, int(e["tg_id"]))
            sent += 1
        except Exception as ex:
            logger.warning(f"checkstatuses_now {e['tg_id']}: {ex}")

    dept_label = f" отдела «{dept_filter}»" if dept_filter else " команды"
    await query.message.reply_text(f"📨 Запрос статусов отправлен {sent} сотрудникам{dept_label}.")

# ── ADMIN CALLBACKS ───────────────────────────────────────────────────────────
async def cb_learn_main(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """
    Главное меню обучения — строгая последовательность по роли. Показывает
    прогресс-бар (✅ пройдено, ▶️ текущий доступный шаг, 🔒 заблокировано
    до прохождения предыдущих). Кликабельна только кнопка текущего шага.
    Если всё пройдено — показывает итог и предлагает новые сценарии, если
    появились после завершения (без повтора всей цепочки).
    """
    query = update.callback_query; await query.answer()
    tg_id = query.from_user.id
    u = emp_by_id(tg_id)
    if not u:
        await query.message.reply_text("Сначала /start"); return

    role = u["role"]
    sequence = learn_scenarios_for_role(role)
    current_idx = learning_current_step_index(tg_id, role)

    if current_idx >= len(sequence):
        # обучение полностью пройдено — проверяем, нет ли новых сценариев
        new_ones = learning_new_scenarios_since_completion(tg_id, role)
        lines = [f"🎉 <b>Обучение завершено!</b> Пройдено {len(sequence)}/{len(sequence)} шагов."]
        if new_ones:
            lines.append("\nПоявились новые функции бота:")
            buttons = [[InlineKeyboardButton(s["label"], callback_data=f"learn_{s['id']}")] for s in new_ones]
            await query.message.reply_text("\n".join(lines), parse_mode="HTML",
                                            reply_markup=InlineKeyboardMarkup(buttons))
        else:
            await query.message.reply_text("\n".join(lines), parse_mode="HTML")
        return

    lines = ["📚 <b>Твой прогресс обучения</b>"]
    lines.append(f"Пройдено: {current_idx} из {len(sequence)}\n")
    for i, s in enumerate(sequence):
        # убираем эмодзи самого шага из общего списка, чтобы не было
        # двух значков подряд (статус + эмодзи названия) — оставляем
        # только текстовое название без иконки-приставки
        label_text = s["label"].split(" ", 1)[-1] if " " in s["label"] else s["label"]
        if i < current_idx:
            lines.append(f"✅  {i+1}. {label_text}")
        elif i == current_idx:
            lines.append(f"▶️  {i+1}. {label_text}")
        else:
            lines.append(f"⚪  {i+1}. {label_text}")

    current = sequence[current_idx]
    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton(f"▶️ Начать: {current['label']}", callback_data=f"learn_{current['id']}"),
    ]])
    await query.message.reply_text("\n".join(lines), parse_mode="HTML", reply_markup=keyboard)

async def cb_learn_scenario(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """
    Открывает карточку текущего доступного шага с кнопкой 'Попробовать сейчас'.
    ВАЖНО: просмотр карточки сам по себе НЕ засчитывается как прохождение —
    шаг отмечается пройденным только когда практическое действие реально
    завершено (см. pop_learning_continue_keyboard и места его вызова).
    Кнопки 'назад' нет — порядок строгий, нельзя перепрыгнуть или вернуться
    к выбору другого шага.
    """
    query = update.callback_query; await query.answer()
    scenario_id = query.data.replace("learn_", "", 1)
    scenario = next((s for s in LEARN_SCENARIOS if s["id"] == scenario_id), None)
    if not scenario:
        await query.message.reply_text("Сценарий не найден."); return

    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton(f"▶️ Попробовать {scenario['try_cmd']}", callback_data=f"learntry_{scenario['id']}")],
    ])
    await query.message.reply_text(scenario["text"], parse_mode="HTML", reply_markup=keyboard)

async def cb_learn_try(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """
    Кнопка 'Попробовать сейчас' — запускает соответствующую команду напрямую.
    /plan и /task зарегистрированы в ConversationHandler с дополнительной
    callback-точкой входа (см. main()), поэтому прямой вызов корректно
    регистрирует состояние диалога — следующее сообщение пользователя попадёт
    в recv_plan/recv_task как положено.
    """
    await update.callback_query.answer()  # сразу гасим "загрузку" на кнопке;
    # повторный answer() внутри cmd_plan/cmd_task безопасен — Telegram просто
    # проигнорирует второй вызов для того же callback_query.

    scenario_id = update.callback_query.data.replace("learntry_", "", 1)
    scenario = next((s for s in LEARN_SCENARIOS if s["id"] == scenario_id), None)
    if not scenario:
        return
    cmd = scenario["try_cmd"].lstrip("/")

    mark_learning_action(update.callback_query.from_user.id, scenario_id)

    dispatch = {
        "startday": cmd_startday, "plan": cmd_plan, "eod": cmd_eod, "endday": cmd_endday,
        "task": cmd_task, "tag": cmd_tag, "changestatus": cmd_changestatus, "menu": cmd_menu,
    }
    handler = dispatch.get(cmd)
    if handler:
        await handler(update, ctx)

async def cb_noop(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Кнопки-заголовки разделов внутри /menu — чисто визуальная группировка,
    нажатие не делает ничего, кроме обязательного answer() (иначе кнопка
    у пользователя зависнет в состоянии 'загрузка')."""
    await update.callback_query.answer()

async def cb_back_main(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    await query.message.reply_text("Выбери действие:", reply_markup=menu_keyboard_for(query.from_user.id))

async def cb_overdue_pick(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Список просроченных задач — каждая с кнопкой выбора действия."""
    query = update.callback_query; await query.answer()
    if not is_admin_check(query): return
    dept_filter = dept_filter_for(query)
    over = tasks_overdue()
    if dept_filter:
        over = [t for t in over if get_dept(t["assigned_to_id"]) == dept_filter]
    if not over:
        await query.message.reply_text("✅ Нет просроченных задач!"); return

    dept_label = f" — {dept_filter}" if dept_filter else ""
    await query.message.reply_text(f"⚠️ <b>Просроченные задачи{dept_label} ({len(over)}):</b>", parse_mode="HTML")
    for t in over[:20]:
        text = (f"🔴 <b>{t['title']}</b>\n"
                f"Исполнитель: {t['assigned_to_name']}\n"
                f"Срок был: {fmt_dl(t['deadline'])}")
        keyboard = InlineKeyboardMarkup([[
            InlineKeyboardButton("✅ Закрыть", callback_data=f"overduedone_{t['task_id']}"),
            InlineKeyboardButton("📅 Перенести", callback_data=f"overduemove_{t['task_id']}"),
        ]])
        await query.message.reply_text(text, parse_mode="HTML", reply_markup=keyboard)
    if len(over) > 20:
        await query.message.reply_text(f"… и ещё {len(over)-20}. Используй 📥 Экспорт отдела для полного списка.")

    notify_kb = InlineKeyboardMarkup([[
        InlineKeyboardButton("📨 Попросить команду актуализировать", callback_data="overdue_notify_team"),
    ]])
    await query.message.reply_text(
        "Хочешь, чтобы каждый сам обновил статус своих просроченных задач?",
        reply_markup=notify_kb
    )

async def cb_overdue_done(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Закрыть просроченную задачу прямо из списка."""
    query = update.callback_query; await query.answer("✅ Закрыто")
    if not is_admin_check(query): return
    tid = query.data.replace("overduedone_", "", 1)
    task_update_status(tid, "done")
    task = task_by_id(tid)
    await query.edit_message_text(f"✅ Закрыто:\n{task['title'] if task else tid}", parse_mode="HTML")

async def cb_overdue_move_pick(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Выбор нового срока для просроченной задачи."""
    query = update.callback_query; await query.answer()
    if not is_admin_check(query): return
    tid = query.data.replace("overduemove_", "", 1)
    today = today_date()
    options = [
        ("Сегодня", today),
        ("Завтра", today + timedelta(days=1)),
        ("+3 дня", today + timedelta(days=3)),
        ("+неделя", today + timedelta(days=7)),
    ]
    buttons = [[InlineKeyboardButton(label, callback_data=f"overduesetdl_{tid}_{d.strftime('%Y-%m-%d')}")]
               for label, d in options]
    await query.message.reply_text("На какой срок перенести?", reply_markup=InlineKeyboardMarkup(buttons))

async def cb_overdue_set_deadline(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Применяет новый срок к просроченной задаче."""
    query = update.callback_query; await query.answer("📅 Перенесено")
    if not is_admin_check(query): return
    rest = query.data.replace("overduesetdl_", "", 1)
    tid, new_deadline = rest.rsplit("_", 1)
    task_update_deadline(tid, new_deadline)
    task = task_by_id(tid)
    await query.edit_message_text(
        f"📅 Перенесено на {fmt_dl(new_deadline)}:\n{task['title'] if task else tid}",
        parse_mode="HTML"
    )

async def cb_overdue_notify_team(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """
    Массовая актуализация: каждому сотруднику с просроченными задачами
    приходит личное сообщение со списком ЕГО просроченных задач, каждая —
    с кнопкой выбора актуального статуса. Переиспользует тот же безопасный
    путь смены статуса, что /changestatus (chstatustask_), где уже проверено,
    что менять статус может только исполнитель или admin — никаких новых
    прав создавать не нужно.
    """
    query = update.callback_query; await query.answer("📨 Рассылаю...")
    if not is_admin_check(query): return
    dept_filter = dept_filter_for(query)
    over = tasks_overdue()
    if dept_filter:
        over = [t for t in over if get_dept(t["assigned_to_id"]) == dept_filter]
    if not over:
        await query.message.reply_text("✅ Нет просроченных задач — рассылать некому."); return

    by_assignee = {}
    for t in over:
        by_assignee.setdefault(t["assigned_to_id"], []).append(t)

    sent = 0
    for assignee_id, tasks in by_assignee.items():
        try:
            buttons = []
            for t in tasks[:15]:
                label = f"{t['title'][:38]}" + ("…" if len(t['title']) > 38 else "") + f" (до {fmt_dl(t['deadline'])})"
                buttons.append([InlineKeyboardButton(label, callback_data=f"chstatustask_{t['task_id']}")])
            overflow = f"\n…и ещё {len(tasks)-15}, используй /changestatus" if len(tasks) > 15 else ""
            await ctx.bot.send_message(
                int(assignee_id),
                f"⚠️ <b>Актуализируй статус по просроченным задачам ({len(tasks)}):</b>\n"
                f"Запросил(а): {emp_by_id(query.from_user.id)['full_name']}{overflow}",
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup(buttons)
            )
            sent += 1
        except Exception as ex:
            logger.warning(f"overdue_notify_team {assignee_id}: {ex}")

    await query.message.reply_text(f"📨 Разослано {sent} сотрудникам ({len(over)} задач всего).")

async def cb_notify_learning_pick(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Admin видит, у каких сценариев есть непройденные версии в команде,
    и может разослать конкретный сценарий всем сотрудникам сразу."""
    query = update.callback_query; await query.answer()
    if not is_strict_admin_check(query): return

    employees = emp_employees()
    buttons = []
    for s in LEARN_SCENARIOS:
        pending_count = 0
        for e in employees:
            role = e["role"] if e["role"] in ("employee", "dept_head", "admin") else "employee"
            if s not in learn_scenarios_for_role(role):
                continue
            completed = learning_completed_versions(int(e["tg_id"]))
            if completed.get(s["id"], 0) < s["version"]:
                pending_count += 1
        if pending_count > 0:
            buttons.append([InlineKeyboardButton(
                f"{s['label']} — не прошли {pending_count}",
                callback_data=f"notifylearn_{s['id']}"
            )])

    if not buttons:
        await query.message.reply_text("✅ Все сотрудники прошли все актуальные сценарии обучения."); return

    await query.message.reply_text(
        "📣 Выбери сценарий, чтобы разослать его всем, кто ещё не прошёл:",
        reply_markup=InlineKeyboardMarkup(buttons)
    )

async def cb_notify_learning_send(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Рассылает карточку сценария всем, у кого она не пройдена в актуальной версии."""
    query = update.callback_query; await query.answer("Рассылаю...")
    if not is_strict_admin_check(query): return
    scenario_id = query.data.replace("notifylearn_", "", 1)
    scenario = next((s for s in LEARN_SCENARIOS if s["id"] == scenario_id), None)
    if not scenario:
        await query.message.reply_text("Сценарий не найден."); return

    employees = emp_employees()
    sent = 0
    for e in employees:
        tg_id = int(e["tg_id"])
        completed = learning_completed_versions(tg_id)
        if completed.get(scenario["id"], 0) >= scenario["version"]:
            continue  # уже прошёл актуальную версию
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton(f"▶️ Попробовать {scenario['try_cmd']}", callback_data=f"learntry_{scenario['id']}")],
            [InlineKeyboardButton("📚 Всё обучение", callback_data="learn_main")],
        ])
        try:
            await ctx.bot.send_message(
                tg_id,
                f"🆕 <b>В боте новая или обновлённая функция!</b>\n\n{scenario['text']}",
                parse_mode="HTML", reply_markup=keyboard
            )
            sent += 1
        except Exception as ex:
            logger.warning(f"notify_learning_send {tg_id}: {ex}")

    await query.message.reply_text(f"📣 Разослано {sent} сотрудникам.")

async def cmd_learningreport(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """
    Текстовый обзор прогресса обучения по всей команде — кто что прошёл,
    без необходимости открывать саму Google-таблицу learning_progress.
    Доступно admin и dept_head (dept_head видит только свой отдел).
    """
    tg_id = update.effective_user.id
    if not emp_has_management_access(tg_id):
        await update.effective_message.reply_text("⛔ Только для руководителей."); return
    dept_filter = emp_managed_dept(tg_id)

    employees = emp_employees()
    if dept_filter:
        employees = [e for e in employees if get_dept(e["tg_id"]) == dept_filter]

    dept_label = f" — {dept_filter}" if dept_filter else ""
    lines = [f"📚 <b>Прогресс обучения{dept_label}</b>\n"]
    for e in employees:
        role = e["role"] if e["role"] in ("employee", "dept_head", "admin") else "employee"
        available = learn_scenarios_for_role(role)
        completed = learning_completed_versions(int(e["tg_id"]))
        done_count = sum(1 for s in available if completed.get(s["id"], 0) >= s["version"])
        total = len(available)
        icon = "✅" if done_count == total else ("🟡" if done_count > 0 else "🔴")
        pending = [s["label"] for s in available if completed.get(s["id"], 0) < s["version"]]
        line = f"{icon} <b>{e['full_name']}</b>: {done_count}/{total}"
        if pending:
            line += f"\n    не пройдено: {', '.join(pending)}"
        lines.append(line)

    await reply_long_text(update.effective_message, lines, parse_mode="HTML")

async def cb_remind_task_pick(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Руководитель выбирает задачу своего отдела (или любую, если admin) для напоминания."""
    query = update.callback_query; await query.answer()
    if not is_admin_check(query): return
    dept_filter = dept_filter_for(query)
    open_t = tasks_open()
    if dept_filter:
        open_t = [t for t in open_t if get_dept(t["assigned_to_id"]) == dept_filter]
    if not open_t:
        await query.message.reply_text("Нет активных задач для напоминания."); return
    buttons = []
    for t in open_t[:30]:
        label = f"{t['assigned_to_name'].split()[0]}: {t['title'][:30]}" + ("…" if len(t['title']) > 30 else "")
        buttons.append([InlineKeyboardButton(label, callback_data=f"remindtask_{t['task_id']}")])
    await query.message.reply_text(
        "О какой задаче напомнить?",
        reply_markup=InlineKeyboardMarkup(buttons)
    )

async def cb_remind_task_send(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Отправляет напоминание исполнителю выбранной задачи."""
    query = update.callback_query; await query.answer("📨 Напоминание отправлено")
    if not is_admin_check(query): return
    tid = query.data.replace("remindtask_", "", 1)
    task = task_by_id(tid)
    if not task:
        await query.message.reply_text("Задача не найдена."); return

    tg_id = query.from_user.id
    sender = emp_by_id(tg_id)
    # Если человек ещё проходит этот шаг обучения — напоминание уходит ему
    # самому, не реальному исполнителю задачи. Без этой защиты предыдущая
    # версия отправляла настоящее уведомление живому коллеге о реальной
    # рабочей задаче просто потому, что кто-то изучал функцию.
    in_remind_lesson = _learning_in_progress.get(tg_id) == "remind"
    recipient_id = tg_id if in_remind_lesson else int(task["assigned_to_id"])
    recipient_name = "себе (тренажёр)" if in_remind_lesson else task["assigned_to_name"]

    try:
        await ctx.bot.send_message(
            recipient_id,
            f"🔔 Напоминание от {sender['full_name']}!\n\n"
            f"<b>{task['title']}</b>\n"
            f"Срок: {fmt_dl(task['deadline'])}\n"
            f"ID: <code>{tid}</code>\n\n"
            f"/done {tid} — отметить выполненной",
            parse_mode="HTML"
        )
        await query.message.reply_text(f"✅ Напомнила {recipient_name} о задаче «{task['title']}».")
    except Exception as ex:
        logger.warning(f"remind_task_send: {ex}")
        await query.message.reply_text("❌ Не удалось отправить напоминание.")

    if in_remind_lesson:
        learn_kb = pop_learning_continue_keyboard(tg_id)
        if learn_kb:
            await query.message.reply_text("Готово, идём дальше:", reply_markup=learn_kb)

async def cb_recover_period_pick(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Выбор периода для восстановления задач из планов."""
    query = update.callback_query; await query.answer()
    if not is_strict_admin_check(query): return
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("Сегодня",      callback_data="recover_run_1")],
        [InlineKeyboardButton("3 дня",        callback_data="recover_run_3")],
        [InlineKeyboardButton("Неделя",       callback_data="recover_run_7")],
        [InlineKeyboardButton("Месяц",        callback_data="recover_run_30")],
        [InlineKeyboardButton("◀️ Назад",     callback_data="back_main")],
    ])
    await query.message.reply_text(
        "За какой период проверить планы и досоздать пропущенные задачи?",
        reply_markup=keyboard
    )

async def cb_recover_run(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Запускает восстановление задач за выбранный период."""
    query = update.callback_query; await query.answer("Сканирую планы...")
    if not is_strict_admin_check(query): return
    tg_id = query.from_user.id
    dept_filter = dept_filter_for(query)

    days_back = int(query.data.replace("recover_run_", "", 1))
    date_to = today_date().strftime("%Y-%m-%d")
    date_from = (today_date() - timedelta(days=days_back)).strftime("%Y-%m-%d")

    recovered, by_person, affected_dates = recover_tasks_from_plans(date_from, date_to, dept_filter)

    if recovered == 0:
        await query.message.reply_text("✅ Расхождений не найдено — все задачи из планов на месте.")
        return

    lines = [f"🔧 <b>Восстановлено задач: {recovered}</b>\n"]
    for person, cnt in sorted(by_person.items()):
        lines.append(f"  {person}: +{cnt}")
    lines.append(f"\nЗатронутые даты: {', '.join(affected_dates)}")
    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("📨 Запросить статусы сейчас", callback_data="checkstatuses_now")
    ]])
    await reply_long_text(query.message, lines, parse_mode="HTML", reply_markup=keyboard)

    # уведомляем остальных администраторов
    notify = (
        f"🔧 <b>Восстановление задач выполнено</b>\n"
        f"Инициатор: {emp_by_id(tg_id)['full_name']}\n"
        f"Период: {date_from} — {date_to}\n"
        f"Восстановлено задач: {recovered}\n"
    )
    for adm in emp_admins():
        if int(adm["tg_id"]) == tg_id:
            continue
        try:
            await ctx.bot.send_message(int(adm["tg_id"]), notify, parse_mode="HTML")
        except Exception as ex:
            logger.warning(f"recover notify admin {adm['tg_id']}: {ex}")

async def cb_tasks_today(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    if not is_admin_check(query): return
    dept_filter = dept_filter_for(query)
    today = today_date().strftime("%Y-%m-%d")
    today_fmt = today_date().strftime("%d.%m.%Y")
    tasks = tasks_for_date(today)
    over = tasks_overdue()
    if dept_filter:
        tasks = [t for t in tasks if get_dept(t["assigned_to_id"]) == dept_filter]
        over  = [t for t in over  if get_dept(t["assigned_to_id"]) == dept_filter]
    dept_label = f" — {dept_filter}" if dept_filter else ""
    lines = [f"📋 <b>Задачи на сегодня{dept_label} ({today_fmt}):</b>\n"]
    if not tasks and not over:
        lines.append("✅ Нет задач на сегодня!")
    SHOW_LIMIT = 30
    if over:
        lines.append(f"🔴 <b>Просроченные ({len(over)}):</b>")
        for t in over[:SHOW_LIMIT]:
            lines.append(f"  🔴 <code>{t['task_id']}</code> {t['assigned_to_name']}: {t['title']} (срок {fmt_dl(t['deadline'])})")
        if len(over) > SHOW_LIMIT:
            lines.append(f"  … и ещё {len(over)-SHOW_LIMIT}. Используй 📥 Экспорт отдела для полного списка.")
        lines.append("")
    if tasks:
        lines.append(f"🔵 <b>На сегодня ({len(tasks)}):</b>")
        for t in tasks[:SHOW_LIMIT]:
            src = " <i>(план)</i>" if t.get("source") == "plan" else ""
            sl = {"open":"⬜","in_progress":"🔄","paused":"⏸","done":"✅","cancelled":"🚫"}
            icon = sl.get(t["status"],"🔵")
            lines.append(f"  {icon} <code>{t['task_id']}</code> {t['assigned_to_name']}: {t['title']}{src}")
        if len(tasks) > SHOW_LIMIT:
            lines.append(f"  … и ещё {len(tasks)-SHOW_LIMIT}. Используй 📥 Экспорт отдела для полного списка.")
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("🔄 Обновить", callback_data="tasks_today")],
        [InlineKeyboardButton("◀️ Назад", callback_data="back_main")],
    ])
    await reply_long_text(query.message, lines, parse_mode="HTML", reply_markup=keyboard)

async def cb_show_all_tasks(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    if not is_admin_check(query): return
    dept_filter = dept_filter_for(query)
    open_t = tasks_open(); over_t = tasks_overdue()
    if dept_filter:
        open_t = [t for t in open_t if get_dept(t["assigned_to_id"]) == dept_filter]
        over_t = [t for t in over_t if get_dept(t["assigned_to_id"]) == dept_filter]
    dept_label = f" — {dept_filter}" if dept_filter else ""
    lines = [f"📋 <b>Активные задачи{dept_label} ({len(open_t)}):</b>\n"]
    if not open_t:
        lines.append("✅ Нет активных задач!")
    else:
        by_dept: dict = {}
        for t in open_t:
            dept = get_dept(t["assigned_to_id"])
            by_dept.setdefault(dept, []).append(t)
        for dept, dt_list in sorted(by_dept.items()):
            lines.append(f"<b>— {dept} —</b>")
            for t in dt_list:
                icon = "🔴" if t in over_t else "🔵"
                src = " <i>(план)</i>" if t.get("source") == "plan" else ""
                lines.append(f"  {icon} <code>{t['task_id']}</code> {t['assigned_to_name']}: {t['title']} — до {fmt_dl(t['deadline'])}{src}")
            lines.append("")
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("🔄 Обновить", callback_data="show_all_tasks")],
        [InlineKeyboardButton("◀️ Назад", callback_data="back_main")],
    ])
    await reply_long_text(query.message, lines, parse_mode="HTML", reply_markup=keyboard)

async def cb_summary_depts(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    if not is_admin_check(query): return
    dept_filter = dept_filter_for(query)
    employees = emp_employees()
    if dept_filter:
        employees = [e for e in employees if get_dept(e["tg_id"]) == dept_filter]
    plan_ids   = {str(p["tg_id"]) for p in plans_today()}
    report_ids = {str(r["tg_id"]) for r in reports_today()}
    all_t = tasks_all_real()
    today = today_str(); today_fmt = today_date().strftime("%d.%m.%Y")
    dept_employees: dict = {}
    for e in employees:
        dept = get_dept(e["tg_id"])
        dept_employees.setdefault(dept, []).append(e)
    dept_label = f" — {dept_filter}" if dept_filter else ""
    lines = [f"📊 <b>Сводка по отделам{dept_label} — {today_fmt}</b>\n"]
    for dept, emps in sorted(dept_employees.items()):
        submitted = sum(1 for e in emps if str(e["tg_id"]) in report_ids)
        planned   = sum(1 for e in emps if str(e["tg_id"]) in plan_ids)
        overdue_c = sum(1 for t in all_t
                        if get_dept(t["assigned_to_id"]) == dept
                        and t["status"] not in ("done", "cancelled") and is_overdue(t.get("deadline",""), today))
        icon = "✅" if submitted == len(emps) else ("🟡" if submitted > 0 else "🔴")
        lines.append(f"{icon} <b>{dept}</b> ({len(emps)} чел.)")
        lines.append(f"  планов {planned}/{len(emps)}  EOD закрыт {submitted}/{len(emps)}"
                     + (f"  ⚠️{overdue_c}" if overdue_c else ""))
        for e in emps:
            tid = str(e["tg_id"])
            lines.append(f"    {e['full_name']}  {'✅' if tid in plan_ids else '❌'}план  {'✅' if tid in report_ids else '—'}отчёт")
        lines.append("")
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("👤 По сотруднику", callback_data="summary_person_list")],
        [InlineKeyboardButton("◀️ Назад", callback_data="back_main")],
    ])
    await reply_long_text(query.message, lines, parse_mode="HTML", reply_markup=keyboard)

async def cb_summary_person_list(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    if not is_admin_check(query): return
    dept_filter = dept_filter_for(query)
    employees = emp_employees()
    if dept_filter:
        employees = [e for e in employees if get_dept(e["tg_id"]) == dept_filter]
    buttons = []
    row = []
    for e in employees:
        row.append(InlineKeyboardButton(e["full_name"].split()[0], callback_data=f"person_{e['tg_id']}"))
        if len(row) == 2:
            buttons.append(row); row = []
    if row: buttons.append(row)
    buttons.append([InlineKeyboardButton("◀️ Назад", callback_data="back_main")])
    await query.message.reply_text("Выбери сотрудника:", reply_markup=InlineKeyboardMarkup(buttons))

async def cb_workday_summary(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """
    Сводка по команде: кто во сколько начал и закончил рабочий день сегодня.
    Доступна admin (вся компания), dept_head (свой отдел), founder (вся
    компания, read-only — кнопка та же, просто без управляющих соседей).
    """
    query = update.callback_query; await query.answer()
    if not is_admin_check(query): return
    dept_filter = dept_filter_for(query)
    employees = emp_employees()
    if dept_filter:
        employees = [e for e in employees if get_dept(e["tg_id"]) == dept_filter]

    today_rows = {str(r["tg_id"]): r for r in workday_today()}
    today_fmt = today_date().strftime("%d.%m.%Y")
    dept_label = f" — {dept_filter}" if dept_filter else ""

    closed, in_progress, not_started = [], [], []
    for e in employees:
        tid = str(e["tg_id"])
        row = today_rows.get(tid)
        if not row or not row.get("start_at"):
            not_started.append(e)
        elif not row.get("end_at"):
            in_progress.append((e, row))
        else:
            closed.append((e, row))

    def fmt_time(ts: str) -> str:
        # ts вида "2026-06-24 11:06" -> "11:06"
        return ts.split(" ")[-1] if ts else "—"

    lines = [f"🕐 <b>Рабочий день{dept_label} — {today_fmt}</b>\n"]

    if in_progress:
        lines.append(f"🟢 <b>Сейчас на смене ({len(in_progress)}):</b>")
        for e, row in sorted(in_progress, key=lambda x: x[1].get("start_at", "")):
            lines.append(f"  {e['full_name']}: начал {fmt_time(row['start_at'])}")
        lines.append("")

    if closed:
        lines.append(f"⚪ <b>День закрыт ({len(closed)}):</b>")
        for e, row in sorted(closed, key=lambda x: x[1].get("start_at", "")):
            lines.append(f"  {e['full_name']}: {fmt_time(row['start_at'])}–{fmt_time(row['end_at'])}")
        lines.append("")

    if not_started:
        lines.append(f"🔴 <b>Не начали день ({len(not_started)}):</b>")
        for e in not_started:
            lines.append(f"  {e['full_name']}")

    if not employees:
        lines.append("Нет сотрудников для отображения.")

    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("🔄 Обновить", callback_data="workday_summary")],
        [InlineKeyboardButton("◀️ Назад", callback_data="back_main")],
    ])
    await reply_long_text(query.message, lines, parse_mode="HTML", reply_markup=keyboard)

async def cb_summary_person(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    if not is_admin_check(query): return
    dept_filter = dept_filter_for(query)
    tg_id = query.data.split("_", 1)[1]
    emp = emp_by_id(int(tg_id))
    if not emp:
        await query.message.reply_text("Не найден."); return
    if dept_filter and get_dept(tg_id) != dept_filter:
        await query.answer("⛔ Сотрудник не из твоего отдела.", show_alert=True); return
    today = today_str(); today_fmt = today_date().strftime("%d.%m.%Y")
    plan_ids   = {str(p["tg_id"]) for p in plans_today()}
    report_ids = {str(r["tg_id"]) for r in reports_today()}
    all_t = tasks_all_real()
    emp_tasks = [t for t in all_t if str(t["assigned_to_id"]) == tg_id and t["status"] not in ("done", "cancelled")]
    done_today = [t for t in all_t if str(t["assigned_to_id"]) == tg_id
                  and t["status"] == "done" and t.get("done_at","").startswith(today)]
    overdue = [t for t in emp_tasks if is_overdue(t.get("deadline",""), today)]
    sl = {"open":"⬜","in_progress":"🔄","paused":"⏸","done":"✅","cancelled":"🚫"}
    lines = [
        f"👤 <b>{emp['full_name']}</b>  ({get_dept(tg_id)})",
        f"Сегодня {today_fmt}:  план {'✅' if tg_id in plan_ids else '❌'}  EOD {'✅' if tg_id in report_ids else '❌'}\n",
    ]
    if emp_tasks:
        lines.append(f"<b>Активные ({len(emp_tasks)}):</b>")
        for t in emp_tasks:
            icon = "🔴" if t in overdue else sl.get(t["status"],"🔵")
            src = " <i>(план)</i>" if t.get("source") == "plan" else ""
            comment = f"\n    💬 {t['comment']}" if t.get("comment") else ""
            link = ""
            if t.get("result_link"):
                link = "\n" + "\n".join(f"    🔗 {l}" for l in t["result_link"].split("\n") if l)
            lines.append(f"  {icon} {t['title']} — до {fmt_dl(t['deadline'])}{src}{comment}{link}")
    if done_today:
        lines.append(f"\n<b>Выполнено сегодня ({len(done_today)}):</b>")
        for t in done_today:
            lines.append(f"  ✅ {t['title']}")
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("👥 Другой", callback_data="summary_person_list")],
        [InlineKeyboardButton("◀️ Назад", callback_data="back_main")],
    ])
    await reply_long_text(query.message, lines, parse_mode="HTML", reply_markup=keyboard)

async def cb_period_week(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    if not is_admin_check(query): return
    dept_filter = dept_filter_for(query)
    today = today_date()
    week_start = (today - timedelta(days=today.weekday())).strftime("%Y-%m-%d")
    week_end   = today.strftime("%Y-%m-%d")
    week_start_fmt = (today - timedelta(days=today.weekday())).strftime("%d.%m")
    week_end_fmt   = today.strftime("%d.%m.%Y")
    employees  = emp_employees()
    if dept_filter:
        employees = [e for e in employees if get_dept(e["tg_id"]) == dept_filter]
    reps       = reports_for_period(week_start, week_end)
    all_t      = tasks_all_real()
    done_w     = [t for t in all_t if t["status"] == "done"
                  and t.get("done_at","")[:10] >= week_start]
    over       = tasks_overdue()
    if dept_filter:
        done_w = [t for t in done_w if get_dept(t["assigned_to_id"]) == dept_filter]
        over   = [t for t in over   if get_dept(t["assigned_to_id"]) == dept_filter]
    dept_label = f" — {dept_filter}" if dept_filter else ""
    lines = [
        f"📅 <b>Неделя{dept_label} ({week_start_fmt} — {week_end_fmt})</b>\n",
        f"Уникальных отчётов: {len(set(r['tg_id'] for r in reps))} / {len(employees)}",
        f"Задач выполнено: {len(done_w)}",
        f"Просроченных: {len(over)}\n",
        "<b>По сотрудникам:</b>",
    ]
    for e in employees:
        tid = str(e["tg_id"])
        r_c = sum(1 for r in reps if str(r["tg_id"]) == tid)
        d_c = sum(1 for t in done_w if str(t["assigned_to_id"]) == tid)
        o_c = sum(1 for t in over if str(t["assigned_to_id"]) == tid)
        icon = "✅" if r_c >= 4 else ("🟡" if r_c >= 1 else "🔴")
        lines.append(f"{icon} <b>{e['full_name']}</b> ({get_dept(e['tg_id'])}): "
                     f"отчётов {r_c}, задач ✅{d_c}" + (f" ⚠️{o_c}" if o_c else ""))
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("📅 За месяц", callback_data="period_month")],
        [InlineKeyboardButton("◀️ Назад",    callback_data="back_main")],
    ])
    await reply_long_text(query.message, lines, parse_mode="HTML", reply_markup=keyboard)

async def cb_period_month(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    if not is_admin_check(query): return
    dept_filter = dept_filter_for(query)
    today = today_date()
    month_start = today.replace(day=1).strftime("%Y-%m-%d")
    month_end   = today.strftime("%Y-%m-%d")
    month_label = today.strftime("%B %Y")
    _, last = calendar.monthrange(today.year, today.month)
    wd = sum(1 for d in range(1, today.day+1) if date(today.year, today.month, d).weekday() < 5)
    employees = emp_employees()
    if dept_filter:
        employees = [e for e in employees if get_dept(e["tg_id"]) == dept_filter]
    reps      = reports_for_period(month_start, month_end)
    all_t     = tasks_all_real()
    done_m    = [t for t in all_t if t["status"] == "done"
                 and t.get("done_at","")[:10] >= month_start]
    over      = tasks_overdue()
    if dept_filter:
        done_m = [t for t in done_m if get_dept(t["assigned_to_id"]) == dept_filter]
        over   = [t for t in over   if get_dept(t["assigned_to_id"]) == dept_filter]
    dept_label = f" — {dept_filter}" if dept_filter else ""
    lines = [
        f"📆 <b>{month_label}{dept_label}</b>\n",
        f"Рабочих дней прошло: {wd}",
        f"Задач выполнено: {len(done_m)}",
        f"Просроченных: {len(over)}\n",
        "<b>По сотрудникам:</b>",
    ]
    for e in employees:
        tid = str(e["tg_id"])
        r_c = sum(1 for r in reps if str(r["tg_id"]) == tid)
        d_c = sum(1 for t in done_m if str(t["assigned_to_id"]) == tid)
        pct = round(r_c / max(wd, 1) * 100)
        icon = "✅" if pct >= 80 else ("🟡" if pct >= 40 else "🔴")
        lines.append(f"{icon} <b>{e['full_name']}</b> ({get_dept(e['tg_id'])}): "
                     f"отчётов {r_c}/{wd} ({pct}%), задач ✅{d_c}")
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("📅 За неделю", callback_data="period_week")],
        [InlineKeyboardButton("◀️ Назад",     callback_data="back_main")],
    ])
    await reply_long_text(query.message, lines, parse_mode="HTML", reply_markup=keyboard)

async def cb_closed_period_pick(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Выбор периода для просмотра закрытых (выполненных) задач."""
    query = update.callback_query; await query.answer()
    if not is_admin_check(query): return
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("Сегодня",   callback_data="closed_today")],
        [InlineKeyboardButton("Неделя",    callback_data="closed_week")],
        [InlineKeyboardButton("Месяц",     callback_data="closed_month")],
        [InlineKeyboardButton("◀️ Назад",  callback_data="back_main")],
    ])
    await query.message.reply_text("За какой период показать закрытые задачи?", reply_markup=keyboard)

async def cb_closed_tasks(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Список выполненных задач за выбранный период, с фильтром по отделу."""
    query = update.callback_query; await query.answer()
    if not is_admin_check(query): return
    dept_filter = dept_filter_for(query)
    period = query.data.replace("closed_", "", 1)  # today|week|month
    today = today_date()
    if period == "today":
        start = today.strftime("%Y-%m-%d")
        label = f"сегодня ({today.strftime('%d.%m.%Y')})"
    elif period == "week":
        start = (today - timedelta(days=today.weekday())).strftime("%Y-%m-%d")
        label = f"за неделю (с {(today - timedelta(days=today.weekday())).strftime('%d.%m')})"
    else:
        start = today.replace(day=1).strftime("%Y-%m-%d")
        label = f"за {today.strftime('%B %Y')}"

    all_t = tasks_all_real()
    done = [t for t in all_t if t["status"] == "done" and t.get("done_at","")[:10] >= start]
    if dept_filter:
        done = [t for t in done if get_dept(t["assigned_to_id"]) == dept_filter]
    dept_label = f" — {dept_filter}" if dept_filter else ""

    lines = [f"✅ <b>Закрытые задачи{dept_label}, {label}</b>\n", f"Всего: {len(done)}\n"]
    by_person: dict = {}
    for t in done:
        by_person.setdefault(t["assigned_to_name"], []).append(t)
    for person, tlist in sorted(by_person.items()):
        lines.append(f"<b>{person}</b> ({len(tlist)}):")
        for t in tlist:
            src = " <i>(план)</i>" if t.get("source") == "plan" else ""
            comment = f" — {t['comment']}" if t.get("comment") else ""
            lines.append(f"  ✅ {t['title']}{src}{comment}")
        lines.append("")
    if not done:
        lines.append("Нет закрытых задач за этот период.")

    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("Сегодня", callback_data="closed_today"),
         InlineKeyboardButton("Неделя",  callback_data="closed_week"),
         InlineKeyboardButton("Месяц",   callback_data="closed_month")],
        [InlineKeyboardButton("◀️ Назад", callback_data="back_main")],
    ])
    await reply_long_text(query.message, lines, parse_mode="HTML", reply_markup=keyboard)

async def cb_export_dept_pick(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Выбор периода для экспорта в Excel."""
    query = update.callback_query; await query.answer()
    if not is_readonly_admin_check(query): return
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("Неделя", callback_data="export_week")],
        [InlineKeyboardButton("Месяц",  callback_data="export_month")],
        [InlineKeyboardButton("◀️ Назад", callback_data="back_main")],
    ])
    await query.message.reply_text("За какой период сформировать Excel-отчёт?", reply_markup=keyboard)

def build_export_workbook(dept_filter: str, period: str):
    """Строит .xlsx с задачами за период, фильтр по отделу опционален."""
    today = today_date()
    if period == "week":
        start = (today - timedelta(days=today.weekday())).strftime("%Y-%m-%d")
        label = f"неделя с {(today - timedelta(days=today.weekday())).strftime('%d.%m.%Y')}"
    else:
        start = today.replace(day=1).strftime("%Y-%m-%d")
        label = today.strftime("%B %Y")

    all_t = tasks_all_real()
    rows = [t for t in all_t if t.get("created_at","")[:10] >= start
            or t.get("done_at","")[:10] >= start]
    if dept_filter:
        rows = [t for t in rows if get_dept(t["assigned_to_id"]) == dept_filter]

    wb = Workbook()
    ws = wb.active
    ws.title = "Задачи"

    headers = ["ID", "Сотрудник", "Отдел", "Название", "Срок", "Статус", "Комментарий", "Ссылка", "Источник"]
    ws.append(headers)
    header_fill = PatternFill(start_color="185FA5", end_color="185FA5", fill_type="solid")
    for col_i, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_i)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    status_labels = {"open":"Не начато","in_progress":"В работе","paused":"Приостановлено","done":"Выполнено","cancelled":"Отменено"}
    link_col_idx = headers.index("Ссылка") + 1 if "Ссылка" in headers else None
    for t in rows:
        ws.append([
            t["task_id"], t["assigned_to_name"], get_dept(t["assigned_to_id"]),
            t["title"], t.get("deadline",""), status_labels.get(t["status"], t["status"]),
            t.get("comment",""), t.get("result_link",""), t.get("source","manual"),
        ])
        if link_col_idx and t.get("result_link") and "\n" in t["result_link"]:
            # несколько ссылок в одной ячейке — включаем перенос строк,
            # иначе Excel сжимает их в нечитаемую сплошную строку
            ws.cell(row=ws.max_row, column=link_col_idx).alignment = Alignment(wrap_text=True, vertical="top")

    widths = [10, 20, 18, 40, 12, 14, 30, 30, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, label

async def cb_export_period(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Формирует и отправляет xlsx-файл."""
    query = update.callback_query; await query.answer()
    if not is_readonly_admin_check(query): return
    dept_filter = dept_filter_for(query)
    period = query.data.replace("export_", "", 1)  # week|month

    buf, label = build_export_workbook(dept_filter, period)
    dept_part = dept_filter.replace(" ", "_") if dept_filter else "all"
    filename = f"report_{dept_part}_{today_date().strftime('%Y%m%d')}.xlsx"

    await query.message.reply_document(
        document=InputFile(buf, filename=filename),
        caption=f"📥 Отчёт за {label}" + (f", отдел «{dept_filter}»" if dept_filter else ", все отделы")
    )

    tg_id = query.from_user.id
    if _learning_in_progress.get(tg_id) == "export":
        learn_kb = pop_learning_continue_keyboard(tg_id)
        if learn_kb:
            await query.message.reply_text("Готово, идём дальше:", reply_markup=learn_kb)

async def cb_dynamics_dept(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Динамика выполнения задач по неделям — текстовый спарклайн за последние 4 недели."""
    query = update.callback_query; await query.answer()
    if not is_readonly_admin_check(query): return
    dept_filter = dept_filter_for(query)
    all_t = tasks_all_real()
    if dept_filter:
        all_t = [t for t in all_t if get_dept(t["assigned_to_id"]) == dept_filter]

    today = today_date()
    weeks = []
    for i in range(3, -1, -1):
        w_start = today - timedelta(days=today.weekday() + i*7)
        w_end = w_start + timedelta(days=6)
        weeks.append((w_start, w_end))

    dept_label = f" — {dept_filter}" if dept_filter else ""
    lines = [f"📈 <b>Динамика выполнения{dept_label}</b>\n"]
    counts = []
    for w_start, w_end in weeks:
        s = w_start.strftime("%Y-%m-%d")
        e = w_end.strftime("%Y-%m-%d")
        done_count = sum(1 for t in all_t if t["status"] == "done"
                         and s <= t.get("done_at","")[:10] <= e)
        counts.append(done_count)
        bar = "█" * min(done_count, 20) + ("" if done_count <= 20 else f" (+{done_count-20})")
        lines.append(f"{w_start.strftime('%d.%m')}–{w_end.strftime('%d.%m')}: {bar} {done_count}")

    if len(counts) >= 2 and counts[-2] > 0:
        delta = counts[-1] - counts[-2]
        trend = "📈 рост" if delta > 0 else ("📉 спад" if delta < 0 else "➡️ без изменений")
        lines.append(f"\n{trend} к прошлой неделе ({'+' if delta>=0 else ''}{delta})")

    keyboard = InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Назад", callback_data="back_main")]])
    await reply_long_text(query.message, lines, parse_mode="HTML", reply_markup=keyboard)

# ── NOTIFICATIONS ─────────────────────────────────────────────────────────────
# Счётчик отправленных напоминаний о старте дня за сегодня, чтобы не слать
# больше 2 повторов с интервалом 30 минут (module-level, сбрасывается сам
# по себе раз в день, так как ключ — today_str(), старые даты просто не
# совпадают и накопления не происходит при разумном перезапуске).
_startday_reminder_count = {}  # f"{tg_id}:{date}" -> count

async def job_check_workday_start(bot: Bot):
    """
    Каждые 30 минут проверяет, у кого наступило время старта рабочего дня
    по их графику (10:00 стандартный, 11:00 сдвинутый) и старт не отмечен.
    Шлёт вопрос с кнопкой 'Начать рабочий день'. Эскалация: максимум 2
    напоминания с интервалом 30 минут, дальше не дёргает — взрослые люди,
    дальнейшие настойчивые пинги контрпродуктивны.
    """
    now = datetime.now(APZ)
    today = today_str()
    for e in emp_employees():
        tg_id = int(e["tg_id"])
        if workday_started(tg_id):
            continue
        shift = emp_shift(tg_id)
        start_dt = now.replace(hour=shift["start_hour"], minute=shift["start_minute"], second=0, microsecond=0)
        if now < start_dt:
            continue  # рабочий день для этого человека ещё не наступил

        key = f"{tg_id}:{today}"
        count = _startday_reminder_count.get(key, 0)
        if count >= 2:
            continue  # уже напомнили дважды, дальше не дёргаем

        try:
            keyboard = InlineKeyboardMarkup([[
                InlineKeyboardButton("☀️ Начать рабочий день", callback_data="startday_btn"),
            ]])
            text = ("☀️ Доброе утро! Отметь начало рабочего дня:" if count == 0
                    else "🔔 Напоминание: ты ещё не отметил(а) начало рабочего дня.")
            await bot.send_message(tg_id, text, reply_markup=keyboard)
            _startday_reminder_count[key] = count + 1
        except Exception as ex:
            logger.warning(f"check_workday_start {tg_id}: {ex}")

_endday_reminder_count = {}  # f"{tg_id}:{date}" -> count, аналогично старту дня

async def job_check_workday_end(bot: Bot):
    """
    Каждые 30 минут проверяет, у кого наступило время конца рабочего дня по
    их графику (19:00 стандартный, 22:00 сдвинутый) и день ещё не закрыт.
    Запускает EOD-опрос автоматически — это заменяет старый единый
    job_request_eod в 19:00 для всех. Эскалация: максимум 2 запуска с
    интервалом 30 минут, если сотрудник не реагирует на первый.
    """
    now = datetime.now(APZ)
    today = today_str()
    for e in emp_employees():
        tg_id = int(e["tg_id"])
        if workday_ended(tg_id):
            continue
        shift = emp_shift(tg_id)
        end_dt = now.replace(hour=shift["end_hour"], minute=shift["end_minute"], second=0, microsecond=0)
        if now < end_dt:
            continue

        key = f"{tg_id}:{today}"
        count = _endday_reminder_count.get(key, 0)
        if count >= 2:
            continue

        try:
            # Та же защита от утечки состояния обучения, что в cmd_eod/cmd_endday.
            _learning_in_progress.pop(tg_id, None)
            await start_eod_flow(bot, tg_id)
            _endday_reminder_count[key] = count + 1
        except Exception as ex:
            logger.warning(f"check_workday_end {tg_id}: {ex}")

async def job_learning_reminder(bot: Bot):
    """
    14:00 МСК каждый день — напоминает сотрудникам, которые начали обучение,
    но не завершили его (started_at есть, completed_at пусто). После 5 рабочих
    дней с начала — сбрасывает прогресс и удаляет тренажёрные задачи, чтобы
    не копился незавершённый мусор в листах tasks и learning_progress.
    """
    today = today_date()
    for r in learning_abandoned_employees():
        tg_id = int(r["tg_id"])
        working_days = count_working_days_between(r.get("started_at", ""), today)

        if working_days >= 5:
            delete_training_tasks(tg_id)
            # сбрасываем прогресс: очищаем все колонки сценариев и started_at,
            # чтобы при следующем обращении обучение началось заново с шага 1
            ws = learning_sheet()
            columns = learning_columns()
            records = safe_records(ws, columns)
            for i, rec in enumerate(records, start=2):
                if str(rec["tg_id"]) == str(tg_id):
                    blank = [tg_id, rec["full_name"], rec["role"], rec["department"], "", "", ""]
                    blank += [""] * len(LEARN_SCENARIOS)
                    ws.update(f"A{i}", [blank])
                    invalidate_cache("learning_progress")
                    break
            try:
                await bot.send_message(
                    tg_id,
                    "⏳ Обучение не было завершено в течение 5 рабочих дней.\n"
                    "Прогресс сброшен, тренажёрные задачи удалены.\n\n"
                    "Напиши /start, чтобы начать заново."
                )
            except Exception as ex:
                logger.warning(f"learning_reminder reset {tg_id}: {ex}")
            continue

        try:
            await bot.send_message(
                tg_id,
                "📚 Напоминание: ты начал(а) обучение, но не закончил(а).\n"
                f"Если не завершить за {5 - working_days} рабочих дней — "
                "прогресс и тестовые задачи будут сброшены.\n\n"
                "Напиши /start или нажми «📚 Обучение» в /menu, чтобы продолжить."
            )
            learning_set_field(tg_id, "last_reminder_at", datetime.now(TZ).strftime("%Y-%m-%d %H:%M"))
        except Exception as ex:
            logger.warning(f"learning_reminder {tg_id}: {ex}")

async def job_ping_deadlines(bot: Bot):
    for t in tasks_due_tomorrow():
        try:
            await bot.send_message(int(t["assigned_to_id"]),
                f"⏰ Завтра срок задачи!\n<b>{t['title']}</b>\n"
                f"ID: <code>{t['task_id']}</code>\n/done {t['task_id']}",
                parse_mode="HTML")
        except Exception as ex: logger.warning(f"ping_tomorrow: {ex}")
    for t in tasks_due_today_list():
        try:
            await bot.send_message(int(t["assigned_to_id"]),
                f"🚨 Срок задачи сегодня!\n<b>{t['title']}</b>\n"
                f"ID: <code>{t['task_id']}</code>",
                parse_mode="HTML")
        except Exception as ex: logger.warning(f"ping_today: {ex}")

    # уведомление руководителю отдела о новых просрочках его сотрудников
    overdue = tasks_overdue()
    if not overdue:
        return
    by_dept: dict = {}
    for t in overdue:
        by_dept.setdefault(get_dept(t["assigned_to_id"]), []).append(t)

    for head in emp_dept_heads():
        head_dept = get_dept(head["tg_id"])
        dept_overdue = by_dept.get(head_dept, [])
        if not dept_overdue:
            continue
        lines = [f"⚠️ <b>Просроченные задачи в отделе «{head_dept}»</b>\n"]
        for t in dept_overdue:
            lines.append(f"🔴 {t['assigned_to_name']}: {t['title']} (срок {fmt_dl(t['deadline'])})")
        try:
            await send_long_text(bot, int(head["tg_id"]), lines, parse_mode="HTML")
        except Exception as ex:
            logger.warning(f"overdue notify dept_head {head['tg_id']}: {ex}")

async def job_daily_digest(bot: Bot):
    today = datetime.now(TZ).strftime("%d.%m.%Y")
    employees = emp_employees()
    p_list = plans_today(); r_list = reports_today()
    plan_ids   = {str(p["tg_id"]) for p in p_list}
    report_ids = {str(r["tg_id"]) for r in r_list}
    all_t = tasks_all_real()
    over = tasks_overdue()
    done_today = [t for t in all_t if t["status"] == "done"
                  and t.get("done_at","").startswith(datetime.now(TZ).strftime("%Y-%m-%d"))]

    # группируем сотрудников по отделам
    dept_employees: dict = {}
    for e in employees:
        dept_employees.setdefault(get_dept(e["tg_id"]), []).append(e)

    # ── Полная сводка в общий чат руководителей ──
    lines = [
        f"📊 <b>Сводка команды — {today}</b>",
        f"Планов: {len(plan_ids)}/{len(employees)}   EOD закрыт: {len(report_ids)}/{len(employees)}   "
        f"Закрыто задач: {len(done_today)}\n",
    ]
    for dept, emps in sorted(dept_employees.items()):
        submitted = sum(1 for e in emps if str(e["tg_id"]) in report_ids)
        icon = "✅" if submitted == len(emps) else ("🟡" if submitted > 0 else "🔴")
        dept_done = sum(1 for t in done_today if get_dept(t["assigned_to_id"]) == dept)
        dept_over = sum(1 for t in over if get_dept(t["assigned_to_id"]) == dept)
        lines.append(f"{icon} <b>{dept}</b>: отчётов {submitted}/{len(emps)}, закрыто {dept_done}"
                     + (f", ⚠️ просрочено {dept_over}" if dept_over else ""))
    if over:
        lines.append(f"\n⚠️ <b>Всего просроченных: {len(over)}</b>")
        for t in over[:3]:
            lines.append(f"  🔴 {t['assigned_to_name']}: {t['title']}")
        if len(over) > 3: lines.append(f"  ...и ещё {len(over)-3}")
    try:
        await send_long_text(bot, ADMIN_CHAT_ID, lines,
                               parse_mode="HTML", reply_markup=build_admin_keyboard())
    except Exception as ex:
        logger.warning(f"daily_digest: {ex}")
        # дублируем ошибку лично каждому admin — иначе провал отправки в
        # группу остаётся видимым только в логах Railway, которые никто
        # не читает в реальном времени
        for adm in emp_admins():
            try:
                await bot.send_message(
                    int(adm["tg_id"]),
                    f"⚠️ Не удалось отправить сводку в группу руководителей.\n"
                    f"Причина: {ex}\n\nПроверь /testdigest для диагностики."
                )
            except Exception:
                pass

    # ── Персональная сводка по отделу для каждого dept_head ──
    for head in emp_dept_heads():
        head_dept = get_dept(head["tg_id"])
        dept_emps = dept_employees.get(head_dept, [])
        if not dept_emps:
            continue
        no_plan   = [e["full_name"] for e in dept_emps if str(e["tg_id"]) not in plan_ids]
        no_report = [e["full_name"] for e in dept_emps if str(e["tg_id"]) not in report_ids]
        dept_done = [t for t in done_today if get_dept(t["assigned_to_id"]) == head_dept]
        dept_over = [t for t in over if get_dept(t["assigned_to_id"]) == head_dept]
        h_lines = [
            f"📊 <b>Сводка отдела «{head_dept}» — {today}</b>\n",
            f"Закрыто задач: {len(dept_done)}",
        ]
        if no_plan:   h_lines.append(f"❌ <b>Нет плана:</b> {', '.join(no_plan)}")
        if no_report: h_lines.append(f"❌ <b>День не закрыт (EOD):</b> {', '.join(no_report)}")
        if dept_over:
            h_lines.append(f"\n⚠️ <b>Просрочено в отделе: {len(dept_over)}</b>")
            for t in dept_over[:5]:
                h_lines.append(f"  🔴 {t['assigned_to_name']}: {t['title']}")
        try:
            await bot.send_message(int(head["tg_id"]), "\n".join(h_lines),
                                   parse_mode="HTML", reply_markup=build_dept_head_keyboard())
        except Exception as ex:
            logger.warning(f"daily_digest dept_head {head['tg_id']}: {ex}")

    # ── Ежедневная сводка учредителям, выбравшим digest/both ──
    founders = [r for r in emp_all() if r["role"] == "founder"
                and r.get("founder_digest_pref") in ("digest", "both")]
    for f in founders:
        try:
            await send_long_text(bot, int(f["tg_id"]), lines,
                                 parse_mode="HTML", reply_markup=build_founder_keyboard())
        except Exception as ex:
            logger.warning(f"daily_digest founder {f['tg_id']}: {ex}")

async def job_weekly_audit(bot: Bot):
    now = datetime.now(TZ)
    ws = now.strftime("%Y-W%W")
    w0 = (now - timedelta(days=now.weekday())).strftime("%d.%m")
    w1 = now.strftime("%d.%m.%Y")
    employees = emp_employees()
    plans_w   = records_for_week("plans", PH)
    reports_w = records_for_week("reports", RH)
    all_t     = tasks_all_real()
    done_w    = [t for t in all_t if t["status"]=="done" and _is_this_week(t.get("done_at",""), ws)]
    over      = tasks_overdue()
    lines = [f"📅 <b>Еженедельный аудит ({w0} — {w1})</b>\n",
             f"EOD закрыт: {len(set(r['tg_id'] for r in reports_w))}/{len(employees)}",
             f"Задач выполнено: {len(done_w)}", f"Просроченных: {len(over)}\n",
             "<b>По сотрудникам:</b>"]
    for e in employees:
        tid = str(e["tg_id"])
        r_c = sum(1 for r in reports_w if str(r["tg_id"]) == tid)
        d_c = sum(1 for t in done_w if str(t["assigned_to_id"]) == tid)
        o_c = sum(1 for t in over if str(t["assigned_to_id"]) == tid)
        icon = "✅" if r_c >= 4 else ("🟡" if r_c >= 1 else "🔴")
        lines.append(f"{icon} <b>{e['full_name']}</b> ({get_dept(e['tg_id'])}): "
                     f"отчётов {r_c}, задач ✅{d_c}" + (f" ⚠️{o_c}" if o_c else ""))
    try:
        await send_long_text(bot, ADMIN_CHAT_ID, lines, parse_mode="HTML")
    except Exception as ex: logger.warning(f"weekly_audit: {ex}")

async def job_monthly_audit(bot: Bot):
    now = datetime.now(TZ)
    month = now.strftime("%Y-%m"); month_label = now.strftime("%B %Y")
    _, last = calendar.monthrange(now.year, now.month)
    wd = sum(1 for d in range(1, now.day+1) if datetime(now.year, now.month, d).weekday() < 5)
    employees = emp_employees()
    reps      = records_for_month("reports", RH)
    all_t     = tasks_all_real()
    done_m    = [t for t in all_t if t["status"]=="done" and t.get("done_at","").startswith(month)]
    over      = tasks_overdue()
    lines = [f"📆 <b>Ежемесячный аудит — {month_label}</b>\n",
             f"Рабочих дней: {wd}", f"Задач выполнено: {len(done_m)}", f"Просроченных: {len(over)}\n",
             "<b>По сотрудникам:</b>"]
    for e in employees:
        tid = str(e["tg_id"])
        r_c = sum(1 for r in reps if str(r["tg_id"]) == tid)
        d_c = sum(1 for t in done_m if str(t["assigned_to_id"]) == tid)
        pct = round(r_c / max(wd, 1) * 100)
        icon = "✅" if pct >= 80 else ("🟡" if pct >= 40 else "🔴")
        lines.append(f"{icon} <b>{e['full_name']}</b> ({get_dept(e['tg_id'])}): "
                     f"отчётов {r_c}/{wd} ({pct}%), задач ✅{d_c}")
    try:
        await send_long_text(bot, ADMIN_CHAT_ID, lines, parse_mode="HTML")
    except Exception as ex: logger.warning(f"monthly_audit: {ex}")

def _is_this_week(dt_str, iso_week):
    try:
        return datetime.strptime(dt_str[:10], "%Y-%m-%d").strftime("%Y-W%W") == iso_week
    except Exception: return False

# ── SCHEDULER ─────────────────────────────────────────────────────────────────
def build_scheduler(bot: Bot):
    s = AsyncIOScheduler(timezone=APZ)
    def j(fn, **kw):
        # misfire_grace_time=900 (15 минут): если контейнер перезапускался (redeploy)
        # ровно в момент срабатывания задания, APScheduler всё равно выполнит его
        # в течение 15 минут после восстановления, а не молча пропустит до следующего дня.
        s.add_job(fn, CronTrigger(timezone=APZ, **kw), args=[bot], misfire_grace_time=900)
    # Старт/конец рабочего дня теперь проверяются каждые 30 минут в рабочем
    # окне 10:00–22:30, чтобы покрыть и стандартный график (10:00–19:00),
    # и сдвинутый (11:00–22:00) — раньше было одно фиксированное время для всех.
    j(job_check_workday_start, day_of_week="mon-fri", hour="10-21", minute="0,30")
    j(job_check_workday_end,   day_of_week="mon-fri", hour="19-22", minute="0,30")
    j(job_ping_deadlines,  day_of_week="mon-fri", hour=10, minute=0)
    j(job_learning_reminder, day_of_week="mon-fri", hour=14, minute=0)
    j(job_daily_digest,    day_of_week="mon-fri", hour=19, minute=35)
    j(job_weekly_audit,    day_of_week="fri",     hour=18, minute=0)
    j(job_monthly_audit,   day=28,                hour=17, minute=0)
    return s

# ── MAIN ──────────────────────────────────────────────────────────────────────
async def post_init(app):
    scheduler = build_scheduler(app.bot)
    scheduler.start()
    logger.info("Scheduler started")
    await app.bot.set_my_commands([
        ("start",     "Регистрация / меню"),
        ("ping",      "Проверка что бот на связи"),
        ("help",      "Список команд и обучение"),
        ("startday",  "Начать рабочий день — запросит план"),
        ("plan",      "План на день — поддерживает [ДД.ММ.ГГГГ] для другого срока"),
        ("endday",    "Актуализировать статусы и закрыть рабочий день"),
        ("eod",       "Закрыть день — статусы задач"),
        ("task",      "Поставить задачу — выбор отдела и сотрудника кнопками"),
        ("tag",       "Тег канала для задачи: /tag — выбор кнопками"),
        ("mytasks",   "Мои активные задачи"),
        ("done",      "Выполнено: /done ID"),
        ("changestatus", "Сменить статус задачи — выбор кнопками"),
        ("status",    "Статус: /status ID"),
        ("edit",      "Изменить задачу — выбор кнопками"),
        ("menu",      "Панель руководителя с кнопками"),
        ("team",      "Сводка команды"),
        ("workday",   "Кто во сколько начал/закончил рабочий день"),
        ("tasks_all", "Все задачи"),
        ("checkstatuses", "Запросить статусы у команды/отдела прямо сейчас"),
        ("learningreport", "Прогресс обучения команды"),
        ("recovertasks", "Восстановить пропавшие задачи из планов: /recovertasks [дней]"),
        ("deleteuser", "Удалить сотрудника из команды"),
        ("makeadmin", "Стать администратором"),
        ("setdepthead", "Назначить руководителя отдела: /setdepthead @username"),
        ("setfounder", "Назначить учредителя: /setfounder @username"),
        ("setdept", "Сменить отдел сотрудника: /setdept @username"),
        ("setshift", "Назначить график работы сотруднику"),
        ("fixname", "Исправить имя: /fixname @username Имя Фамилия"),
        ("fixallnames", "Нормализовать имена всех сотрудников"),
    ])

async def global_error_handler(update: object, ctx: ContextTypes.DEFAULT_TYPE):
    """Логирует все необработанные исключения, не даёт им тихо остановить polling."""
    logger.error(f"Unhandled exception: {ctx.error}", exc_info=ctx.error)
    try:
        if isinstance(update, Update) and update.effective_message:
            await update.effective_message.reply_text(
                "⚠️ Произошла ошибка при обработке запроса. Попробуй ещё раз или напиши /menu."
            )
        elif isinstance(update, Update) and update.callback_query:
            await update.callback_query.answer("⚠️ Ошибка, попробуй снова.", show_alert=True)
    except Exception:
        pass

def main():
    app = Application.builder().token(BOT_TOKEN).post_init(post_init).build()
    app.add_error_handler(global_error_handler)
    cancel = CommandHandler("cancel", lambda u,c: ConversationHandler.END)

    app.add_handler(ConversationHandler(
        entry_points=[CommandHandler("start", cmd_start)],
        states={S_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, recv_name)]},
        fallbacks=[cancel],
    ))
    app.add_handler(ConversationHandler(
        entry_points=[
            CommandHandler("plan", cmd_plan),
            CallbackQueryHandler(cmd_plan, pattern="^learntry_plan$"),
            CommandHandler("startday", cmd_startday),
            CallbackQueryHandler(cmd_startday, pattern="^startday_btn$"),
            CallbackQueryHandler(cmd_startday, pattern="^learntry_startday$"),
        ],
        states={S_PLAN: [MessageHandler(filters.TEXT & ~filters.COMMAND, recv_plan)]},
        fallbacks=[cancel],
        per_message=False,
    ))
    app.add_handler(ConversationHandler(
        entry_points=[
            CommandHandler("task", cmd_task),
            CallbackQueryHandler(cmd_task, pattern="^learntry_task$"),
        ],
        states={S_TASK: [MessageHandler(filters.TEXT & ~filters.COMMAND, recv_task)]},
        fallbacks=[cancel],
        per_message=False,
    ))

    # EOD text handlers (не ConversationHandler — работают через bot_data флаги)
    app.add_handler(MessageHandler(
        filters.TEXT & ~filters.COMMAND & filters.ChatType.PRIVATE,
        eod_text_router
    ), group=1)

    app.add_handler(CommandHandler("eod",       cmd_eod))
    app.add_handler(CommandHandler("endday",    cmd_endday))
    app.add_handler(CommandHandler("done",      cmd_done))
    app.add_handler(CommandHandler("changestatus", cmd_changestatus))
    app.add_handler(CommandHandler("status",    cmd_status))
    app.add_handler(CommandHandler("mytasks",   cmd_mytasks))
    app.add_handler(CommandHandler("menu",      cmd_menu))
    app.add_handler(CommandHandler("makeadmin", cmd_makeadmin))
    app.add_handler(CommandHandler("setadmin",  cmd_setadmin))
    app.add_handler(CommandHandler("setdepthead", cmd_setdepthead))
    app.add_handler(CommandHandler("setfounder", cmd_setfounder))
    app.add_handler(CommandHandler("setdept", cmd_setdept))
    app.add_handler(CommandHandler("setshift", cmd_setshift))
    app.add_handler(CommandHandler("deleteuser", cmd_deleteuser))
    app.add_handler(CommandHandler("fixsheets", cmd_fixsheets))
    app.add_handler(CommandHandler("fixname", cmd_fixname))
    app.add_handler(CommandHandler("fixallnames", cmd_fixallnames))
    app.add_handler(CommandHandler("edit", cmd_edit))
    app.add_handler(CommandHandler("tag", cmd_tag))
    app.add_handler(CommandHandler("ping", cmd_ping))
    app.add_handler(CommandHandler("testdigest", cmd_testdigest))
    app.add_handler(CommandHandler("help", cmd_help))
    app.add_handler(CommandHandler("learningreport", cmd_learningreport))
    app.add_handler(CommandHandler("checkstatuses", cmd_checkstatuses))
    app.add_handler(CommandHandler("recovertasks", cmd_recovertasks))
    app.add_handler(CommandHandler("team",      cmd_team))
    app.add_handler(CommandHandler("workday",   cmd_workday))
    app.add_handler(CommandHandler("tasks_all", cmd_tasks_all))

    # Callbacks
    app.add_handler(CallbackQueryHandler(cb_learn_main,          pattern="^learn_main$"))
    app.add_handler(CallbackQueryHandler(cb_notify_learning_pick, pattern="^notify_learning_pick$"))
    app.add_handler(CallbackQueryHandler(cb_notify_learning_send, pattern="^notifylearn_"))
    app.add_handler(CallbackQueryHandler(cb_learn_try,           pattern="^learntry_"))
    app.add_handler(CallbackQueryHandler(cb_learn_scenario,      pattern="^learn_"))
    app.add_handler(CallbackQueryHandler(cb_confirm_plan,        pattern="^confirm_plan_"))
    app.add_handler(CallbackQueryHandler(cb_tasknew_dept,        pattern="^tasknew_dept_"))
    app.add_handler(CallbackQueryHandler(cb_tasknew_emp,         pattern="^tasknew_emp_"))
    app.add_handler(CallbackQueryHandler(cb_tagtask_pick,        pattern="^tagtask_"))
    app.add_handler(CallbackQueryHandler(cb_tagchannel_pick,     pattern="^tagchannel_"))
    app.add_handler(CallbackQueryHandler(cb_sethead_dept,        pattern="^sethead_dept_"))
    app.add_handler(CallbackQueryHandler(cb_sethead_emp,         pattern="^sethead_emp_"))
    app.add_handler(CallbackQueryHandler(cb_founder_pref,        pattern="^founder_pref_"))
    app.add_handler(CallbackQueryHandler(cb_reg_dept,            pattern="^reg_dept_"))
    app.add_handler(CallbackQueryHandler(cb_admset_head,         pattern="^admset_head_"))
    app.add_handler(CallbackQueryHandler(cb_admset_dept,         pattern="^admset_dept_"))
    app.add_handler(CallbackQueryHandler(cb_setdeptemp_pick,     pattern="^setdeptemp_"))
    app.add_handler(CallbackQueryHandler(cb_setshiftemp_pick,    pattern="^setshiftemp_"))
    app.add_handler(CallbackQueryHandler(cb_setshiftval_pick,    pattern="^setshiftval_"))
    app.add_handler(CallbackQueryHandler(cmd_deleteuser,          pattern="^deleteuser_init$"))
    app.add_handler(CallbackQueryHandler(cb_deluser_pick,        pattern="^deluser_pick_"))
    app.add_handler(CallbackQueryHandler(cb_deluser_confirm,     pattern="^deluser_confirm_"))
    app.add_handler(CallbackQueryHandler(cb_donetask_pick,       pattern="^donetask_"))
    app.add_handler(CallbackQueryHandler(cb_changestatus_task_pick, pattern="^chstatustask_"))
    app.add_handler(CallbackQueryHandler(cb_changestatus_apply,     pattern="^chstatus_"))
    app.add_handler(CallbackQueryHandler(cb_edittask_pick,       pattern="^edittask_"))
    app.add_handler(CallbackQueryHandler(cb_editfield_pick,      pattern="^editfield_"))
    app.add_handler(CallbackQueryHandler(cb_eod_status,          pattern="^eods_"))
    app.add_handler(CallbackQueryHandler(cb_eod_channel,         pattern="^eodchannel_"))
    app.add_handler(CallbackQueryHandler(cb_eod_extra_yes,       pattern="^eod_extra_yes$"))
    app.add_handler(CallbackQueryHandler(cb_eod_extra_no,        pattern="^eod_extra_no$"))
    app.add_handler(CallbackQueryHandler(cb_noop,                 pattern="^noop$"))
    app.add_handler(CallbackQueryHandler(cb_back_main,           pattern="^back_main$"))
    app.add_handler(CallbackQueryHandler(cb_tasks_today,         pattern="^tasks_today$"))
    app.add_handler(CallbackQueryHandler(cb_show_all_tasks,      pattern="^show_all_tasks$"))
    app.add_handler(CallbackQueryHandler(cb_summary_depts,       pattern="^summary_depts$"))
    app.add_handler(CallbackQueryHandler(cb_summary_person_list, pattern="^summary_person_list$"))
    app.add_handler(CallbackQueryHandler(cb_workday_summary,     pattern="^workday_summary$"))
    app.add_handler(CallbackQueryHandler(cb_summary_person,      pattern="^person_"))
    app.add_handler(CallbackQueryHandler(cb_period_week,         pattern="^period_week$"))
    app.add_handler(CallbackQueryHandler(cb_period_month,        pattern="^period_month$"))
    app.add_handler(CallbackQueryHandler(cb_closed_period_pick,  pattern="^closed_period_pick$"))
    app.add_handler(CallbackQueryHandler(cb_closed_tasks,        pattern="^closed_(today|week|month)$"))
    app.add_handler(CallbackQueryHandler(cb_export_dept_pick,    pattern="^export_dept_pick$"))
    app.add_handler(CallbackQueryHandler(cb_export_period,       pattern="^export_(week|month)$"))
    app.add_handler(CallbackQueryHandler(cb_dynamics_dept,       pattern="^dynamics_dept$"))
    app.add_handler(CallbackQueryHandler(cb_overdue_pick,         pattern="^overdue_pick$"))
    app.add_handler(CallbackQueryHandler(cb_overdue_done,         pattern="^overduedone_"))
    app.add_handler(CallbackQueryHandler(cb_overdue_move_pick,    pattern="^overduemove_"))
    app.add_handler(CallbackQueryHandler(cb_overdue_set_deadline, pattern="^overduesetdl_"))
    app.add_handler(CallbackQueryHandler(cb_overdue_notify_team,  pattern="^overdue_notify_team$"))
    app.add_handler(CallbackQueryHandler(cb_remind_task_pick,    pattern="^remind_task_pick$"))
    app.add_handler(CallbackQueryHandler(cb_remind_task_send,    pattern="^remindtask_"))
    app.add_handler(CallbackQueryHandler(cb_recover_period_pick, pattern="^recover_period_pick$"))
    app.add_handler(CallbackQueryHandler(cb_recover_run,         pattern=r"^recover_run_\d+$"))
    app.add_handler(CallbackQueryHandler(cb_checkstatuses_now,   pattern="^checkstatuses_now$"))

    logger.info("Bot starting...")
    app.run_polling(drop_pending_updates=True)

# ── EOD TEXT ROUTER ───────────────────────────────────────────────────────────
async def eod_text_router(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Роутер для текстовых ответов в EOD-потоке, в /task и в /edit."""
    tg_id = update.effective_user.id
    # Обязательный комментарий к задаче после выбора канала
    if tg_id in ctx.bot_data.get("eod_pending_comment", {}):
        await recv_eod_comment(update, ctx)
        return
    # Доп. задачи
    if tg_id in ctx.bot_data.get("eod_extra_pending", set()):
        await recv_eod_extra(update, ctx)
        return
    # Новая задача после выбора сотрудника кнопками в /task
    if tg_id in ctx.bot_data.get("tasknew_pending", {}):
        await recv_tasknew_text(update, ctx)
        return
    # Новое значение поля после выбора кнопками в /edit
    if tg_id in ctx.bot_data.get("edit_pending", {}):
        await recv_edit_value(update, ctx)
        return

if __name__ == "__main__":
    main()
